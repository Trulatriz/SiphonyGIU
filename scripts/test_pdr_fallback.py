import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from modules.pdr_module import PDRModule
import os
from openpyxl import Workbook
import tempfile

# Create a minimal registros workbook
tmpdir = tempfile.gettempdir()
registros_path = os.path.join(tmpdir, 'test_registros.xlsx')
wb = Workbook()
ws = wb.active
ws.title = 'Registros'
headers = ["Filename", "Pi (MPa)", "Pf (MPa)", "PDR (MPa/s)", "Chart"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)
wb.save(registros_path)

# Create a minimal output processed workbook name (it doesn't need to exist for fallback)
processed_path = os.path.join(tmpdir, 'HDPE 20250911 procesado.xlsx')

# Instantiate module with a dummy Tk root (we won't show UI)
import tkinter as tk
root = tk.Tk()
root.withdraw()
module = PDRModule(root)
module.registros_file = registros_path
module.current_foam_type = 'HDPE'

# Call fallback update method
success = module._fallback_update_registros_file('HDPE 20250911.csv', processed_path)
print('Fallback success:', success)
print('Registros file:', registros_path)

# Open registro to show last row
from openpyxl import load_workbook
wb2 = load_workbook(registros_path)
ws2 = wb2['Registros']
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=True):
    print(row)
