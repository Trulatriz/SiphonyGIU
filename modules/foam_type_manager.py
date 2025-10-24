import json
import os
from datetime import datetime
from typing import Dict, List, Optional

class FoamTypeManager:
    def __init__(self, config_file: str = "foam_types_config.json"):
        self.config_file = config_file
        self.config = self.load_config()
        # Special key used to store module paths at paper scope (no foam split)
        self._GLOBAL_KEY = "__ALL__"
    
    def load_config(self) -> Dict:
        """Load configuration from JSON file"""
        default_config = {
            # Papers/workspaces (top-level context)
            "papers": ["Paper 1"],
            "current_paper": "Paper 1",
            # Foam types (secondary context)
            "foam_types": ["Foam_Type_1", "Foam_Type_2", "Foam_Type_3"],
            "current_foam_type": "Foam_Type_1",
            # Association of foam types per paper
            "paper_foam_types": {
                "Paper 1": ["Foam_Type_1", "Foam_Type_2", "Foam_Type_3"]
            },
            # Absolute paper root paths selected by the user (per paper)
            "paper_root_paths": {},
            # Module scope registry: 'paper' or 'paper+foam'
            "module_scopes": {
                "Combine": "paper",
                "Analysis": "paper",
                "PDR": "paper+foam",
                "OC": "paper+foam",
                "SEM": "paper+foam",
                "DSC": "paper+foam",
            },
            # Module paths organized by module -> paper -> foam_type -> paths
            "module_paths": {},
            # Per-paper All_Results.xlsx path (optional)
            "all_results_paths": {}
        }
        
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                # Merge with defaults to ensure all keys exist
                for key in default_config:
                    if key not in config:
                        config[key] = default_config[key]

                # Backward-compatibility: if module_paths lacks paper level, wrap it into current_paper
                # Expected new shape: module_paths[module][paper][foam_type] -> dict
                try:
                    for module, by_key in list(config.get("module_paths", {}).items()):
                        # If values look like per-foam dicts (old shape) rather than per-paper dicts
                        # we detect by checking if any value has known path keys (e.g., 'input_folder')
                        # directly under foam type. Then we wrap under current_paper.
                        needs_wrap = False
                        for k, v in by_key.items():
                            if isinstance(v, dict) and any(s in v for s in ("input_folder", "output_folder", "last_used", "results_file", "registros_file", "density_file")):
                                needs_wrap = True
                                break
                        if needs_wrap:
                            paper = config.get("current_paper", default_config["current_paper"])
                            wrapped = {paper: by_key}
                            config["module_paths"][module] = wrapped
                except Exception as _:
                    pass

                # Backward-compatibility: ensure module_scopes exist and valid
                try:
                    scopes = config.get("module_scopes") or {}
                    defaults = default_config["module_scopes"]
                    for k, v in defaults.items():
                        scopes.setdefault(k, v)
                    # sanitize values
                    for k in list(scopes.keys()):
                        if scopes[k] not in ("paper", "paper+foam"):
                            scopes[k] = defaults.get(k, "paper+foam")
                    config["module_scopes"] = scopes
                except Exception as _:
                    pass
                return config
            except Exception as e:
                print(f"Error loading config: {e}")
                return default_config
        else:
            return default_config
    
    def save_config(self):
        """Save configuration to JSON file"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving config: {e}")
    
    def get_foam_types(self) -> List[str]:
        """Get list of all foam types"""
        return self.config["foam_types"]
    
    def get_foam_types_for_paper(self, paper: Optional[str] = None) -> List[str]:
        """Get foam types associated with a specific paper"""
        if paper is None:
            paper = self.get_current_paper()
        return self.config.get("paper_foam_types", {}).get(paper, self.config["foam_types"])
    
    def get_current_foam_type(self) -> str:
        """Get currently selected foam type"""
        return self.config["current_foam_type"]
    
    def set_current_foam_type(self, foam_type: str):
        """Set current foam type"""
        if foam_type not in self.config["foam_types"]:
            self.add_foam_type(foam_type)
        self.config["current_foam_type"] = foam_type
        self.save_config()

    def set_foam_types_for_paper(self, paper: str, foam_types: List[str]):
        """Set foam types for a specific paper"""
        if "paper_foam_types" not in self.config:
            self.config["paper_foam_types"] = {}
        self.config["paper_foam_types"][paper] = foam_types
        self.save_config()

    def is_foam_type_in_use(self, foam_type: str) -> bool:
        """Check if a foam type is being used in any paper"""
        paper_foam_types = self.config.get("paper_foam_types", {})
        active_papers = set(self.config.get("papers", []))
        for paper, foams in paper_foam_types.items():
            if paper not in active_papers:
                continue
            if foam_type in foams:
                return True
        return False

    def get_unused_foam_types(self) -> List[str]:
        """Get foam types that are not used in any paper"""
        all_foams = self.get_foam_types()
        unused = []
        for foam in all_foams:
            if not self.is_foam_type_in_use(foam):
                unused.append(foam)
        return unused

    # ----- Paper management -----
    def get_papers(self) -> List[str]:
        """Get list of papers (workspaces)"""
        return self.config["papers"]

    def add_paper(self, paper: str):
        """Add a new paper/workspace"""
        if paper not in self.config["papers"]:
            self.config["papers"].append(paper)
            self.save_config()

    def remove_paper(self, paper: str) -> bool:
        """Remove a paper and its saved paths. Keeps at least one paper."""
        if paper in self.config["papers"] and len(self.config["papers"]) > 1:
            self.config["papers"].remove(paper)
            # Clean up module paths for this paper
            for module in list(self.config["module_paths"].keys()):
                if isinstance(self.config["module_paths"][module], dict) and paper in self.config["module_paths"][module]:
                    del self.config["module_paths"][module][paper]
            # Clean up paper-specific associations
            paper_foams = self.config.get("paper_foam_types") or {}
            if paper in paper_foams:
                del paper_foams[paper]
            if paper in self.config.get("all_results_paths", {}):
                del self.config["all_results_paths"][paper]
            if paper in self.config.get("paper_root_paths", {}):
                del self.config["paper_root_paths"][paper]
            # Adjust current paper if needed
            if self.config["current_paper"] == paper:
                self.config["current_paper"] = self.config["papers"][0]
            self.save_config()
            return True
        return False

    def get_current_paper(self) -> str:
        return self.config["current_paper"]

    def set_current_paper(self, paper: str):
        if paper not in self.config["papers"]:
            self.add_paper(paper)
        self.config["current_paper"] = paper
        self.save_config()
    
    def add_foam_type(self, foam_type: str):
        """Add a new foam type"""
        if foam_type not in self.config["foam_types"]:
            self.config["foam_types"].append(foam_type)
            # Auto-create folder structure if in a paper context
            self.create_foam_folders_if_needed(foam_type)
            self.save_config()
    
    def create_foam_folders_if_needed(self, foam_type: str):
        """Create folder structure for new foam type in current paper if paper structure exists"""
        try:
            import os
            from pathlib import Path
            
            current_paper = self.get_current_paper()
            # Use stored paper root; do not assume defaults
            paper_root = self.config.get("paper_root_paths", {}).get(current_paper)
            if not paper_root:
                return
            base_path = Path(paper_root)
            
            # Check if paper folder exists (indicating it was created via NewPaper)
            if base_path.exists():
                # Create foam folders with Input/Output subfolders
                foam_path = base_path / foam_type
                foam_folders = [
                    # PDR folders
                    foam_path / "PDR" / "Input",
                    foam_path / "PDR" / "Output",
                    
                    # DSC folders
                    foam_path / "DSC" / "Input", 
                    foam_path / "DSC" / "Output",
                    
                    # SEM folders
                    foam_path / "SEM" / "Input",
                    foam_path / "SEM" / "Output",
                    
                    # Open-cell content folders
                    foam_path / "Open-cell content" / "Input",
                    foam_path / "Open-cell content" / "Output"
                ]
                
                for folder in foam_folders:
                    folder.mkdir(parents=True, exist_ok=True)
                    
                print(f"Created folder structure for {foam_type} in {current_paper}")
        except Exception as e:
            print(f"Could not auto-create folders for {foam_type}: {e}")
    
    def remove_foam_type(self, foam_type: str):
        """Remove a foam type"""
        if foam_type in self.config["foam_types"] and len(self.config["foam_types"]) > 1:
            self.config["foam_types"].remove(foam_type)

            # Clean up module paths for this foam type across all papers
            for module, by_paper in list(self.config.get("module_paths", {}).items()):
                if not isinstance(by_paper, dict):
                    continue
                for paper, paths_by_foam in list(by_paper.items()):
                    if isinstance(paths_by_foam, dict) and foam_type in paths_by_foam:
                        del paths_by_foam[foam_type]
                # Drop empty paper entries
                for paper in list(by_paper.keys()):
                    if isinstance(by_paper[paper], dict) and not by_paper[paper]:
                        del by_paper[paper]

            # Remove from paper foam assignments
            paper_foams = self.config.get("paper_foam_types") or {}
            for paper, foams in paper_foams.items():
                if foam_type in foams:
                    foams.remove(foam_type)
            
            # Change current if needed
            if self.config["current_foam_type"] == foam_type:
                self.config["current_foam_type"] = self.config["foam_types"][0]

            self.save_config()
            return True
        return False
    
    def get_module_paths(self, module: str, foam_type: Optional[str] = None, paper: Optional[str] = None) -> Dict:
        """Get saved paths for a specific module, scoped by paper and foam type"""
        if paper is None:
            paper = self.get_current_paper()
        scope = self.get_module_scope(module)
        if scope == "paper":
            # Store/retrieve under a global key for this paper (no foam split)
            foam_key = self._GLOBAL_KEY
        else:
            if foam_type is None:
                foam_type = self.get_current_foam_type()
            foam_key = foam_type

        mp = self.config.setdefault("module_paths", {})
        mod = mp.setdefault(module, {})
        by_paper = mod.setdefault(paper, {})
        # Primary lookup
        by_ft = by_paper.setdefault(foam_key, {})

        # Backward-compatibility: for paper-scoped modules, fall back to current foam_type entry if exists
        if scope == "paper" and not by_ft:
            cur_ft = self.get_current_foam_type()
            legacy = by_paper.get(cur_ft)
            if isinstance(legacy, dict) and legacy:
                # don't migrate automatically; just return legacy as view
                return legacy
        return by_ft
    
    def save_module_paths(self, module: str, foam_type: Optional[str], paths: Dict, paper: Optional[str] = None):
        """Save paths for a specific module, scoped by paper and foam type"""
        if paper is None:
            paper = self.get_current_paper()
        scope = self.get_module_scope(module)
        if scope == "paper":
            foam_key = self._GLOBAL_KEY
        else:
            if foam_type is None:
                foam_type = self.get_current_foam_type()
            foam_key = foam_type

        # Add timestamp
        paths = dict(paths)
        paths["last_used"] = datetime.now().isoformat()

        mp = self.config.setdefault("module_paths", {})
        mod = mp.setdefault(module, {})
        by_paper = mod.setdefault(paper, {})
        by_ft = by_paper.setdefault(foam_key, {})
        by_ft.update(paths)
        self.save_config()
    
    def get_suggested_paths(self, module: str, foam_type: Optional[str] = None, paper: Optional[str] = None) -> Dict:
        """Get suggested paths based on paper, foam type, and module"""
        if paper is None:
            paper = self.get_current_paper()
        scope = self.get_module_scope(module)
        if scope == "paper+foam":
            if foam_type is None:
                foam_type = self.get_current_foam_type()
            foam_key = foam_type
        else:
            # paper-scoped: ignore foam in path suggestions
            foam_key = None

        # Prefer saved paths
        saved_paths = self.get_module_paths(module, foam_type=foam_key if foam_key else None, paper=paper)
        if saved_paths:
            return saved_paths

        # Determine paper root (no hardcoded defaults)
        base_path = self.config.get("paper_root_paths", {}).get(paper)
        if not base_path:
            return {}
        if scope == "paper+foam":
            suggestions = {
                "input_folder": os.path.join(base_path, foam_key, module, "Input"),
                "output_folder": os.path.join(base_path, foam_key, module, "Output"),
            }
        else:
            suggestions = {
                "input_folder": os.path.join(base_path, module, "Input"),
                "output_folder": os.path.join(base_path, module, "Output"),
            }

        # Module-specific suggestions
        if module == "PDR":
            # PDR is foam-scoped by default
            ft = foam_key if foam_key else self.get_current_foam_type()
            # Suggest typical folders and results file using new naming standard
            suggestions["input_folder"] = os.path.join(base_path, ft, "PDR", "Input")
            suggestions["output_folder"] = os.path.join(base_path, ft, "PDR", "Output")
            suggestions["registros_file"] = os.path.join(base_path, ft, "PDR", "Output", f"PDR_Results_{ft}.xlsx")
        elif module == "OC":
            # OC is foam-scoped, but Density.xlsx is paper-level
            ft = foam_key if foam_key else self.get_current_foam_type()
            # Use the real folder name created by New paper
            suggestions["input_folder"] = os.path.join(base_path, ft, "Open-cell content", "Input")
            suggestions["output_folder"] = os.path.join(base_path, ft, "Open-cell content", "Output")
            suggestions["results_file"] = os.path.join(base_path, ft, "Open-cell content", "Output", f"OC_Results_{ft}.xlsx")
            suggestions["density_file"] = os.path.join(base_path, "Density.xlsx")  # Paper-level file
        elif module == "DSC":
            # DSC is foam-scoped with analysis type specific naming
            ft = foam_key if foam_key else self.get_current_foam_type()
            suggestions["input_folder"] = os.path.join(base_path, ft, "DSC", "Input")
            suggestions["output_folder"] = os.path.join(base_path, ft, "DSC", "Output")
            # Standard DSC results file naming
            suggestions["results_file"] = os.path.join(base_path, ft, "DSC", "Output", f"DSC_Results_{ft}.xlsx")
        elif module == "Combine":
            # Combine is paper-scoped (combines all foam types)
            # Default output to paper-level Results folder
            suggestions["output_folder"] = os.path.join(base_path, "Results")
        elif module == "SEM":
            # SEM is foam-scoped
            ft = foam_key if foam_key else self.get_current_foam_type()
            suggestions["input_folder"] = os.path.join(base_path, ft, "SEM", "Input")
            suggestions["output_folder"] = os.path.join(base_path, ft, "SEM", "Output")
        elif module in ("Combine", "Analysis"):
            # Per-paper All_Results default (in Results folder for Combine)
            results_base = os.path.join(base_path, "Results") if module == "Combine" else base_path
            suggestions["results_file"] = os.path.join(results_base, "All_Results.xlsx")
            suggestions["density_file"] = os.path.join(base_path, "Density.xlsx")  # Paper-level file

        return suggestions

    # ----- Paper root path helpers -----
    def get_paper_root_path(self, paper: Optional[str] = None) -> Optional[str]:
        if paper is None:
            paper = self.get_current_paper()
        return self.config.get("paper_root_paths", {}).get(paper)

    def set_paper_root_path(self, paper: str, path: str):
        self.config.setdefault("paper_root_paths", {})[paper] = path
        self.save_config()
    
    def relocate_paper_root(self, paper: str, new_root: str) -> dict:
        """Update the paper base directory and remap stored absolute paths."""
        if not new_root:
            raise ValueError("New root path cannot be empty.")
        normalized_new = os.path.normpath(new_root)
        old_root = self.get_paper_root_path(paper)
        self.config.setdefault("paper_root_paths", {})[paper] = normalized_new

        if not old_root:
            self.save_config()
            return {"updated": 0, "skipped": 0, "old_root": None, "new_root": normalized_new}

        old_norm = os.path.normpath(old_root)

        def remap_path(value: str) -> str | None:
            if not value or not isinstance(value, str):
                return None
            try:
                value_norm = os.path.normpath(value)
                common = os.path.commonpath([old_norm, value_norm])
            except Exception:
                return None
            if common != old_norm:
                return None
            try:
                rel = os.path.relpath(value_norm, old_norm)
            except Exception:
                return None
            return os.path.normpath(os.path.join(normalized_new, rel))

        updated = 0
        skipped = 0

        module_paths = self.config.get("module_paths", {})
        for module, by_paper in module_paths.items():
            if not isinstance(by_paper, dict):
                continue
            paper_paths = by_paper.get(paper)
            if not isinstance(paper_paths, dict):
                continue
            for foam_key, paths in paper_paths.items():
                if not isinstance(paths, dict):
                    continue
                for key, value in list(paths.items()):
                    if key == "last_used" or not isinstance(value, str):
                        continue
                    new_value = remap_path(value)
                    if new_value and new_value != value:
                        paths[key] = new_value
                        updated += 1
                    else:
                        skipped += 1

        all_results = self.config.setdefault("all_results_paths", {})
        current_results = all_results.get(paper)
        if isinstance(current_results, str):
            new_value = remap_path(current_results)
            if new_value and new_value != current_results:
                all_results[paper] = new_value
                updated += 1
            elif current_results:
                skipped += 1

        self.save_config()
        return {"updated": updated, "skipped": skipped, "old_root": old_norm, "new_root": normalized_new}



    def get_recent_foam_types(self, limit: int = 5) -> List[str]:
        """Get recently used foam types across papers"""
        foam_usage = []
        for module, by_paper in self.config.get("module_paths", {}).items():
            if not isinstance(by_paper, dict):
                continue
            for paper, by_ft in by_paper.items():
                for foam_type, paths in by_ft.items():
                    if isinstance(paths, dict) and "last_used" in paths:
                        foam_usage.append({
                            "foam_type": foam_type,
                            "last_used": paths["last_used"]
                        })

        foam_usage.sort(key=lambda x: x["last_used"], reverse=True)
        recent_types: List[str] = []
        for item in foam_usage:
            if item["foam_type"] not in recent_types:
                recent_types.append(item["foam_type"])
                if len(recent_types) >= limit:
                    break
        return recent_types

    # ----- All_Results per paper helpers -----
    def get_all_results_path(self, paper: Optional[str] = None) -> Optional[str]:
        if paper is None:
            paper = self.get_current_paper()
        return self.config.get("all_results_paths", {}).get(paper)

    def set_all_results_path(self, path: str, paper: Optional[str] = None):
        if paper is None:
            paper = self.get_current_paper()
        self.config.setdefault("all_results_paths", {})[paper] = path
        self.save_config()

    # ----- Module scope management -----
    def get_module_scope(self, module: str) -> str:
        """Return scope for module: 'paper' or 'paper+foam' (default 'paper+foam')."""
        scopes = self.config.get("module_scopes") or {}
        scope = scopes.get(module)
        if scope in ("paper", "paper+foam"):
            return scope
        # default for unknown modules
        return "paper+foam"

    def set_module_scope(self, module: str, scope: str):
        """Set scope for a module and persist."""
        if scope not in ("paper", "paper+foam"):
            raise ValueError("scope must be 'paper' or 'paper+foam'")
        scopes = self.config.setdefault("module_scopes", {})
        scopes[module] = scope
        self.save_config()


class FoamTypeSelector:
    """GUI component for selecting foam type"""
    def __init__(self, parent, foam_manager: FoamTypeManager, on_change_callback=None):
        self.parent = parent
        self.foam_manager = foam_manager
        self.on_change_callback = on_change_callback
        
        self.create_widgets()
        
    def create_widgets(self):
        """Create foam type selection widgets"""
        import tkinter as tk
        from tkinter import ttk, messagebox, simpledialog
        
        # Foam type frame
        foam_frame = ttk.LabelFrame(self.parent, text="Foam Type", padding="5")
        foam_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        foam_frame.columnconfigure(1, weight=1)
        
        # Foam type selection
        ttk.Label(foam_frame, text="Current Type:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        self.foam_var = tk.StringVar(value=self.foam_manager.get_current_foam_type())
        self.foam_combo = ttk.Combobox(foam_frame, textvariable=self.foam_var, values=self.foam_manager.get_foam_types_for_paper(), state="readonly")
        self.foam_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        self.foam_combo.bind("<<ComboboxSelected>>", self.on_foam_type_changed)
        
    # Note: Removed disabled Add New / Remove buttons per UI cleanup
        
        # Recent types label
        recent_types = self.foam_manager.get_recent_foam_types(3)
        if recent_types:
            recent_text = "Recent: " + ", ".join(recent_types[:3])
            ttk.Label(foam_frame, text=recent_text, font=('Arial', 8), foreground='gray').grid(row=1, column=1, sticky=tk.W, pady=(5, 0))
    
    def on_foam_type_changed(self, event=None):
        """Handle foam type change"""
        new_type = self.foam_var.get()
        self.foam_manager.set_current_foam_type(new_type)
        
        if self.on_change_callback:
            self.on_change_callback(new_type)
    
    def add_foam_type(self):
        """Add a new foam type"""
        import tkinter.simpledialog as simpledialog
        from tkinter import messagebox
        
        new_type = simpledialog.askstring("New Foam Type", "Enter foam type name:")
        if new_type and new_type.strip():
            new_type = new_type.strip()
            if new_type not in self.foam_manager.get_foam_types():
                self.foam_manager.add_foam_type(new_type)
                self.foam_combo['values'] = self.foam_manager.get_foam_types()
                self.foam_var.set(new_type)
                self.foam_manager.set_current_foam_type(new_type)
                
                if self.on_change_callback:
                    self.on_change_callback(new_type)
                    
                messagebox.showinfo("Success", f"Foam type '{new_type}' added successfully!")
            else:
                messagebox.showwarning("Warning", f"Foam type '{new_type}' already exists!")
    
    def remove_foam_type(self):
        """Remove current foam type"""
        from tkinter import messagebox
        
        current_type = self.foam_var.get()
        if len(self.foam_manager.get_foam_types()) <= 1:
            messagebox.showwarning("Warning", "Cannot remove the last foam type!")
            return
        
        result = messagebox.askyesno("Confirm Remove", f"Remove foam type '{current_type}' and all its saved paths?")
        if result:
            if self.foam_manager.remove_foam_type(current_type):
                self.foam_combo['values'] = self.foam_manager.get_foam_types()
                self.foam_var.set(self.foam_manager.get_current_foam_type())
                
                if self.on_change_callback:
                    self.on_change_callback(self.foam_manager.get_current_foam_type())
                    
                messagebox.showinfo("Success", f"Foam type '{current_type}' removed successfully!")
    
    def get_selected_foam_type(self) -> str:
        """Get currently selected foam type"""
        return self.foam_var.get()
    
    def refresh(self):
        """Refresh the foam type selector"""
        self.foam_combo['values'] = self.foam_manager.get_foam_types_for_paper()
        self.foam_var.set(self.foam_manager.get_current_foam_type())


class FoamTypeDialog:
    """Modal dialog to select or manage foam type at app startup"""
    def __init__(self, parent, foam_manager: FoamTypeManager):
        import tkinter as tk
        from tkinter import ttk
        self.parent = parent
        self.top = tk.Toplevel(parent)
        self.top.title("Select Foam Type")
        self.top.transient(parent)
        self.top.grab_set()
        self.top.resizable(False, False)

        self.foam_manager = foam_manager

        frame = ttk.Frame(self.top, padding=15)
        frame.grid(row=0, column=0)

        ttk.Label(frame, text="Select Foam Type", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=3, pady=(0, 10))

        # Reuse selector widget
        self.selector = FoamTypeSelector(frame, foam_manager, on_change_callback=None)

        # Buttons
        btns = ttk.Frame(frame)
        btns.grid(row=2, column=0, columnspan=4, pady=(10, 0), sticky=(tk.E))
        ttk.Button(btns, text="New Paper", command=self.new_paper).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btns, text="Manage Papers", command=self.manage_papers).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(btns, text="Manage Foams", command=self.manage_foams).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(btns, text="OK", command=self.on_ok).grid(row=0, column=3, padx=(0, 8))
        ttk.Button(btns, text="Cancel", command=self.on_cancel).grid(row=0, column=4)

        # Center on parent
        self.top.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.top.winfo_width() // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.top.winfo_height() // 2)
        self.top.geometry(f"+{x}+{y}")

        self.result = None

    def on_ok(self):
        # Persist current selection
        self.foam_manager.set_current_foam_type(self.selector.get_selected_foam_type())
        self.result = self.selector.get_selected_foam_type()
        self.top.grab_release()
        self.top.destroy()

    def on_cancel(self):
        # Keep current foam type
        self.result = self.foam_manager.get_current_foam_type()
        self.top.grab_release()
        self.top.destroy()


class PaperSelector:
    """GUI component for selecting paper/workspace"""
    def __init__(self, parent, foam_manager: FoamTypeManager, on_change_callback=None):
        self.parent = parent
        self.foam_manager = foam_manager
        self.on_change_callback = on_change_callback
        self.create_widgets()

    def create_widgets(self):
        import tkinter as tk
        from tkinter import ttk

        frame = ttk.LabelFrame(self.parent, text="Paper", padding="5")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="Current Paper:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.paper_var = tk.StringVar(value=self.foam_manager.get_current_paper())
        self.paper_combo = ttk.Combobox(frame, textvariable=self.paper_var, values=self.foam_manager.get_papers(), state="readonly")
        self.paper_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        self.paper_combo.bind("<<ComboboxSelected>>", self.on_paper_changed)

        ttk.Button(frame, text="❌ Remove", command=self.remove_paper).grid(row=0, column=2)

    def on_paper_changed(self, event=None):
        new_paper = self.paper_var.get()
        self.foam_manager.set_current_paper(new_paper)
        if self.on_change_callback:
            self.on_change_callback(new_paper)

    def remove_paper(self):
        from tkinter import messagebox
        current = self.paper_var.get()
        if len(self.foam_manager.get_papers()) <= 1:
            messagebox.showwarning("Warning", "Cannot remove the last paper!")
            return
        foams = self.foam_manager.get_foam_types_for_paper(current)
        warning = f"Remove paper '{current}' and its saved paths?"
        if foams:
            warning += f"\n\nFoams: {', '.join(foams)}"
        result = messagebox.askyesno("Confirm Remove", warning)
        if result:
            if self.foam_manager.remove_paper(current):
                self.paper_combo['values'] = self.foam_manager.get_papers()
                self.paper_var.set(self.foam_manager.get_current_paper())
                if self.on_change_callback:
                    self.on_change_callback(self.foam_manager.get_current_paper())
                messagebox.showinfo("Success", f"Paper '{current}' removed successfully!")


class PaperDialog:
    """Modal dialog to select or manage paper at app startup"""
    def __init__(self, parent, foam_manager: FoamTypeManager):
        import tkinter as tk
        from tkinter import ttk

        self.parent = parent
        self.foam_manager = foam_manager

        self.top = tk.Toplevel(parent)
        self.top.title("Select Paper")
        self.top.transient(parent)
        self.top.grab_set()
        self.top.resizable(False, False)

        frame = ttk.Frame(self.top, padding=15)
        frame.grid(row=0, column=0)

        ttk.Label(frame, text="Select Paper", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=4, pady=(0, 10))
        self.selector = PaperSelector(frame, foam_manager, on_change_callback=None)

        # Buttons frame (include New Paper)
        btns = ttk.Frame(frame)
        btns.grid(row=2, column=0, columnspan=4, pady=(10, 0), sticky=(tk.E))

        ttk.Button(btns, text="🆕 New Paper", command=self.new_paper).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btns, text="⚙️ Manage Foams", command=self.manage_foams).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(btns, text="OK", command=self.on_ok).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(btns, text="Cancel", command=self.on_cancel).grid(row=0, column=3)

        # Center on parent
        self.top.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.top.winfo_width() // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.top.winfo_height() // 2)
        self.top.geometry(f"+{x}+{y}")

        self.result = None

    def _refresh_paper_combo(self):
        papers = self.foam_manager.get_papers()
        self.selector.paper_combo['values'] = papers
        current = self.foam_manager.get_current_paper()
        if current not in papers and papers:
            current = papers[0]
            self.foam_manager.set_current_paper(current)
        if papers:
            self.selector.paper_var.set(self.foam_manager.get_current_paper())
        else:
            self.selector.paper_var.set("")

    def manage_papers(self):
        """Open manage-papers dialog and refresh combo."""
        try:
            dialog = ManagePapersDialog(self.parent, self.foam_manager)
            self.top.wait_window(dialog.top)
            self._refresh_paper_combo()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Failed to manage papers: {str(e)}")

    def manage_foams(self):
        """Manage foam types for existing paper"""
        try:
            current_paper = self.selector.paper_var.get()
            dialog = ManageFoamsDialog(self.parent, self.foam_manager, current_paper)
            self.top.wait_window(dialog.top)
            self._refresh_paper_combo()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Failed to manage foams: {str(e)}")

    def on_ok(self):
        self.foam_manager.set_current_paper(self.selector.paper_var.get())
        self.result = self.selector.paper_var.get()
        self.top.grab_release()
        self.top.destroy()

    def on_cancel(self):
        self.result = self.foam_manager.get_current_paper()
        self.top.grab_release()
        self.top.destroy()

    def new_paper(self):
        """Open NewPaperDialog; upon success, select the new paper in this dialog."""
        try:
            dialog = NewPaperDialog(self.top, self.foam_manager)
            self.top.wait_window(dialog.top)
            if getattr(dialog, 'result', None) and dialog.result.get('paper_name'):
                new_paper = dialog.result['paper_name']
                # Ensure it is set current
                self.foam_manager.set_current_paper(new_paper)
                # Refresh combobox values and select new paper
                self._refresh_paper_combo()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Failed to create new paper: {str(e)}")


class NewPaperDialog:
    """Dialog to create a new paper with folder structure and templates"""
    def __init__(self, parent, foam_manager: FoamTypeManager):
        import tkinter as tk
        from tkinter import ttk
        import os
        import pandas as pd
        from pathlib import Path
        
        self.parent = parent
        self.foam_manager = foam_manager
        self.result = None
        
        self.top = tk.Toplevel(parent)
        self.top.title("Create New Paper")
        self.top.transient(parent)
        self.top.grab_set()
        self.top.resizable(False, False)
        
        main_frame = ttk.Frame(self.top, padding=20)
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        ttk.Label(main_frame, text="Create New Paper", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 15))
        
        # Paper name
        ttk.Label(main_frame, text="Paper Name:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        self.paper_name_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.paper_name_var, width=30).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Base path
        ttk.Label(main_frame, text="Base Path:").grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        # No default path; user must choose explicitly
        self.base_path_var = tk.StringVar(value="")
        path_frame = ttk.Frame(main_frame)
        path_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        path_frame.columnconfigure(0, weight=1)
        ttk.Entry(path_frame, textvariable=self.base_path_var, width=25).grid(row=0, column=0, sticky=(tk.W, tk.E))
        ttk.Button(path_frame, text="Browse", command=self.browse_base_path).grid(row=0, column=1, padx=(5, 0))
        
        # Foam types selection
        ttk.Label(main_frame, text="Select Foam Types:").grid(row=3, column=0, sticky=(tk.W, tk.N), pady=(15, 5))
        
        foam_frame = ttk.LabelFrame(main_frame, text="Available Foam Types", padding=10)
        foam_frame.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(15, 5))
        
        self.foam_vars = {}
        available_foams = self.foam_manager.get_foam_types()
        
        for i, foam in enumerate(available_foams):
            var = tk.BooleanVar(value=True)  # Default selected
            self.foam_vars[foam] = var
            ttk.Checkbutton(foam_frame, text=foam, variable=var).grid(row=i//2, column=i%2, sticky=tk.W, padx=(0, 15), pady=2)
        
        # Add custom foam type
        custom_frame = ttk.Frame(foam_frame)
        custom_frame.grid(row=(len(available_foams)//2) + 1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        custom_frame.columnconfigure(0, weight=1)
        
        ttk.Label(custom_frame, text="Add Custom:").grid(row=0, column=0, sticky=tk.W)
        self.custom_foam_var = tk.StringVar()
        ttk.Entry(custom_frame, textvariable=self.custom_foam_var, width=15).grid(row=0, column=1, padx=(5, 5))
        ttk.Button(custom_frame, text="Add", command=self.add_custom_foam).grid(row=0, column=2)
        
        # What will be created
        info_frame = ttk.LabelFrame(main_frame, text="What will be created", padding=10)
        info_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(15, 0))
        
        info_text = """• Paper folder structure under base path
• DoE.xlsx template
• Density.xlsx template  
• Folder structure per foam type with Input/Output subfolders:
  - PDR: Input, Output
  - DSC: Input, Output  
  - SEM: Input (images), Output (histogram results from web analysis)
  - Open-cell content: Input, Output
• Templates stored at paper root level

Note: Foam names with "/" create nested folders (e.g., "Foam_A/Type_1" → Foam_A/Type_1/)"""
        
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).grid(row=0, column=0, sticky=tk.W)
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(20, 0))
        
        ttk.Button(btn_frame, text="Create Paper", command=self.create_paper).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(btn_frame, text="Cancel", command=self.cancel).grid(row=0, column=1)
        
        # Center dialog
        self.top.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.top.winfo_width() // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.top.winfo_height() // 2)
        self.top.geometry(f"+{x}+{y}")
        
        # Focus on paper name
        self.paper_name_var.set("Paper " + str(len(self.foam_manager.get_papers()) + 1))
        
    def browse_base_path(self):
        import tkinter.filedialog as filedialog
        folder = filedialog.askdirectory(title="Select base folder for papers")
        if folder:
            self.base_path_var.set(folder)
    
    def add_custom_foam(self):
        import tkinter as tk
        from tkinter import ttk
        
        custom = self.custom_foam_var.get().strip()
        if custom and custom not in self.foam_vars:
            # Add to foam manager
            self.foam_manager.add_foam_type(custom)
            
            # Add checkbox
            foam_frame = None
            for widget in self.top.winfo_children():
                if isinstance(widget, ttk.Frame):
                    for child in widget.winfo_children():
                        if isinstance(child, ttk.LabelFrame) and "Available Foam Types" in str(child.cget('text')):
                            foam_frame = child
                            break
            
            if foam_frame:
                row = len(self.foam_vars) // 2
                col = len(self.foam_vars) % 2
                var = tk.BooleanVar(value=True)
                self.foam_vars[custom] = var
                ttk.Checkbutton(foam_frame, text=custom, variable=var).grid(row=row, column=col, sticky=tk.W, padx=(0, 15), pady=2)
                
            self.custom_foam_var.set("")
    
    def create_paper(self):
        from tkinter import messagebox
        import os
        import pandas as pd
        from pathlib import Path
        
        paper_name = self.paper_name_var.get().strip()
        if not paper_name:
            messagebox.showerror("Error", "Please enter a paper name")
            return
            
        # Get selected foam types
        selected_foams = [foam for foam, var in self.foam_vars.items() if var.get()]
        if not selected_foams:
            messagebox.showerror("Error", "Please select at least one foam type")
            return
            
        try:
            base_path = Path(self.base_path_var.get())
            paper_path = base_path / paper_name
            
            # Create folder structure
            self.create_folder_structure(paper_path, selected_foams)
            
            # Create templates (new minimal format)
            self.create_doe_template_v2(paper_path, selected_foams)
            self.create_density_template_v2(paper_path, selected_foams)
            
            # Add paper to manager and persist its root path
            self.foam_manager.add_paper(paper_name)
            self.foam_manager.set_paper_root_path(paper_name, str(paper_path))
            
            # Associate selected foam types with this paper
            self.foam_manager.set_foam_types_for_paper(paper_name, selected_foams)
            
            self.result = {
                'paper_name': paper_name,
                'foam_types': selected_foams,
                'paper_path': str(paper_path)
            }
            
            messagebox.showinfo("Success", f"Paper '{paper_name}' created successfully!\n\nPath: {paper_path}\nFoam types: {', '.join(selected_foams)}")
            self.top.grab_release()
            self.top.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create paper: {str(e)}")
    
    def create_folder_structure(self, paper_path, foam_types):
        """Create the folder structure for the paper"""
        # Main paper folders (Analysis only, Combine is per foam type)
        folders = [
            paper_path / "Analysis" / "Input",
            paper_path / "Analysis" / "Output"
        ]
        
        # Per-foam folders with Input/Output subfolders
        for foam in foam_types:
            # Use Path to handle foam names with "/" properly
            from pathlib import Path
            foam_path = paper_path / foam
            
            folders.extend([
                # PDR folders
                foam_path / "PDR" / "Input",
                foam_path / "PDR" / "Output",
                
                # DSC folders  
                foam_path / "DSC" / "Input",
                foam_path / "DSC" / "Output",
                
                # SEM folders
                foam_path / "SEM" / "Input",
                foam_path / "SEM" / "Output",
                
                # Open-cell content folders
                foam_path / "Open-cell content" / "Input",
                foam_path / "Open-cell content" / "Output",
                
                # Combine folder (per foam type)
                foam_path / "Combine",
                foam_path / "Combine" / "Previous results"
            ])
        
        # Create all folders
        for folder in folders:
            folder.mkdir(parents=True, exist_ok=True)
    
    def create_doe_template(self, paper_path, foam_types):
        """Create DoE.xlsx template with separate sheet for each foam type"""
        import pandas as pd
        
        # Save to Excel at paper root
        doe_path = paper_path / "DoE.xlsx"
        with pd.ExcelWriter(doe_path, engine='openpyxl') as writer:
            
            # Create a sheet for each foam type
            for foam in foam_types:
                # Create sample data for this foam type
                data = [{
                    'Label': '20250204',
                    'Plastic name': foam,
                    'Enterprise': '',
                    'm (g)': '10',
                    'Water (g)': 0,
                    'T (ºC)': 120,
                    'P CO2 (bar)': 200,
                    't (min)': 30,
                    'SEM measured?': 'Yes',
                    'Density measured?': 'Yes',
                    'DSC measured?': '',
                    'Picnometry measured?': '',
                    'Conductivity measured?': '',
                    'Comments': ''
                }]
                
                df = pd.DataFrame(data)
                
                # Use foam type name as sheet name (Excel sheet names have character limits)
                sheet_name = foam[:31]  # Excel sheet name limit is 31 characters
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format headers for this sheet
                worksheet = writer.sheets[sheet_name]
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
    
    def create_density_template(self, paper_path, foam_types):
        """Create Density.xlsx template with example rows"""
        import pandas as pd
        
        # Create sample data with one row per foam type
        data = []
        for i, foam in enumerate(foam_types, 1):
            data.append({
                'Label': f'{foam}_001',
                'Av Exp ρ foam (g/cm3)': 0.100,
                'Desvest Exp ρ foam (g/cm3)': 0.005,
                '%DER Exp ρ foam (g/cm3)': 5.0,
                'ρr': 0.10,
                'X': 10.0,
                'Porosity (%)': 90.0,
                'OC (%)': 85.0,
                'Polymer': foam
            })
        
        df = pd.DataFrame(data)
        
        # Save to Excel at paper root
        density_path = paper_path / "Density.xlsx"
        with pd.ExcelWriter(density_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Density', index=False)
            
            # Format headers
            worksheet = writer.sheets['Density']
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column].width = adjusted_width
    
    def cancel(self):
        self.result = None
        self.top.grab_release()
        self.top.destroy()

    # --- New minimal templates ---
    def create_doe_template_v2(self, paper_path, foam_types):
        """Create DoE.xlsx with minimal headers, one sheet per foam type."""
        import pandas as pd
        cols = ['Label', 'm (g)', 'Water (g)', 'T (°C)', 'P CO2 (bar)', 't (min)']
        doe_path = paper_path / "DoE.xlsx"
        with pd.ExcelWriter(doe_path, engine='openpyxl') as writer:
            for foam in foam_types:
                df = pd.DataFrame([{c: '' for c in cols}])
                sheet = foam[:31]
                df.to_excel(writer, sheet_name=sheet, index=False)
                ws = writer.sheets[sheet]
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value is not None:
                                max_len = max(max_len, len(str(cell.value)))
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    def create_density_template_v2(self, paper_path, foam_types):
        """Create Density.xlsx per foam with formulas on row 2."""
        import pandas as pd
        from openpyxl.utils import get_column_letter
        headers = [
            'Polymer', 'Label', 'Measure 1', 'Measure 2', 'Measure 3',
            'ρ foam (g/cm^3)', 'Desvest ρ foam (g/cm^3)', '%DER ρ foam (g/cm^3)',
            'ρr', 'X', 'Porosity (%)', '', 'ρ solid polymer (g/cm^3)',
            'Measure 1', 'Measure 2', 'Measure 3'
        ]
        density_path = paper_path / "Density.xlsx"
        with pd.ExcelWriter(density_path, engine='openpyxl') as writer:
            for foam in foam_types:
                df = pd.DataFrame([[''] * len(headers)], columns=headers)
                sheet = foam[:31]
                df.to_excel(writer, sheet_name=sheet, index=False)
                ws = writer.sheets[sheet]
                # Formulas row 2
                ws['F2'] = "=AVERAGE(C2:E2)"
                ws['G2'] = "=STDEV(C2:E2)"
                ws['M2'] = "=AVERAGE(N2:P2)"
                ws['H2'] = "=G2/F2*100"
                ws['I2'] = "=F2/$M$2"  # ρr = ρfoam / ρsolid
                ws['J2'] = "=1/I2"
                ws['K2'] = "=(1-F2/$M$2)*100"
                # Autosize
                for i in range(1, len(headers)+1):
                    letter = get_column_letter(i)
                    max_len = len(headers[i-1])
                    v = ws[f"{letter}2"].value
                    if v:
                        try:
                            max_len = max(max_len, len(str(v)))
                        except Exception:
                            pass
                    ws.column_dimensions[letter].width = min(max_len + 2, 50)


class ManagePapersDialog:
    """Dialog to review, relocate, and delete papers."""

    def __init__(self, parent, foam_manager: FoamTypeManager):
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog

        self.parent = parent
        self.foam_manager = foam_manager
        self.paper_items = {}

        self.top = tk.Toplevel(parent)
        self.top.title("Manage Papers")
        self.top.transient(parent)
        self.top.grab_set()
        self.top.resizable(True, True)

        main = ttk.Frame(self.top, padding=20)
        main.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        self.top.columnconfigure(0, weight=1)
        self.top.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(0, weight=1)

        tree_frame = ttk.LabelFrame(main, text="Papers and foams", padding=10)
        tree_frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(tree_frame, show="tree")
        self.tree.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=scroll.set)

        detail = ttk.LabelFrame(main, text="Details", padding=10)
        detail.grid(row=0, column=1, sticky=(tk.N, tk.S, tk.W, tk.E), padx=(15, 0))
        detail.columnconfigure(1, weight=1)

        self.base_var = tk.StringVar(value="–")
        self.foams_var = tk.StringVar(value="")
        self.results_var = tk.StringVar(value="–")

        ttk.Label(detail, text="Base directory:").grid(row=0, column=0, sticky=tk.W)
        self.base_label = ttk.Label(detail, textvariable=self.base_var, wraplength=260, justify=tk.LEFT)
        self.base_label.grid(row=0, column=1, sticky=(tk.W, tk.E))

        ttk.Label(detail, text="Foams:").grid(row=1, column=0, sticky=tk.NW, pady=(8, 0))
        self.foams_label = ttk.Label(detail, textvariable=self.foams_var, wraplength=260, justify=tk.LEFT)
        self.foams_label.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(8, 0))

        ttk.Label(detail, text="All_Results path:").grid(row=2, column=0, sticky=tk.W, pady=(8, 0))
        self.results_label = ttk.Label(detail, textvariable=self.results_var, wraplength=260, justify=tk.LEFT)
        self.results_label.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=(8, 0))

        btns = ttk.Frame(main)
        btns.grid(row=1, column=1, sticky=tk.E, padx=(15, 0), pady=(12, 0))
        self.foams_btn = ttk.Button(btns, text="Manage Foams...", command=self.manage_foams)
        self.foams_btn.grid(row=0, column=0, padx=(0, 6))
        self.change_btn = ttk.Button(btns, text="Change Directory...", command=self.change_directory)
        self.change_btn.grid(row=0, column=1, padx=(0, 6))
        self.open_btn = ttk.Button(btns, text="Open Folder", command=self.open_folder)
        self.open_btn.grid(row=0, column=2, padx=(0, 6))
        self.delete_btn = ttk.Button(btns, text="Delete Paper", command=self.delete_paper)
        self.delete_btn.grid(row=0, column=3)

        close_frame = ttk.Frame(main)
        close_frame.grid(row=2, column=0, columnspan=2, sticky=tk.E, pady=(12, 0))
        ttk.Button(close_frame, text="Close", command=self.close).grid(row=0, column=0)

        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self._populate_tree()
        self._update_buttons_state(bool(self.paper_items))

        self.top.update_idletasks()
        try:
            x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.top.winfo_width() // 2)
            y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.top.winfo_height() // 2)
            self.top.geometry(f"+{x}+{y}")
        except Exception:
            pass
        self.top.protocol("WM_DELETE_WINDOW", self.close)

    def _populate_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.paper_items.clear()

        papers = self.foam_manager.get_papers()
        for paper in papers:
            foams = self.foam_manager.get_foam_types_for_paper(paper)
            label = f"{paper} ({len(foams)} foams)" if foams else f"{paper} (0 foams)"
            node = self.tree.insert("", "end", text=label, open=True)
            self.paper_items[node] = paper
            for foam in foams:
                self.tree.insert(node, "end", text=foam)

        if self.paper_items:
            first_item = next(iter(self.paper_items))
            self.tree.selection_set(first_item)
            self.tree.focus(first_item)
            self._update_details(self.paper_items[first_item])
        else:
            self._update_details(None)
        self._update_buttons_state(bool(self.paper_items))

    def _current_selection(self):
        selection = self.tree.selection()
        if not selection:
            return None, None
        item = selection[0]
        paper = self.paper_items.get(item)
        if paper is None:
            parent = self.tree.parent(item)
            paper = self.paper_items.get(parent)
            item = parent if paper else None
        return item, paper

    def _on_select(self, _event=None):
        _item, paper = self._current_selection()
        self._update_details(paper)
        self._update_buttons_state(bool(paper))

    def _update_details(self, paper):
        if not paper:
            self.base_var.set("-")
            self.results_var.set("-")
            self.foams_var.set("")
            return
        base = self.foam_manager.get_paper_root_path(paper) or "Not set"
        results = self.foam_manager.get_all_results_path(paper) or "Not saved"
        foams = self.foam_manager.get_foam_types_for_paper(paper)
        self.base_var.set(base)
        self.results_var.set(results)
        self.foams_var.set(", ".join(foams) if foams else "No foams assigned")

    def _update_buttons_state(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        for btn in (self.foams_btn, self.change_btn, self.open_btn, self.delete_btn):
            btn.configure(state=state)

    def change_directory(self):
        from tkinter import filedialog, messagebox
        item, paper = self._current_selection()
        if not paper:
            return
        current = self.foam_manager.get_paper_root_path(paper)
        initialdir = current if current and os.path.exists(current) else None
        new_path = filedialog.askdirectory(title=f"Select directory for {paper}", initialdir=initialdir)
        if not new_path:
            return
        try:
            result = self.foam_manager.relocate_paper_root(paper, new_path)
            msg = [f"Paper '{paper}' directory updated."]
            msg.append(f"Updated {result.get('updated', 0)} stored paths.")
            skipped = result.get('skipped', 0)
            if skipped:
                msg.append(f"Skipped {skipped} paths outside the previous base.")
            messagebox.showinfo("Paper updated", "\n".join(msg))
            self._update_details(paper)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update paper directory: {e}")

    def open_folder(self):
        from tkinter import messagebox
        import os
        _item, paper = self._current_selection()
        if not paper:
            return
        base = self.foam_manager.get_paper_root_path(paper)
        if not base:
            messagebox.showwarning("No folder", "This paper does not have a base directory assigned.")
            return
        if not os.path.exists(base):
            messagebox.showwarning("Missing folder", f"Folder not found:\n{base}")
            return
        try:
            os.startfile(base)
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open folder: {e}")

    def delete_paper(self):
        from tkinter import messagebox
        item, paper = self._current_selection()
        if not paper:
            return
        foams = self.foam_manager.get_foam_types_for_paper(paper)
        warning = f"Delete paper '{paper}' from PressTech?"
        if foams:
            warning += f"\n\nFoams: {', '.join(foams)}"
        if not messagebox.askyesno("Confirm delete", warning):
            return
        if not self.foam_manager.remove_paper(paper):
            messagebox.showwarning("Cannot delete", "At least one paper must remain in the workspace.")
            return
        messagebox.showinfo("Paper deleted", f"Paper '{paper}' has been removed.")
        self._populate_tree()
        self._update_buttons_state(bool(self.paper_items))

    def manage_foams(self):
        from tkinter import messagebox
        _item, paper = self._current_selection()
        if not paper:
            return
        try:
            dialog = ManageFoamsDialog(self.top, self.foam_manager, paper)
            self.top.wait_window(dialog.top)
            previous_paper = paper if paper in self.foam_manager.get_papers() else None
            self._populate_tree()
            if previous_paper:
                for node, value in self.paper_items.items():
                    if value == previous_paper:
                        self.tree.selection_set(node)
                        self.tree.focus(node)
                        self._update_details(previous_paper)
                        break
        except Exception as e:
            messagebox.showerror("Error", f"Failed to manage foams: {e}")

    def close(self):
        try:
            self.top.grab_release()
        except Exception:
            pass
        self.top.destroy()


class ManageFoamsDialog:
    """Dialog to manage foam types for an existing paper"""
    def __init__(self, parent, foam_manager: FoamTypeManager, paper_name: str):
        import tkinter as tk
        from tkinter import ttk
        import os
        from pathlib import Path
        
        self.parent = parent
        self.foam_manager = foam_manager
        self.paper_name = paper_name
        self.result = None
        
        self.top = tk.Toplevel(parent)
        self.top.title(f"Manage Foams - {paper_name}")
        self.top.transient(parent)
        self.top.grab_set()
        self.top.resizable(False, False)
        
        main_frame = ttk.Frame(self.top, padding=20)
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        ttk.Label(main_frame, text=f"Manage Foams for '{paper_name}'", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 15))
        
        # Current foam types for this paper
        ttk.Label(main_frame, text="Current Foams:").grid(row=1, column=0, sticky=(tk.W, tk.N), pady=(0, 5))
        
        current_frame = ttk.LabelFrame(main_frame, text="Currently Assigned", padding=10)
        current_frame.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        current_frame.columnconfigure(0, weight=1)
        self.current_frame = current_frame
        self.current_checks_frame = ttk.Frame(current_frame)
        self.current_checks_frame.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        self.current_foam_vars = {}
        current_foams = self.foam_manager.get_foam_types_for_paper(paper_name)
        self.original_foams = list(current_foams)
        
        for foam in current_foams:
            var = tk.BooleanVar(value=True)
            self.current_foam_vars[foam] = var
        self._render_current_foams()
        
        # Available foam types to add
        ttk.Label(main_frame, text="Add More:").grid(row=2, column=0, sticky=(tk.W, tk.N), pady=(15, 5))
        
        available_frame = ttk.LabelFrame(main_frame, text="Available to Add", padding=10)
        available_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(15, 5))
        available_frame.columnconfigure(0, weight=1)
        self.available_frame = available_frame
        self.available_checks_frame = ttk.Frame(available_frame)
        self.available_checks_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        self.available_foam_vars = {}
        all_foams = self.foam_manager.get_foam_types()
        available_foams = [foam for foam in all_foams if foam not in current_foams]
        
        for foam in available_foams:
            self.available_foam_vars[foam] = tk.BooleanVar(value=False)
        self._render_available_foams()
        
        self.add_selected_button = ttk.Button(available_frame, text="Add to the paper", command=self.add_selected_foams_to_paper)
        self.add_selected_button.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(8, 0))
        
        # Add custom foam type
        custom_frame = ttk.Frame(available_frame)
        custom_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        custom_frame.columnconfigure(0, weight=1)
        
        ttk.Label(custom_frame, text="Add Custom:").grid(row=0, column=0, sticky=tk.W)
        self.custom_foam_var = tk.StringVar()
        ttk.Entry(custom_frame, textvariable=self.custom_foam_var, width=15).grid(row=0, column=1, padx=(5, 5))
        ttk.Button(custom_frame, text="Add", command=self.add_custom_foam).grid(row=0, column=2)
        
        # Delete unused foam types section
        unused_foams = self.foam_manager.get_unused_foam_types()
        if unused_foams:
            ttk.Label(main_frame, text="Delete Unused:").grid(row=3, column=0, sticky=(tk.W, tk.N), pady=(15, 5))
            
            unused_frame = ttk.LabelFrame(main_frame, text="Unused Foam Types", padding=10)
            unused_frame.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(15, 5))
            
            self.unused_frame = unused_frame
            self.unused_checks_frame = ttk.Frame(unused_frame)
            self.unused_checks_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
            self.unused_foam_vars = {}
            for foam in unused_foams:
                self.unused_foam_vars[foam] = tk.BooleanVar(value=False)
            self._render_unused_foams()
            
            self.delete_unused_button = ttk.Button(unused_frame, text="Delete Selected", command=self.delete_selected_unused_foams)
            self.delete_unused_button.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(8, 0))
        else:
            self.unused_foam_vars = {}
            self.unused_frame = None
            self.unused_checks_frame = None
            self.delete_unused_button = None
        
        # Information
        info_frame = ttk.LabelFrame(main_frame, text="Actions", padding=10)
        info_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(15, 0))
        
        info_text = (
            "* Uncheck foams to remove from this paper\n"
            "* Select foams under 'Available to Add' and click 'Add to the paper'\n"
            "* New foam folder structures will be created if they don't exist\n"
            "* Templates (DoE.xlsx, Density.xlsx) will be updated with new foams\n"
            "* Use 'Delete Selected' to remove unused foams from the application"
        )
        
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).grid(row=0, column=0, sticky=tk.W)
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(20, 0))
        
        ttk.Button(btn_frame, text="Apply Changes", command=self.apply_changes).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(btn_frame, text="Cancel", command=self.cancel).grid(row=0, column=1)
        
        # Center dialog
        self.top.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.top.winfo_width() // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.top.winfo_height() // 2)
        self.top.geometry(f"+{x}+{y}")
    
    def _render_current_foams(self):
        import tkinter as tk
        from tkinter import ttk
        for widget in self.current_checks_frame.winfo_children():
            widget.destroy()
        for i, (foam, var) in enumerate(self.current_foam_vars.items()):
            ttk.Checkbutton(self.current_checks_frame, text=foam, variable=var).grid(
                row=i // 2, column=i % 2, sticky=tk.W, padx=(0, 15), pady=2
            )

    def _render_available_foams(self):
        import tkinter as tk
        from tkinter import ttk
        for widget in self.available_checks_frame.winfo_children():
            widget.destroy()
        if not self.available_foam_vars:
            ttk.Label(self.available_checks_frame, text="No additional foams available", foreground="gray").grid(row=0, column=0, sticky=tk.W)
            if hasattr(self, "add_selected_button") and self.add_selected_button:
                self.add_selected_button.configure(state="disabled")
            return
        for i, (foam, var) in enumerate(self.available_foam_vars.items()):
            ttk.Checkbutton(self.available_checks_frame, text=foam, variable=var).grid(
                row=i // 2, column=i % 2, sticky=tk.W, padx=(0, 15), pady=2
            )
        if hasattr(self, "add_selected_button") and self.add_selected_button:
            self.add_selected_button.configure(state="normal")

    def _render_unused_foams(self):
        import tkinter as tk
        from tkinter import ttk
        if not getattr(self, "unused_checks_frame", None):
            return
        for widget in self.unused_checks_frame.winfo_children():
            widget.destroy()
        if not self.unused_foam_vars:
            ttk.Label(self.unused_checks_frame, text="No unused foams", foreground="gray").grid(row=0, column=0, sticky=tk.W)
            if getattr(self, "delete_unused_button", None):
                self.delete_unused_button.configure(state="disabled")
            return
        style = ttk.Style()
        style.configure("RedFoam.TCheckbutton", foreground="red")
        for i, (foam, var) in enumerate(self.unused_foam_vars.items()):
            ttk.Checkbutton(
                self.unused_checks_frame,
                text=f"{foam} (unused)",
                variable=var,
                style="RedFoam.TCheckbutton",
            ).grid(row=i // 2, column=i % 2, sticky=tk.W, padx=(0, 15), pady=2)
        if getattr(self, "delete_unused_button", None):
            self.delete_unused_button.configure(state="normal")

    def add_selected_foams_to_paper(self):
        import tkinter as tk
        from tkinter import messagebox

        selected = [foam for foam, var in self.available_foam_vars.items() if var.get()]
        if not selected:
            messagebox.showinfo("No Selection", "Select at least one foam to add.")
            return

        for foam in selected:
            if foam not in self.current_foam_vars:
                self.current_foam_vars[foam] = tk.BooleanVar(value=True)
        for foam in selected:
            self.available_foam_vars.pop(foam, None)

        self._render_current_foams()
        self._render_available_foams()

    def delete_selected_unused_foams(self):
        from tkinter import messagebox

        selected = [foam for foam, var in self.unused_foam_vars.items() if var.get()]
        if not selected:
            messagebox.showinfo("No Selection", "Select at least one foam to delete.")
            return

        if not messagebox.askyesno(
            "Confirm Deletion",
            "Permanently delete these unused foam types from the application?\n\n"
            + ", ".join(selected)
            + "\n\nThis action cannot be undone.",
        ):
            return

        deleted = []
        for foam in selected:
            if self.foam_manager.remove_foam_type(foam):
                deleted.append(foam)

        if deleted:
            messagebox.showinfo("Foams Deleted", f"Deleted: {', '.join(deleted)}")

        for foam in deleted:
            self.unused_foam_vars.pop(foam, None)
            self.available_foam_vars.pop(foam, None)
            self.current_foam_vars.pop(foam, None)

        self._render_unused_foams()
        self._render_available_foams()
        self._render_current_foams()

    def add_custom_foam(self):
        import tkinter as tk
        from tkinter import messagebox
        
        custom = self.custom_foam_var.get().strip()
        if not custom:
            return
        if custom in self.current_foam_vars or custom in self.available_foam_vars:
            messagebox.showwarning("Duplicate", f"Foam type '{custom}' is already listed.")
            return

        self.foam_manager.add_foam_type(custom)
        self.available_foam_vars[custom] = tk.BooleanVar(value=True)
        self._render_available_foams()
        self.custom_foam_var.set("")
    
    def apply_changes(self):
        from tkinter import messagebox
        from pathlib import Path
        import os
        
        try:
            # Get final foam list for this paper
            final_foams = []
            
            # Keep checked current foams
            for foam, var in self.current_foam_vars.items():
                if var.get():
                    final_foams.append(foam)
            
            # Add checked new foams
            for foam, var in self.available_foam_vars.items():
                if var.get():
                    final_foams.append(foam)
            
            if not final_foams:
                messagebox.showerror("Error", "Please select at least one foam type")
                return
            
            added_foams = [foam for foam in final_foams if foam not in self.original_foams]
            removed_foams = [foam for foam in self.original_foams if foam not in final_foams]

            foams_to_delete = []
            for foam, var in self.unused_foam_vars.items():
                if var.get():
                    foams_to_delete.append(foam)

            if not added_foams and not removed_foams and not foams_to_delete:
                messagebox.showinfo("No Changes", "No foam assignments were modified.")
                return

            if added_foams or removed_foams:
                change_lines = []
                if added_foams:
                    change_lines.append(f"Add to '{self.paper_name}': {', '.join(added_foams)}")
                if removed_foams:
                    change_lines.append(f"Remove from '{self.paper_name}': {', '.join(removed_foams)}")
                confirm_message = "Apply these changes?\n\n" + "\n".join(change_lines)
                confirm_message += "\n\nThis will update the foam list for the paper."
                if not messagebox.askyesno("Confirm Foam Changes", confirm_message):
                    return

            if foams_to_delete:
                result = messagebox.askyesno(
                    "Confirm Deletion",
                    f"Permanently delete these unused foam types from the application?\n\n{', '.join(foams_to_delete)}\n\nThis action cannot be undone."
                )
                if result:
                    for foam in foams_to_delete:
                        if self.foam_manager.remove_foam_type(foam):
                            print(f"Permanently deleted foam type: {foam}")
                else:
                    foams_to_delete = []
            
            # Update foam types for this paper
            self.foam_manager.set_foam_types_for_paper(self.paper_name, final_foams)
            
            # Try to create folder structure for new foams (if paper folder exists)
            try:
                paper_root = self.foam_manager.get_paper_root_path(self.paper_name)
                if paper_root:
                    base_path = Path(paper_root)
                    if base_path.exists():
                        self.create_folder_structure_for_foams(base_path, final_foams)
                        self.update_templates(base_path, final_foams)
            except Exception as e:
                print(f"Could not update folder structure: {e}")
            
            self.result = {
                'paper_name': self.paper_name,
                'foam_types': final_foams,
                'deleted_foams': foams_to_delete
            }
            
            status_msg = f"Foam types for '{self.paper_name}' updated!\n\nFoam types: {', '.join(final_foams)}"
            if foams_to_delete:
                status_msg += f"\n\nDeleted globally: {', '.join(foams_to_delete)}"
            
            messagebox.showinfo("Success", status_msg)
            self.top.grab_release()
            self.top.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update foams: {str(e)}")
    
    def create_folder_structure_for_foams(self, paper_path, foam_types):
        """Create folder structure for any missing foam types"""
        for foam in foam_types:
            foam_path = paper_path / foam
            
            folders = [
                # PDR folders
                foam_path / "PDR" / "Input",
                foam_path / "PDR" / "Output",
                
                # DSC folders  
                foam_path / "DSC" / "Input",
                foam_path / "DSC" / "Output",
                
                # SEM folders
                foam_path / "SEM" / "Input",
                foam_path / "SEM" / "Output",
                
                # Open-cell content folders
                foam_path / "Open-cell content" / "Input",
                foam_path / "Open-cell content" / "Output"
            ]
            
            for folder in folders:
                folder.mkdir(parents=True, exist_ok=True)
    
    def update_templates(self, paper_path, foam_types):
        """Update DoE.xlsx and Density.xlsx templates with new foam types"""
        import pandas as pd
        
        try:
            # Update DoE.xlsx at paper root
            doe_path = paper_path / "DoE.xlsx"
            if doe_path.exists():
                # Read existing data
                df = pd.read_excel(doe_path, sheet_name='DoE')
                existing_polymers = set(df['Polymer'].values) if 'Polymer' in df.columns else set()
                
                # Add missing foam types
                new_rows = []
                for foam in foam_types:
                    if foam not in existing_polymers:
                        new_rows.append({
                            'Label': f'{foam}_001',
                            'm(g)': 2.0,
                            'Water (g)': 0.0,
                            'T (ºC)': 140,
                            'P CO2 (bar)': 200,
                            't (min)': 3,
                            'Polymer': foam
                        })
                
                if new_rows:
                    new_df = pd.DataFrame(new_rows)
                    df = pd.concat([df, new_df], ignore_index=True)
                    
                    with pd.ExcelWriter(doe_path, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='DoE', index=False)
            
            # Update Density.xlsx at paper root
            density_path = paper_path / "Density.xlsx"
            if density_path.exists():
                # Read existing data
                df = pd.read_excel(density_path, sheet_name='Density')
                existing_polymers = set(df['Polymer'].values) if 'Polymer' in df.columns else set()
                
                # Add missing foam types
                new_rows = []
                for foam in foam_types:
                    if foam not in existing_polymers:
                        new_rows.append({
                            'Label': f'{foam}_001',
                            'Av Exp ρ foam (g/cm3)': 0.100,
                            'Desvest Exp ρ foam (g/cm3)': 0.005,
                            '%DER Exp ρ foam (g/cm3)': 5.0,
                            'ρr': 0.10,
                            'X': 10.0,
                            'Porosity (%)': 90.0,
                            'OC (%)': 85.0,
                            'Polymer': foam
                        })
                
                if new_rows:
                    new_df = pd.DataFrame(new_rows)
                    df = pd.concat([df, new_df], ignore_index=True)
                    
                    with pd.ExcelWriter(density_path, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Density', index=False)
        
        except Exception as e:
            print(f"Could not update templates: {e}")
    
    def cancel(self):
        self.result = None
        self.top.grab_release()
        self.top.destroy()
