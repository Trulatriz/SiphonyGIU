import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import re
import glob
import warnings
from pathlib import Path
from openpyxl.styles import Font, PatternFill
from .foam_type_manager import FoamTypeManager, FoamTypeSelector

# Suppress warnings for better user experience
warnings.filterwarnings("ignore", category=FutureWarning, module="pandas")


class DSCModule:
    def __init__(self, root, paper_path=None, foam_type=None):
        self.root = root
        self.root.title("DSC Analysis")
        self.root.geometry("900x700")

        # Store context
        self.paper_path = paper_path
        self.foam_manager = FoamTypeManager()

        # Current foam type
        if foam_type and foam_type in self.foam_manager.get_foam_types():
            self.current_foam_type = foam_type
        else:
            self.current_foam_type = self.foam_manager.get_current_foam_type()

        # Paths / state
        self.input_folder = None
        self.selected_dsc_files = []
        self.output_folder = None
        
        # Analysis type selection
        self.analysis_type = tk.StringVar(value="semicrystalline")
        
        # Review data after processing
        self.processed_data = []
        self.review_window = None

        self.create_widgets()

    def create_widgets(self):
        main = ttk.Frame(self.root, padding="10")
        main.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(4, weight=1)

        ttk.Label(main, text="DSC Analysis", font=("Arial", 16, "bold")).grid(row=0, column=0, pady=(0, 16))

        # Foam type selector
        self.foam_selector = FoamTypeSelector(main, self.foam_manager, self.on_foam_type_changed)

        # Setup files and folders
        setup = ttk.LabelFrame(main, text="Setup Files and Folders", padding="10")
        setup.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(4, 10))
        setup.columnconfigure(1, weight=1)

        # Input folder
        ttk.Label(setup, text="DSC Files Folder:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.input_folder_var = tk.StringVar()
        ttk.Entry(setup, textvariable=self.input_folder_var, state="readonly", width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5))
        ttk.Button(setup, text="Browse", command=self.browse_input_folder).grid(row=0, column=2)

        # Output folder
        ttk.Label(setup, text="Output Folder:").grid(row=1, column=0, sticky=tk.W, pady=(5, 5))
        self.output_folder_var = tk.StringVar()
        ttk.Entry(setup, textvariable=self.output_folder_var, state="readonly", width=60).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5))
        ttk.Button(setup, text="Browse", command=self.browse_output_folder).grid(row=1, column=2)

        # Results file
        ttk.Label(setup, text="DSC Results:").grid(row=2, column=0, sticky=tk.W, pady=(5, 0))
        self.results_file_var = tk.StringVar()
        results_row = ttk.Frame(setup)
        results_row.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E))
        results_row.columnconfigure(0, weight=1)
        ttk.Entry(results_row, textvariable=self.results_file_var, state="readonly", width=60).grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(10, 5))
        ttk.Button(results_row, text="Open/Create", command=self.open_or_create_results_file).grid(row=0, column=1)

        # Analysis type configuration
        config = ttk.LabelFrame(main, text="Analysis Configuration", padding="10")
        config.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        # Analysis type selection (without specifying polymer examples)
        ttk.Label(config, text="Analysis Type:").grid(row=0, column=0, sticky=tk.W)
        analysis_frame = ttk.Frame(config)
        analysis_frame.grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        
        ttk.Radiobutton(analysis_frame, text="Semicrystalline Analysis", 
                       variable=self.analysis_type, value="semicrystalline").grid(row=0, column=0, sticky=tk.W)
        ttk.Radiobutton(analysis_frame, text="Amorphous Analysis", 
                       variable=self.analysis_type, value="amorphous").grid(row=1, column=0, sticky=tk.W)

        # Load suggested paths
        self.load_suggested_paths()

        # Scan button
        ttk.Button(config, text="🔍 Scan DSC Files", command=self.scan_dsc_files).grid(row=1, column=0, pady=(10, 5))

        # Files table
        file_frame = ttk.LabelFrame(main, text="Selected DSC Files", padding="10")
        file_frame.grid(row=4, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        file_frame.columnconfigure(0, weight=1)
        file_frame.rowconfigure(1, weight=1)

        # Selection buttons
        sel_frame = ttk.Frame(file_frame)
        sel_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        ttk.Button(sel_frame, text="Select All", command=self.select_all_files).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(sel_frame, text="Select None", command=self.select_none_files).grid(row=0, column=1, padx=(0, 5))
        ttk.Label(sel_frame, text="(Ctrl+Click to select/deselect individual files)").grid(row=0, column=2, padx=(10, 0))

        cols = ("filename", "size", "status")
        self.file_tree = ttk.Treeview(file_frame, columns=cols, show="headings", height=8, selectmode="extended")
        self.file_tree.heading("filename", text="DSC File Name")
        self.file_tree.heading("size", text="Size (KB)")
        self.file_tree.heading("status", text="Status")
        self.file_tree.column("filename", width=420)
        self.file_tree.column("size", width=100, anchor="center")
        self.file_tree.column("status", width=160, anchor="center")
        self.file_tree.grid(row=1, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        
        vbar = ttk.Scrollbar(file_frame, orient=tk.VERTICAL, command=self.file_tree.yview)
        vbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.file_tree.configure(yscrollcommand=vbar.set)
        
        # Bind Ctrl+Click for selection
        self.file_tree.bind("<Control-Button-1>", self.on_ctrl_click)

        # Buttons
        btns = ttk.Frame(main)
        btns.grid(row=5, column=0, pady=10)
        ttk.Button(btns, text=" Process Selected", command=self.process_files).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(btns, text="📊 Open/Create Results", command=self.open_or_create_results_file).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(btns, text="💾 Save Paths", command=self.save_current_paths).grid(row=0, column=2, padx=(0, 10))
        ttk.Button(btns, text="❌ Close", command=self.root.destroy).grid(row=0, column=3)

        # Status + progress
        self.status_var = tk.StringVar(value="Ready - Select foam type and configure paths")
        ttk.Label(main, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W).grid(row=6, column=0, sticky=(tk.W, tk.E))
        self.progress = ttk.Progressbar(main, mode="determinate")
        self.progress.grid(row=7, column=0, sticky=(tk.W, tk.E), pady=(5, 0))

    def load_suggested_paths(self):
        sug = self.foam_manager.get_suggested_paths("DSC", self.current_foam_type)
        if not sug:
            return
        # Input (suggested path even if it doesn't exist yet)
        if sug.get("input_folder"):
            if os.path.exists(sug["input_folder"]) or not self.input_folder:
                self.input_folder = sug["input_folder"]
                self.input_folder_var.set(self.input_folder)
        # Output (suggested path even if it doesn't exist yet)
        if sug.get("output_folder"):
            if os.path.exists(sug["output_folder"]) or not self.output_folder:
                self.output_folder = sug["output_folder"]
                self.output_folder_var.set(self.output_folder)
        # Results file path suggestion
        if sug.get("results_file"):
            self.results_file_var.set(sug["results_file"])

    def on_foam_type_changed(self, new_type):
        # Save current paths before switching
        if self.output_folder or self.input_folder:
            self.save_current_paths()
        self.current_foam_type = new_type
        # Clear files table
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        self.selected_dsc_files = []
        # Reload suggestions
        self.load_suggested_paths()
        self.status_var.set(f"Switched to {new_type} - paths updated")

    def save_current_paths(self):
        data = {}
        if self.input_folder:
            data["input_folder"] = self.input_folder
        if self.output_folder:
            data["output_folder"] = self.output_folder
        if self.results_file_var.get():
            data["results_file"] = self.results_file_var.get()
        if data:
            self.foam_manager.save_module_paths("DSC", self.current_foam_type, data)
            self.status_var.set(f"Paths saved for {self.current_foam_type}")

    def on_ctrl_click(self, event):
        """Handle Ctrl+Click for file selection"""
        pass  # Let default treeview selection handle this

    def select_all_files(self):
        """Select all files in the tree"""
        for item in self.file_tree.get_children():
            self.file_tree.selection_add(item)

    def select_none_files(self):
        """Deselect all files in the tree"""
        self.file_tree.selection_remove(self.file_tree.selection())

    def get_selected_files(self):
        """Get list of files that are selected in the treeview"""
        selected = []
        selected_items = self.file_tree.selection()
        for item in selected_items:
            values = self.file_tree.item(item, "values")
            if values:
                filename = values[0]
                # Find full path
                for full_path in self.selected_dsc_files:
                    if os.path.basename(full_path) == filename:
                        selected.append(full_path)
                        break
        return selected

    def browse_input_folder(self):
        # Try suggested as initial dir
        initial = None
        sug = self.foam_manager.get_suggested_paths("DSC", self.current_foam_type)
        if sug and sug.get("input_folder"):
            initial = sug["input_folder"]
        folder = filedialog.askdirectory(title="Select DSC Files Folder", initialdir=initial)
        if folder:
            self.input_folder = folder
            self.input_folder_var.set(folder)
            self.scan_dsc_files()

    def scan_dsc_files(self):
        self.selected_dsc_files = []
        folder = self.input_folder or self.input_folder_var.get()
        if not folder:
            messagebox.showwarning("No folder", "Please select a DSC Files Folder first.")
            return
        
        txt_files = glob.glob(os.path.join(folder, "*.txt"))
        txt_files = sorted(txt_files)
        
        if not txt_files:
            messagebox.showinfo("Info", "No DSC text files found in the selected folder")
            self.update_file_tree([])
            return
        
        self.selected_dsc_files = txt_files
        self.update_file_tree(txt_files)
        self.status_var.set(f"Found {len(txt_files)} DSC files in folder")

    def update_file_tree(self, files):
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        for p in files:
            name = os.path.basename(p)
            try:
                size_kb = round(os.path.getsize(p) / 1024, 1)
            except Exception:
                size_kb = "?"
            item = self.file_tree.insert("", "end", values=(name, size_kb, "Ready"))
            # Select all by default
            self.file_tree.selection_add(item)

    def browse_output_folder(self):
        initial = None
        sug = self.foam_manager.get_suggested_paths("DSC", self.current_foam_type)
        if sug and sug.get("output_folder"):
            initial = os.path.dirname(sug["output_folder"]) or None
        folder = filedialog.askdirectory(title="Select output folder", initialdir=initial)
        if folder:
            self.output_folder = folder
            self.output_folder_var.set(folder)
            # Suggest results file path if not set
            results = os.path.join(folder, self.get_output_filename())
            self.results_file_var.set(results)
            self.status_var.set(f"Output folder set: {os.path.basename(folder)}")

    def open_or_create_results_file(self):
        target = self.results_file_var.get()
        if not target:
            # Try to get from suggested paths first
            sug = self.foam_manager.get_suggested_paths("DSC", self.current_foam_type)
            if sug and sug.get("results_file"):
                target = sug["results_file"]
            elif self.output_folder:
                # If we have output folder, construct the path
                target = os.path.join(self.output_folder, self.get_output_filename())
            else:
                messagebox.showerror("Error", "No results file path configured. Set an Output Folder first.")
                return
            self.results_file_var.set(target)
        
        os.makedirs(os.path.dirname(target), exist_ok=True)
        
        if os.path.exists(target):
            # File exists, open it directly
            try:
                os.startfile(target)
                self.status_var.set(f"Opened {os.path.basename(target)}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open results file: {e}")
        else:
            # File doesn't exist, ask to create
            response = messagebox.askyesno(
                "Create DSC Results File", 
                "No DSC results file found. Would you like to create a new one?"
            )
            
            if response:
                try:
                    # Create initial Excel file based on analysis type
                    if self.analysis_type.get() == "semicrystalline":
                        columns = [
                            'Sample', 'Mass (mg)',
                            '1st Heat Crystallinity (%)', '1st Heat Enthalpy (J/g)', '1st Heat Fusion Onset (°C)', '1st Heat Fusion Peak (°C)',
                            'Cooling Crystallinity (%)', 'Cooling Enthalpy (J/g)', 'Cooling Cryst. Onset (°C)', 'Cooling Cryst. Peak (°C)',
                            '2nd Heat Crystallinity (%)', '2nd Heat Enthalpy (J/g)', '2nd Heat Fusion Onset (°C)', '2nd Heat Fusion Peak (°C)'
                        ]
                        sheet_name = 'DSC_Results'
                    else:
                        columns = [
                            'Sample', 'Mass (mg)',
                            '1st Heat Tg (°C)', '1st Heat Δcp (J/gK)',
                            '2nd Heat Tg (°C)', '2nd Heat Δcp (J/gK)'
                        ]
                        sheet_name = 'DSC_Tg_Results'
                    
                    pd.DataFrame(columns=columns).to_excel(target, index=False, engine="openpyxl", sheet_name=sheet_name)
                    self.status_var.set(f"Created {os.path.basename(target)}")
                    
                    # Open the newly created file
                    os.startfile(target)
                    messagebox.showinfo("Success", f"New DSC results file created and opened:\\n{target}")
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create DSC results file: {str(e)}")
                    self.status_var.set("Error creating DSC results file")

    # ---------- Processing ----------
    def process_files(self):
        selected_files = self.get_selected_files()
        if not selected_files:
            messagebox.showerror("Error", "No files selected for processing. Check files in the list.")
            return
        if not self.output_folder:
            messagebox.showerror("Error", "Output folder not selected.")
            return
        
        os.makedirs(self.output_folder, exist_ok=True)

        self.progress["maximum"] = len(selected_files)
        self.progress["value"] = 0
        self.processed_data = []
        errors = 0

        for i, path in enumerate(selected_files):
            fname = os.path.basename(path)
            self.status_var.set(f"Processing {fname}…")
            self.root.update()
            
            if self.analysis_type.get() == "semicrystalline":
                ok, data = self.process_semicrystalline_file(path)
            else:
                ok, data = self.process_amorphous_file(path)
            
            if ok and data:
                # Add extra info for review
                data['filename'] = fname
                data['filepath'] = path
                self.processed_data.append(data)
                self._set_tree_status(fname, "✅ Processed")
            else:
                errors += 1
                self._set_tree_status(fname, "❌ Error")
            
            self.progress["value"] = i + 1
            self.root.update()

        if self.processed_data:
            # Save results directly (like original scripts)
            self.save_results_to_excel()
            self.status_var.set(f"Processed: {len(self.processed_data)} files saved to Excel")
            messagebox.showinfo("Success", f"DSC files processed successfully!\nProcessed {len(self.processed_data)} files\nSaved to: {self.get_output_filename()}")
        else:
            self.status_var.set("No files were successfully processed")

        self.progress["value"] = 0

    def _set_tree_status(self, filename, status):
        for item in self.file_tree.get_children():
            vals = self.file_tree.item(item)["values"]
            if vals and len(vals) > 0 and vals[0] == filename:
                self.file_tree.set(item, "status", status)
                return

    def _parse_dsc_text(self, full_text: str):
        """Extract sample name, mass, and the results section from a DSC report."""
        sample_name, mass = None, None
        sample_match = re.search(r"Sample:\s*(.*?),\s*([\d,\.]+)\s*mg", full_text)
        if sample_match:
            sample_name_raw = sample_match.group(1).replace("DSC", "").strip()
            mass = float(sample_match.group(2).replace(",", "."))
            foam_types_to_remove = self.foam_manager.get_foam_types()
            sample_name = sample_name_raw
            for foam_type in foam_types_to_remove:
                if sample_name_raw.upper().startswith(foam_type.upper()):
                    sample_name = re.sub(rf"^{re.escape(foam_type)}\s*", "", sample_name_raw, flags=re.IGNORECASE).strip()
                    break

        sections = full_text.split("Results:")
        results_body = sections[1] if len(sections) > 1 else None
        return sample_name, mass, results_body

    def process_semicrystalline_file(self, file_path):
        """Process a single DSC file for semicrystalline polymers"""
        fname = Path(file_path).name
        try:
            with open(file_path, 'r', encoding='latin-1') as f:
                full_text = f.read()

            sample_name, mass, results_body = self._parse_dsc_text(full_text)

            if not results_body:
                print(f"Warning: No 'Results:' section found in {fname}")
                return False, None
                
            pattern = re.compile(
                r"Crystallinity\s+([\d,\.]+)\s*%.*?" +
                r"normalized\s+(-?[\d,\.]+)\s*Jg\^-1.*?" +
                r"Onset\s+([\d,\.]+)\s*°C.*?" +
                r"Peak\s+(-?[\d,\.]+)\s*°C",
                re.DOTALL
            )
            matches = pattern.findall(results_body)
            
            if sample_name and matches and len(matches) == 3:
                results_values = [float(val.replace(',', '.')) for block in matches for val in block]
                
                result = {
                    'Sample': sample_name,
                    'Mass (mg)': mass,
                    '1st Heat Crystallinity (%)': results_values[0],
                    '1st Heat Enthalpy (J/g)': results_values[1],
                    '1st Heat Fusion Onset (°C)': results_values[2],
                    '1st Heat Fusion Peak (°C)': results_values[3],
                    'Cooling Crystallinity (%)': results_values[4],
                    'Cooling Enthalpy (J/g)': results_values[5],
                    'Cooling Cryst. Onset (°C)': results_values[6],
                    'Cooling Cryst. Peak (°C)': results_values[7],
                    '2nd Heat Crystallinity (%)': results_values[8],
                    '2nd Heat Enthalpy (J/g)': results_values[9],
                    '2nd Heat Fusion Onset (°C)': results_values[10],
                    '2nd Heat Fusion Peak (°C)': results_values[11]
                }
                return True, result
            else:
                print(f"Warning: Could not extract 3 blocks of results from {fname}")
                return False, None
                
        except Exception as e:
            print(f"ERROR processing {fname}: {e}")
            return False, None

    def process_amorphous_file(self, file_path):
        """Process a single DSC file for amorphous polymers"""
        fname = Path(file_path).name
        try:
            with open(file_path, 'r', encoding='latin-1') as f:
                full_text = f.read()

            sample_name, mass, results_body = self._parse_dsc_text(full_text)

            if not results_body:
                print(f"Warning: No 'Results:' section found in {fname}")
                return False, None

            pattern = re.compile(
                r"Glass Transition.*?"
                r"Midpoint ISO\s+([\d,\.]+)\s*\u00b0C.*?"
                r"Delta cp\s+([\d,\.]+)\s*Jg\^-1K\^-1",
                re.DOTALL
            )
            matches = pattern.findall(results_body)

            if sample_name and matches and len(matches) >= 1:
                results_values = [float(val.replace(',', '.')) for block in matches for val in block]
                while len(results_values) < 4:  # Fill missing 2nd cycle
                    results_values.extend([None, None])

                result = {
                    'Sample': sample_name,
                    'Mass (mg)': mass,
                    '1st Heat Tg (\u00b0C)': results_values[0],
                    '1st Heat \u0394cp (J/gK)': results_values[1],
                    '2nd Heat Tg (\u00b0C)': results_values[2] if results_values[2] is not None else None,
                    '2nd Heat \u0394cp (J/gK)': results_values[3] if results_values[3] is not None else None
                }
                return True, result
            else:
                print(f"Warning: Could not extract glass transition data from {fname}")
                return False, None

        except Exception as e:
            print(f"ERROR processing {fname}: {e}")
            return False, None


    def get_output_filename(self):
        """Generate output filename based on analysis type and foam type"""
        if not self.current_foam_type:
            print(f"Warning: DSC current_foam_type is empty or None: '{self.current_foam_type}'")
            return "DSC_Results.xlsx"  # Fallback
        
        if self.analysis_type.get() == "semicrystalline":
            filename = f"DSC_Results_{self.current_foam_type}.xlsx"
        else:
            filename = f"DSC_Results_{self.current_foam_type}_Tg.xlsx"
        
        print(f"DSC get_output_filename: '{filename}' (foam_type: '{self.current_foam_type}', analysis: '{self.analysis_type.get()}')")
        return filename

    def save_results_to_excel(self):
        """Save results to Excel file with incremental updates like original scripts"""
        # Determine output file path - prioritize results_file_var, then construct default
        output_filepath = self.results_file_var.get()
        if not output_filepath:
            output_filename = self.get_output_filename()
            output_filepath = os.path.join(self.output_folder, output_filename)
            self.results_file_var.set(output_filepath)

        # Check for existing file and processed samples
        processed_samples = set()
        df_existing = pd.DataFrame()
        if os.path.exists(output_filepath):
            print(f"Existing file found: '{os.path.basename(output_filepath)}'. Reading processed samples...")
            try:
                df_existing = pd.read_excel(output_filepath, engine="openpyxl")
                if 'Sample' in df_existing.columns:
                    processed_samples = set(df_existing['Sample'].astype(str))
                    print(f"{len(processed_samples)} samples already processed.")
            except Exception as e:
                print(f"Error reading existing file: {e}")

        # Filter new data (only samples not already processed)
        new_data = []
        for data in self.processed_data:
            sample_name = data.get('Sample', '')
            if sample_name not in processed_samples:
                # Remove extra fields not needed for Excel
                excel_row = {k: v for k, v in data.items() 
                           if k not in ['filename', 'filepath']}
                new_data.append(excel_row)
                print(f"  + Adding new sample: {sample_name}")

        if not new_data:
            print("No new samples to add. Excel file is already up to date.")
            messagebox.showinfo("Info", "No new samples to add. Excel file is already up to date.")
            return

        # Combine data and save
        if self.analysis_type.get() == "semicrystalline":
            columns = [
                'Sample', 'Mass (mg)',
                '1st Heat Crystallinity (%)', '1st Heat Enthalpy (J/g)', '1st Heat Fusion Onset (°C)', '1st Heat Fusion Peak (°C)',
                'Cooling Crystallinity (%)', 'Cooling Enthalpy (J/g)', 'Cooling Cryst. Onset (°C)', 'Cooling Cryst. Peak (°C)',
                '2nd Heat Crystallinity (%)', '2nd Heat Enthalpy (J/g)', '2nd Heat Fusion Onset (°C)', '2nd Heat Fusion Peak (°C)'
            ]
            sheet_name = 'DSC_Results'
        else:
            columns = [
                'Sample', 'Mass (mg)',
                '1st Heat Tg (°C)', '1st Heat Δcp (J/gK)',
                '2nd Heat Tg (°C)', '2nd Heat Δcp (J/gK)'
            ]
            sheet_name = 'DSC_Tg_Results'
        
        df_new = pd.DataFrame(new_data, columns=columns)
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        df_combined = df_combined.sort_values(by='Sample').reset_index(drop=True)
        
        with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
            df_combined.to_excel(writer, index=False, sheet_name=sheet_name)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Highlight important columns based on analysis type
            if self.analysis_type.get() == "semicrystalline":
                # Highlight crystallinity and fusion peak from 1st heat
                crist_col_name = '1st Heat Crystallinity (%)'
                tm_col_name = '1st Heat Fusion Peak (°C)'
                
                if crist_col_name in df_combined.columns:
                    crist_col_idx = df_combined.columns.get_loc(crist_col_name) + 1
                    for row_idx in range(2, len(df_combined) + 2):
                        worksheet.cell(row=row_idx, column=crist_col_idx).fill = yellow_fill
                
                if tm_col_name in df_combined.columns:
                    tm_col_idx = df_combined.columns.get_loc(tm_col_name) + 1
                    for row_idx in range(2, len(df_combined) + 2):
                        worksheet.cell(row=row_idx, column=tm_col_idx).fill = yellow_fill
            else:
                # Highlight Tg from 1st heat for amorphous
                tg1_col_name = '1st Heat Tg (°C)'
                if tg1_col_name in df_combined.columns:
                    tg1_col_idx = df_combined.columns.get_loc(tg1_col_name) + 1
                    for row_idx in range(2, len(df_combined) + 2):
                        worksheet.cell(row=row_idx, column=tg1_col_idx).fill = yellow_fill
                
        print(f"\n¡Success! File updated at: '{output_filepath}'")
        print(f"Added {len(new_data)} new samples. Total samples: {len(df_combined)}.")


def main():
    root = tk.Tk()
    app = DSCModule(root)
    root.mainloop()


if __name__ == "__main__":
    # For direct execution, use absolute imports
    import sys
    import os
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    
    # Re-import with absolute paths when running directly
    try:
        from modules.foam_type_manager import FoamTypeManager, FoamTypeSelector
        # Update the class to use the absolute imports
        globals()['FoamTypeManager'] = FoamTypeManager
        globals()['FoamTypeSelector'] = FoamTypeSelector
        main()
    except ImportError:
        print("Error: This module should be run from the main application.")
        print("Please run main_gui.py instead.")
        input("Press Enter to exit...")
