import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
import win32com.client as win32
from .foam_type_manager import FoamTypeManager, FoamTypeSelector
import re

class PDRModule:
    def __init__(self, root):
        self.root = root
        self.root.title("Pressure Drop Rate Analysis")
        self.root.geometry("800x750")
        
        # Initialize foam type manager
        self.foam_manager = FoamTypeManager()
        
        # Variables for paths
        self.input_folder = None
        self.output_folder = None
        self.registros_file = None
        self.selected_files = []
        
        # Current foam type
        self.current_foam_type = self.foam_manager.get_current_foam_type()
        
        self.create_widgets()
    
    def create_widgets(self):
        """Create the GUI widgets"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)

        # Title
        title_label = ttk.Label(main_frame, text="Pressure Drop Rate Analysis", font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, pady=(0, 20))

        # Foam type selector
        self.foam_selector = FoamTypeSelector(main_frame, self.foam_manager, self.on_foam_type_changed)

        # Setup frame
        setup_frame = ttk.LabelFrame(main_frame, text="Setup Files and Folders", padding="10")
        setup_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        setup_frame.columnconfigure(1, weight=1)

        # Input folder
        ttk.Label(setup_frame, text="Input CSV Folder:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.input_folder_var = tk.StringVar()
        ttk.Entry(setup_frame, textvariable=self.input_folder_var, state='readonly', width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=(0, 5))
        ttk.Button(setup_frame, text="Browse", command=self.browse_input_folder).grid(row=0, column=2, pady=(0, 5))

        # Output folder
        ttk.Label(setup_frame, text="Output Folder:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        self.output_folder_var = tk.StringVar()
        ttk.Entry(setup_frame, textvariable=self.output_folder_var, state='readonly', width=60).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=(0, 5))
        ttk.Button(setup_frame, text="Browse", command=self.browse_output_folder).grid(row=1, column=2, pady=(0, 5))

        # Registros file (Optional)
        ttk.Label(setup_frame, text="PDR Results File:").grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        self.registros_file_var = tk.StringVar()
        ttk.Entry(setup_frame, textvariable=self.registros_file_var, state='readonly', width=60).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=(0, 5))
        ttk.Button(setup_frame, text="📊 Open/Create", command=self.open_or_create_registros_file).grid(row=2, column=2, pady=(0, 5))

        # Load suggested paths (after variables exist)
        self.load_suggested_paths()

        # Scan button
        ttk.Button(setup_frame, text="🔍 Scan CSV Files", command=self.scan_csv_files).grid(row=3, column=0, pady=(10, 5))

        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="Select CSV Files to Process", padding="10")
        file_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        file_frame.rowconfigure(0, weight=1)

        # Treeview for file selection
        columns = ('filename', 'size', 'status')
        self.file_tree = ttk.Treeview(file_frame, columns=columns, show='headings', height=12)

        # Define headings
        self.file_tree.heading('filename', text='CSV File Name')
        self.file_tree.heading('size', text='Size (KB)')
        self.file_tree.heading('status', text='Status')

        # Column widths
        self.file_tree.column('filename', width=400)
        self.file_tree.column('size', width=100, anchor='center')
        self.file_tree.column('status', width=150, anchor='center')

        self.file_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Scrollbars for treeview
        v_scrollbar = ttk.Scrollbar(file_frame, orient=tk.VERTICAL, command=self.file_tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.file_tree.configure(yscrollcommand=v_scrollbar.set)

        # File selection buttons
        selection_frame = ttk.Frame(file_frame)
        selection_frame.grid(row=1, column=0, pady=(5, 0))

        ttk.Button(selection_frame, text="Select All", command=self.select_all_files).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(selection_frame, text="Select None", command=self.select_no_files).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(selection_frame, text="Toggle Selection", command=self.toggle_selection).grid(row=0, column=2)

        # Process buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, pady=10)

        ttk.Button(button_frame, text="🚀 Process Selected Files", command=self.process_files).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(button_frame, text="📊 Open/Create Registros File", command=self.open_or_create_registros_file).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(button_frame, text="💾 Save Paths", command=self.save_current_paths).grid(row=0, column=2, padx=(0, 10))
        ttk.Button(button_frame, text="❌ Close", command=self.root.destroy).grid(row=0, column=3)

        # Status bar
        self.status_var = tk.StringVar(value="Ready - Select foam type and configure paths")
        ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W).grid(row=5, column=0, sticky=(tk.W, tk.E))

        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='determinate')
        self.progress.grid(row=6, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
    
    def load_suggested_paths(self):
        """Load suggested paths for current foam type"""
        suggested = self.foam_manager.get_suggested_paths("PDR", self.current_foam_type)
        
        # Set input folder (suggested path even if it doesn't exist yet)
        if "input_folder" in suggested:
            if os.path.exists(suggested["input_folder"]) or not self.input_folder:
                self.input_folder = suggested["input_folder"]
                self.input_folder_var.set(suggested["input_folder"])
        
        # Set output folder (suggested path even if it doesn't exist yet)
        if "output_folder" in suggested:
            if os.path.exists(suggested["output_folder"]) or not self.output_folder:
                self.output_folder = suggested["output_folder"]
                self.output_folder_var.set(suggested["output_folder"])
        
        # Set registros file (suggested path even if it doesn't exist yet)
        if "registros_file" in suggested:
            if os.path.exists(suggested["registros_file"]) or not self.registros_file:
                self.registros_file = suggested["registros_file"]
                self.registros_file_var.set(suggested["registros_file"])
        
        if suggested and hasattr(self, 'status_var'):
            self.status_var.set(f"Default paths loaded for {self.current_foam_type}")

    def _extract_label_from_filename(self, filename):
        """Extract label from filename.

        If filename has form "foamtype label.csv" and self.current_foam_type matches
        the foamtype prefix, returns only "label". Otherwise returns base filename
        without extension.
        Supports separators: space, underscore, hyphen, dot.
        """
        base = os.path.splitext(filename)[0]
        foam = (self.current_foam_type or "").strip()
        if foam:
            # Match foam at start followed by separator(s) then the label
            pattern = r'^' + re.escape(foam) + r'[\s_\-\.]+(.+)$'
            m = re.match(pattern, base, flags=re.I)
            if m:
                return m.group(1).strip()
        return base
    
    def on_foam_type_changed(self, new_foam_type):
        """Handle foam type change"""
        # Save current paths before changing
        if self.input_folder or self.output_folder or self.registros_file:
            self.save_current_paths()
        
        # Update current foam type
        self.current_foam_type = new_foam_type
        
        # Load new paths
        self.load_suggested_paths()
        
        # Clear file list
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
            
        self.status_var.set(f"Switched to {new_foam_type} - paths updated")
    
    def save_current_paths(self):
        """Save current paths for the current foam type"""
        paths = {}
        if self.input_folder:
            paths["input_folder"] = self.input_folder
        if self.output_folder:
            paths["output_folder"] = self.output_folder
        if self.registros_file:
            paths["registros_file"] = self.registros_file
            
        if paths:
            self.foam_manager.save_module_paths("PDR", self.current_foam_type, paths)
            self.status_var.set(f"Paths saved for {self.current_foam_type}")
    
    def browse_input_folder(self):
        """Browse for input CSV folder"""
        # Start from suggested path
        initial_dir = None
        suggested = self.foam_manager.get_suggested_paths("PDR", self.current_foam_type)
        if "input_folder" in suggested:
            initial_dir = os.path.dirname(suggested["input_folder"])
        
        folder = filedialog.askdirectory(
            title=f"Select input CSV folder for {self.current_foam_type}",
            initialdir=initial_dir
        )
        if folder:
            self.input_folder = folder
            self.input_folder_var.set(folder)
            self.status_var.set(f"Input folder set: {os.path.basename(folder)}")
            
    def browse_output_folder(self):
        """Browse for output folder"""
        # Start from suggested path
        initial_dir = None
        suggested = self.foam_manager.get_suggested_paths("PDR", self.current_foam_type)
        if "output_folder" in suggested:
            initial_dir = os.path.dirname(suggested["output_folder"])
        
        folder = filedialog.askdirectory(
            title=f"Select output folder for {self.current_foam_type}",
            initialdir=initial_dir
        )
        if folder:
            self.output_folder = folder
            self.output_folder_var.set(folder)
            self.status_var.set(f"Output folder set: {os.path.basename(folder)}")
    
    def get_output_filename(self):
        """Generate dynamic filename based on foam type"""
        return f"PDR_Results_{self.current_foam_type}.xlsx"
            
    def browse_registros_file(self):
        """Browse for PDR results file"""
        # Start from suggested path
        initial_dir = None
        suggested = self.foam_manager.get_suggested_paths("PDR", self.current_foam_type)
        if "registros_file" in suggested:
            initial_dir = os.path.dirname(suggested["registros_file"])
        
        filename = filedialog.askopenfilename(
            title="Select existing PDR results file (Optional)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=self.get_output_filename(),
            initialdir=initial_dir
        )
        if filename:
            self.registros_file = filename
            self.registros_file_var.set(filename)
            self.status_var.set(f"Registros file set: {os.path.basename(filename)}")
            
    def scan_csv_files(self):
        """Scan input folder for CSV files"""
        if not self.input_folder:
            messagebox.showerror("Error", "Please select input folder first")
            return
            
        # Clear existing items
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
            
        try:
            csv_files = [f for f in os.listdir(self.input_folder) if f.endswith('.csv')]
            
            if not csv_files:
                messagebox.showwarning("Warning", "No CSV files found in the selected folder")
                return
                
            for csv_file in csv_files:
                filepath = os.path.join(self.input_folder, csv_file)
                try:
                    size_kb = round(os.path.getsize(filepath) / 1024, 1)
                    self.file_tree.insert('', 'end', values=(csv_file, f"{size_kb}", "Ready"))
                except:
                    self.file_tree.insert('', 'end', values=(csv_file, "Error", "Error reading file"))
                    
            self.status_var.set(f"Found {len(csv_files)} CSV files")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error scanning folder: {str(e)}")
            
    def select_all_files(self):
        """Select all files in the tree"""
        for item in self.file_tree.get_children():
            self.file_tree.selection_add(item)
            
    def select_no_files(self):
        """Deselect all files"""
        self.file_tree.selection_remove(self.file_tree.selection())
        
    def toggle_selection(self):
        """Toggle selection of current item"""
        current_selection = self.file_tree.selection()
        all_items = self.file_tree.get_children()
        
        for item in all_items:
            if item in current_selection:
                self.file_tree.selection_remove(item)
            else:
                self.file_tree.selection_add(item)
                
    def process_files(self):
        """Process selected CSV files"""
        # Validation
        if not self.input_folder or not self.output_folder:
            messagebox.showerror("Error", "Please select both input and output folders")
            return
            
        # Check if registros file is needed
        if not self.registros_file:
            # Ask user if they want to create a new registros file
            result = messagebox.askyesnocancel(
                "No Registros File", 
                "No registros file selected. Would you like to:\\n\\n" +
                "• YES: Create a new registros file\\n" +
                "• NO: Process without updating registros\\n" +
                "• CANCEL: Go back to select a file"
            )
            
            if result is None:  # Cancel
                return
            elif result:  # Yes - create new
                self.create_new_registros()
                if not self.registros_file:  # User cancelled creation
                    return
            # If No - continue without registros (self.registros_file stays None)
            
        # Validate existing registros file if provided
        if self.registros_file and not os.path.exists(self.registros_file):
            messagebox.showerror("Error", "Registros file does not exist. Please select a valid file or create a new one.")
            return
            
        selected_items = self.file_tree.selection()
        if not selected_items:
            messagebox.showerror("Error", "Please select at least one CSV file to process")
            return
            
        # Get selected filenames
        selected_files = []
        for item in selected_items:
            filename = self.file_tree.item(item)['values'][0]
            selected_files.append(filename)
            
        # Create output folder if it doesn't exist
        os.makedirs(self.output_folder, exist_ok=True)
        
        # Process files
        self.progress['maximum'] = len(selected_files)
        self.progress['value'] = 0
        
        processed_count = 0
        error_count = 0
        
        for i, filename in enumerate(selected_files):
            self.status_var.set(f"Processing {filename}...")
            self.root.update()
            
            try:
                success = self.process_single_csv(filename)
                if success:
                    processed_count += 1
                    # Update tree status
                    for item in self.file_tree.get_children():
                        if self.file_tree.item(item)['values'][0] == filename:
                            self.file_tree.set(item, 'status', '✅ Processed')
                            break
                else:
                    error_count += 1
                    # Update tree status
                    for item in self.file_tree.get_children():
                        if self.file_tree.item(item)['values'][0] == filename:
                            self.file_tree.set(item, 'status', '❌ Error')
                            break
                            
            except Exception as e:
                error_count += 1
                print(f"Error processing {filename}: {e}")
                # Update tree status
                for item in self.file_tree.get_children():
                    if self.file_tree.item(item)['values'][0] == filename:
                        self.file_tree.set(item, 'status', f'❌ Error: {str(e)[:20]}')
                        break
            
            self.progress['value'] = i + 1
            self.root.update()
        
        # Show completion message
        self.progress['value'] = 0
        message = f"Processing complete!\n\nProcessed: {processed_count}\nErrors: {error_count}"
        if error_count == 0:
            messagebox.showinfo("Success", message)
        else:
            messagebox.showwarning("Completed with errors", message)
            
        self.status_var.set(f"Processing complete: {processed_count} processed, {error_count} errors")
        
    def process_single_csv(self, filename):
        """Process a single CSV file using the original PDR logic"""
        input_path = os.path.join(self.input_folder, filename)
        
        try:
            # Read CSV file
            df = pd.read_csv(input_path)
            print(f"Archivo {filename} leído correctamente.")
        except Exception as e:
            print(f"Error al leer {filename}: {e}")
            return False
            
        # Apply original processing logic
        df.columns = ['Time', 'T1 (ºC)', 'T2 (ºC)', 'P (bar)']
        df['P (bar)'] = pd.to_numeric(df['P (bar)'].str.replace(',', '.'), errors='coerce')

        df['P (MPa)'] = df['P (bar)'] / 10
        df['ΔP (MPa)'] = df['P (MPa)'].diff().abs()
        df['PDR (MPa/s)'] = df['ΔP (MPa)'] / 0.06

        try:
            first_high_delta_p_index = df[df['ΔP (MPa)'] > 1].index[0]
        except IndexError:
            print(f"No se encontró un valor mayor a 1 en 'ΔP (MPa)' en {filename}. Continuando con el siguiente archivo.")
            return False

        start_index = max(0, first_high_delta_p_index - 4)
        end_index = min(len(df), first_high_delta_p_index + 12) 
        
        df_filtered = df.iloc[start_index:end_index].copy()
        
        df['Time (s)'] = None
        df.loc[start_index:end_index - 1, 'Time (s)'] = [0.06 * i for i in range(len(df_filtered))]

        # Save processed file
        output_path = os.path.join(self.output_folder, f"{filename.replace('.csv', '')} procesado.xlsx")
        
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Datos Procesados', index=False)
            worksheet = writer.sheets['Datos Procesados']
            workbook = writer.book
            
            bold_format = workbook.add_format({'bold': True})
            bold_subscript_format = workbook.add_format({'bold': True, 'font_script': 2})

            # Highlighting logic (from original code)
            yellow_speeds_indices = []
            blue_speeds_indices = []
            tolerance = 0.10

            if first_high_delta_p_index < len(df) and pd.notna(df.loc[first_high_delta_p_index, 'PDR (MPa/s)']):
                last_pdr = df.loc[first_high_delta_p_index, 'PDR (MPa/s)']
                for i in range(first_high_delta_p_index, len(df)):
                    current_pdr = df.loc[i, 'PDR (MPa/s)']
                    if pd.isna(current_pdr):
                        break
                    
                    is_similar = False
                    if last_pdr > 0:
                        if abs(current_pdr - last_pdr) / last_pdr <= tolerance:
                            is_similar = True
                    elif current_pdr == last_pdr:
                        is_similar = True

                    if is_similar:
                        yellow_speeds_indices.append(i)
                        last_pdr = current_pdr
                    else:
                        break
            
            if yellow_speeds_indices:
                start_blue_index = yellow_speeds_indices[-1] + 1
                if start_blue_index < len(df) and pd.notna(df.loc[start_blue_index, 'PDR (MPa/s)']):
                    last_pdr = df.loc[start_blue_index, 'PDR (MPa/s)']
                    for i in range(start_blue_index, len(df)):
                        current_pdr = df.loc[i, 'PDR (MPa/s)']
                        if pd.isna(current_pdr):
                            break

                        is_similar = False
                        if last_pdr > 0:
                            if abs(current_pdr - last_pdr) / last_pdr <= tolerance:
                                is_similar = True
                        elif current_pdr == last_pdr:
                            is_similar = True

                        if is_similar:
                            blue_speeds_indices.append(i)
                            last_pdr = current_pdr
                        else:
                            break

            # Apply formatting
            yellow_format = workbook.add_format({'bg_color': '#FFFF00'})
            for idx in yellow_speeds_indices:
                worksheet.write(idx + 1, df.columns.get_loc('PDR (MPa/s)'), df.loc[idx, 'PDR (MPa/s)'], yellow_format)

            blue_format = workbook.add_format({'bg_color': '#00BFFF'})
            for idx in blue_speeds_indices:
                worksheet.write(idx + 1, df.columns.get_loc('PDR (MPa/s)'), df.loc[idx, 'PDR (MPa/s)'], blue_format)

            if yellow_speeds_indices:
                initial_pressure_index = first_high_delta_p_index - 1
                final_pressure_index = yellow_speeds_indices[-1]

                if initial_pressure_index >= 0:
                    pink_format = workbook.add_format({'bg_color': '#FFC0CB'})
                    
                    worksheet.write(initial_pressure_index + 1, df.columns.get_loc('P (MPa)'), df.loc[initial_pressure_index, 'P (MPa)'], pink_format)
                    worksheet.write(initial_pressure_index + 1, df.columns.get_loc('Time (s)'), df.loc[initial_pressure_index, 'Time (s)'], pink_format)
                    worksheet.write(final_pressure_index + 1, df.columns.get_loc('P (MPa)'), df.loc[final_pressure_index, 'P (MPa)'], pink_format)
                    worksheet.write(final_pressure_index + 1, df.columns.get_loc('Time (s)'), df.loc[final_pressure_index, 'Time (s)'], pink_format)

                    p_inicial_cell = f"E{initial_pressure_index + 2}"
                    t_inicial_cell = f"H{initial_pressure_index + 2}"
                    p_final_cell = f"E{final_pressure_index + 2}"
                    t_final_cell = f"H{final_pressure_index + 2}"
                    
                    worksheet.write_rich_string('U1', bold_format, 'P', bold_subscript_format, 'i', bold_format, ' (MPa)')
                    worksheet.write_rich_string('V1', bold_format, 'P', bold_subscript_format, 'f', bold_format, ' (MPa)')
                    worksheet.write('W1', 'PDR (MPa/s)', bold_format)
                    
                    worksheet.write_formula('U2', f"={p_inicial_cell}")
                    worksheet.write_formula('V2', f"={p_final_cell}")
                    formula = f"=ABS(U2-V2)/({t_final_cell}-{t_inicial_cell})"
                    worksheet.write_formula('W2', formula)
                else:
                    worksheet.write('U2', 'Error: No initial point')

            # Add charts (from original code)
            self.add_charts_to_worksheet(workbook, worksheet, df, start_index, end_index)

        print(f"Archivo guardado: {output_path}")

        # Update registros file if provided
        if self.registros_file:
            try:
                self.update_registros_file(filename, output_path)
                return True
            except Exception as e:
                print(f"Error al actualizar el archivo de registros: {e}")
                return False
        else:
            print("No registros file provided - skipping registros update")
            return True
            
    def add_charts_to_worksheet(self, workbook, worksheet, df, start_index, end_index):
        """Add charts to worksheet (from original code)"""
        # Chart 1: P (MPa) vs time
        chart1 = workbook.add_chart({'type': 'scatter', 'subtype': 'straight'})
        chart1.add_series({
            'name': 'P (MPa)',
            'categories': ['Datos Procesados', start_index + 1, df.columns.get_loc('Time (s)'), end_index, df.columns.get_loc('Time (s)')],
            'values': ['Datos Procesados', start_index + 1, df.columns.get_loc('P (MPa)'), end_index, df.columns.get_loc('P (MPa)')],
            'line': {'color': '#004C99', 'width': 1.5},
            'marker': {'type': 'none'},
        })
        chart1.set_title({'name': 'P (MPa) vs time (s)'})
        chart1.set_x_axis({'name': 'Time (s)', 'min': 0, 'max': 0.54, 'major_gridlines': {'visible': False}})
        chart1.set_y_axis({'name': 'P (MPa)', 'major_gridlines': {'visible': False}})
        worksheet.insert_chart('J5', chart1)

        # Chart 2: PDR vs time
        chart2 = workbook.add_chart({'type': 'scatter', 'subtype': 'straight'})
        chart2.add_series({
            'name': 'PDR (MPa/s)',
            'categories': ['Datos Procesados', start_index + 1, df.columns.get_loc('Time (s)'), end_index, df.columns.get_loc('Time (s)')],
            'values': ['Datos Procesados', start_index + 1, df.columns.get_loc('PDR (MPa/s)'), end_index, df.columns.get_loc('PDR (MPa/s)')],
            'line': {'color': '#CC0000', 'width': 1.5},
            'marker': {'type': 'none'},
        })
        chart2.set_title({'name': 'PDR (MPa/s) vs time (s)'})
        chart2.set_x_axis({'name': 'Time (s)', 'min': 0, 'max': 0.54, 'major_gridlines': {'visible': False}})
        chart2.set_y_axis({'name': 'PDR (MPa/s)', 'major_gridlines': {'visible': False}})
        worksheet.insert_chart('J20', chart2)

        # Chart 3: Combined chart with dual axis
        chart3 = workbook.add_chart({'type': 'scatter', 'subtype': 'straight'})

        chart3.add_series({
            'name':       'P (MPa)',
            'categories': ['Datos Procesados', start_index + 1, df.columns.get_loc('Time (s)'), end_index, df.columns.get_loc('Time (s)')],
            'values':     ['Datos Procesados', start_index + 1, df.columns.get_loc('P (MPa)'), end_index, df.columns.get_loc('P (MPa)')],
            'line':       {'color': '#004C99', 'width': 1.5},
            'marker':     {'type': 'none'},
        })

        chart3.add_series({
            'name':       'PDR (MPa/s)',
            'categories': ['Datos Procesados', start_index + 1, df.columns.get_loc('Time (s)'), end_index, df.columns.get_loc('Time (s)')],
            'values':     ['Datos Procesados', start_index + 1, df.columns.get_loc('PDR (MPa/s)'), end_index, df.columns.get_loc('PDR (MPa/s)')],
            'line':       {'color': '#CC0000', 'width': 1.5},
            'marker':     {'type': 'none'},
            'y2_axis':    True,
        })

        max_pdr = df['PDR (MPa/s)'].max()
        y2_axis_max = max_pdr * 1.20

        # Add reference lines and labels
        worksheet.write('M1', 'line_x_1')
        worksheet.write('N1', 'line_y_1')
        worksheet.write('M2', 0.18)
        worksheet.write('N2', 0)
        worksheet.write('M3', 0.18)
        worksheet.write('N3', y2_axis_max)
        worksheet.write('O1', 'line_x_2')
        worksheet.write('P1', 'line_y_2')
        worksheet.write('O2', 0.36)
        worksheet.write('P2', 0)
        worksheet.write('O3', 0.36)
        worksheet.write('P3', y2_axis_max)

        worksheet.write('Q1', 'label_x')
        worksheet.write('R1', 'label_y')
        worksheet.write('S1', 'label_text')
        worksheet.write('Q2', 0.18)
        worksheet.write('R2', y2_axis_max)
        worksheet.write('S2', '45 degrees')
        worksheet.write('Q3', 0.36)
        worksheet.write('R3', y2_axis_max)
        worksheet.write('S3', '90 degrees')

        # Add reference line series
        chart3.add_series({
            'categories': "='Datos Procesados'!$M$2:$M$3",
            'values':     "='Datos Procesados'!$N$2:$N$3",
            'line':       {'color': '#595959', 'width': 0.75},
            'marker':     {'type': 'none'},
            'y2_axis':    True,
        })
        
        chart3.add_series({
            'categories': "='Datos Procesados'!$O$2:$O$3",
            'values':     "='Datos Procesados'!$P$2:$P$3",
            'line':       {'color': '#595959', 'width': 0.75},
            'marker':     {'type': 'none'},
            'y2_axis':    True,
        })

        # Add labels
        chart3.add_series({
            'categories': "='Datos Procesados'!$Q$2",
            'values':     "='Datos Procesados'!$R$2",
            'marker':     {'type': 'none'},
            'y2_axis':    True,
            'data_labels': {'custom': [{'value': "='Datos Procesados'!$S$2"}], 'position': 'above', 'font': {'size': 9}},
        })

        chart3.add_series({
            'categories': "='Datos Procesados'!$Q$3",
            'values':     "='Datos Procesados'!$R$3",
            'marker':     {'type': 'none'},
            'y2_axis':    True,
            'data_labels': {'custom': [{'value': "='Datos Procesados'!$S$3"}], 'position': 'above', 'font': {'size': 9}},
        })

        chart3.set_title({'name': ' '})
        chart3.set_x_axis({'name': 'Time (s)', 'min': 0, 'max': 0.54, 'major_gridlines': {'visible': False}})
        chart3.set_y_axis({'name': 'P (MPa)', 'major_gridlines': {'visible': False}})
        chart3.set_y2_axis({'name': 'PDR (MPa/s)', 'max': y2_axis_max, 'min': 0, 'major_gridlines': {'visible': False}})
        chart3.set_legend({'position': 'right', 'delete_series': [2, 3, 4, 5]})
        worksheet.insert_chart('J35', chart3)
        
    def update_registros_file(self, filename, output_path):
        """Update registros file with new data (from original code)"""
        # Try to use Excel COM automation, but be tolerant: if any COM operation
        # fails (including setting Excel.Visible), fall back to an openpyxl-based
        # approach that writes the key formulas (but won't copy the chart).
        excel = None
        try:
            try:
                excel = win32.Dispatch("Excel.Application")
            except Exception as e:
                # COM not available: fall back
                print(f"Excel COM not available, using openpyxl fallback: {e}")
                return self._fallback_update_registros_file(filename, output_path)

            # Setting Visible can fail in restricted environments; ignore errors
            try:
                excel.Visible = False
            except Exception:
                # Non-fatal: continue without changing visibility
                pass

            wb_registros = excel.Workbooks.Open(self.registros_file)
            wb_calculado = excel.Workbooks.Open(output_path)

            sheet_registros = wb_registros.Sheets["Registros"]
            sheet_calculado = wb_calculado.Sheets["Datos Procesados"]

            # Determine the last filled row in column A (header at row 1)
            last_filled = sheet_registros.Cells(sheet_registros.Rows.Count, 1).End(3).Row

            # Incremental behavior: append only if not already present by base filename
            try:
                existing_names = set()
                if last_filled >= 2:
                    rng = sheet_registros.Range(f"A2:A{last_filled}").Value
                    if isinstance(rng, tuple):
                        for item in rng:
                            val = item[0] if isinstance(item, tuple) else item
                            if val is not None:
                                existing_names.add(str(val).strip().lower())
                    else:
                        if rng is not None:
                            existing_names.add(str(rng).strip().lower())
                # Use extracted label for duplicate detection (so 'foamtype label' -> 'label')
                base_name = self._extract_label_from_filename(filename).strip().lower()
                if base_name in existing_names:
                    # Duplicate found: close and skip appending
                    wb_calculado.Close(SaveChanges=False)
                    wb_registros.Close(SaveChanges=True)
                    try:
                        excel.Quit()
                    except Exception:
                        pass
                    print(f"Skipped duplicate entry for '{base_name}' in Registros.")
                    return True
            except Exception:
                # If any issue reading existing names, continue with append
                pass

            # Next row to write
            last_row = last_filled + 1
            try:
                sheet_registros.Rows(last_row).RowHeight = 218
                sheet_registros.Columns("E").ColumnWidth = 65
            except Exception:
                # Non-fatal formatting error
                pass

            # Try to copy chart; if it fails, continue and only write formulas
            try:
                chart_object = sheet_calculado.ChartObjects(3)
                chart_object.Copy()

                cell_for_graph = sheet_registros.Cells(last_row, 5)
                sheet_registros.Paste(cell_for_graph)

                pasted_chart = sheet_registros.ChartObjects(sheet_registros.ChartObjects().Count)
                pasted_chart.Width = 360
                pasted_chart.Height = 217.5
            except Exception:
                # Chart copy failed: continue without chart
                print("Warning: Could not copy chart to Registros (continuing without chart)")

            # Add data (filename and formulas referencing the processed workbook)
            try:
                # Write label: if filename is 'foamtype label.csv' and current foam type matches,
                # write only 'label'
                label = self._extract_label_from_filename(filename)
                sheet_registros.Cells(last_row, 1).Value = label

                archivo_procesado_sin_ruta = os.path.basename(output_path)
                formula_u2 = f"='[{archivo_procesado_sin_ruta}]Datos Procesados'!$U$2"
                formula_v2 = f"='[{archivo_procesado_sin_ruta}]Datos Procesados'!$V$2"
                formula_w2 = f"='[{archivo_procesado_sin_ruta}]Datos Procesados'!$W$2"

                sheet_registros.Cells(last_row, 2).Formula = formula_u2
                sheet_registros.Cells(last_row, 3).Formula = formula_v2
                sheet_registros.Cells(last_row, 4).Formula = formula_w2

                try:
                    sheet_registros.Columns("A:D").AutoFit()
                except Exception:
                    pass
            except Exception:
                print("Warning: Could not write formulas into Registros via COM. Trying fallback.")
                # Close COM workbooks before fallback
                try:
                    wb_calculado.Close(SaveChanges=False)
                except Exception:
                    pass
                try:
                    wb_registros.Close(SaveChanges=True)
                except Exception:
                    pass
                try:
                    excel.Quit()
                except Exception:
                    pass
                return self._fallback_update_registros_file(filename, output_path)

            # Save and close
            wb_registros.Save()
            try:
                wb_calculado.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                wb_registros.Close(SaveChanges=True)
            except Exception:
                pass
            try:
                excel.Quit()
            except Exception:
                pass
            print(f"Gráfico copiado al archivo de registros en la fila {last_row}.")
            return True

        except Exception as e:
            # Any unexpected COM error: try fallback instead of raising
            print(f"Error al actualizar el archivo de registros via COM: {e}\nFalling back to openpyxl method.")
            try:
                if excel is not None:
                    try:
                        excel.Quit()
                    except Exception:
                        pass
            except Exception:
                pass
            return self._fallback_update_registros_file(filename, output_path)

    def _fallback_update_registros_file(self, filename, output_path):
        """Fallback method that uses openpyxl to append a row with formulas.
        This will NOT copy charts but will add formulas referencing the processed file
        so Excel can evaluate them when the user opens the workbook.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.registros_file)
            if "Registros" not in wb.sheetnames:
                ws = wb.active
                ws.title = "Registros"
            else:
                ws = wb["Registros"]

            # Find last filled row in column A
            # Find last filled row in column A and collect existing names for dedup
            last_row = 1
            existing_names = set()
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if val is not None:
                    last_row = row
                    try:
                        existing_names.add(str(val).strip().lower())
                    except Exception:
                        pass

            next_row = last_row + 1

            # Set row height / column width where possible
            try:
                ws.row_dimensions[next_row].height = 218
            except Exception:
                pass
            try:
                ws.column_dimensions['E'].width = 65
            except Exception:
                pass

            # Determine label and skip duplicates
            label = self._extract_label_from_filename(filename)
            if label.strip().lower() in existing_names:
                print(f"Skipped duplicate entry for '{label}' in Registros (fallback).")
                wb.save(self.registros_file)
                return True

            # Write label
            ws.cell(row=next_row, column=1, value=label)

            archivo_procesado_sin_ruta = os.path.basename(output_path)
            formula_u2 = f"='[{archivo_procesado_sin_ruta}]Datos Procesados'!$U$2"
            formula_v2 = f"='[{archivo_procesado_sin_ruta}]Datos Procesados'!$V$2"
            formula_w2 = f"='[{archivo_procesado_sin_ruta}]Datos Procesados'!$W$2"

            ws.cell(row=next_row, column=2, value=formula_u2)
            ws.cell(row=next_row, column=3, value=formula_v2)
            ws.cell(row=next_row, column=4, value=formula_w2)

            # Save workbook
            wb.save(self.registros_file)
            print(f"Registros actualizado (fallback) en la fila {next_row} (sin gráfico).")
            return True

        except Exception as e:
            print(f"Error en fallback al actualizar registros: {e}")
            return False
            
    def open_or_create_registros_file(self):
        """Open existing or create new registros file with smart logic"""
        if self.registros_file and os.path.exists(self.registros_file):
            # File exists, open it directly
            try:
                os.startfile(self.registros_file)
                self.status_var.set("Registros file opened in Excel")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open registros file: {str(e)}")
        else:
            # File doesn't exist or not selected, ask to create
            if not self.output_folder:
                messagebox.showerror("Error", "Please select output folder first")
                return
                
            response = messagebox.askyesno(
                "Create PDR Results File", 
                "No PDR results file found. Would you like to create a new one?"
            )
            
            if response:
                filename = filedialog.asksaveasfilename(
                    title="Create new PDR results file",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    initialfile=self.get_output_filename(),
                    initialdir=self.output_folder
                )
                
                if filename:
                    try:
                        # Create new Excel file with headers
                        import pandas as pd
                        from openpyxl import Workbook
                        
                        # Create workbook and worksheet
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Registros"
                        
                        # Add headers
                        headers = ["Filename", "Pi (MPa)", "Pf (MPa)", "PDR (MPa/s)", "Chart"]
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=header)
                        
                        # Set column widths
                        ws.column_dimensions['A'].width = 25  # Filename
                        ws.column_dimensions['B'].width = 12  # Pi
                        ws.column_dimensions['C'].width = 12  # Pf  
                        ws.column_dimensions['D'].width = 15  # PDR
                        ws.column_dimensions['E'].width = 65  # Chart
                        
                        # Save the file
                        wb.save(filename)
                        
                        # Set as current registros file
                        self.registros_file = filename
                        self.registros_file_var.set(filename)
                        self.status_var.set(f"New PDR results file created: {os.path.basename(filename)}")
                        
                        # Open the newly created file
                        os.startfile(filename)
                        messagebox.showinfo("Success", f"New PDR results file created and opened:\\n{filename}")
                        
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to create PDR results file: {str(e)}")
                        self.status_var.set("Error creating PDR results file")

    def open_registros_file(self):
        """Open the registros file in Excel"""
        if not self.registros_file:
            messagebox.showwarning("Warning", "No registros file selected. Please browse for a file or create a new one.")
            return
            
        if not os.path.exists(self.registros_file):
            messagebox.showerror("Error", "Registros file not found. Please check the file path.")
            return
            
        try:
            os.startfile(self.registros_file)
            self.status_var.set("Registros file opened in Excel")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open registros file: {str(e)}")


def main():
    root = tk.Tk()
    app = PDRModule(root)
    root.mainloop()

if __name__ == "__main__":
    main()
