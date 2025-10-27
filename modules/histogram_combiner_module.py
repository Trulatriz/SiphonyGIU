"""
Histogram Combiner Module for PressTech
Combines multiple histogram Excel files from SEM analysis into a single results file
"""

import os
import re
import glob
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl


def normalize_label(label):
    if pd.isna(label):
        return ""
    return str(label).strip().upper()

def canonical_histogram_key(value):
    """Canonical form of a histogram sheet name (case-insensitive, ignores spaces/underscores)."""
    if value is None:
        return ""
    s = str(value).strip().upper()
    s = re.sub(r"HISTOGRAM[\s_-]*", "HISTOGRAM", s)
    return re.sub(r"[ _]", "", s)

def canonical_label_key(value):
    """Canonical form for comparing labels (preserve hyphen but ignore spaces/underscores)."""
    if value is None:
        return ""
    return re.sub(r"[ _]", "", str(value).strip().upper())

def extract_label_tokens(value):
    """Return all numeric tokens (with optional single hyphen) found in the value."""
    if value is None:
        return []
    return re.findall(r"\d+(?:-\d+)?", str(value))

def primary_label_token(value):
    tokens = extract_label_tokens(value)
    return tokens[0] if tokens else ""


def excel_cell_to_indices(cell_ref: str):
    m_col = re.match(r'([A-Z]+)', cell_ref)
    m_row = re.search(r'(\d+)', cell_ref)
    if not (m_col and m_row):
        raise ValueError(f"Invalid A1 ref: {cell_ref}")
    col_str = m_col.group(1)
    row_num = int(m_row.group(1))
    col_num = 0
    for ch in col_str:
        col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
    return row_num, col_num


class HistogramCombinerModule:
    def __init__(self, parent, foam_manager):
        self.parent = parent
        self.foam_manager = foam_manager
        self.window = None

        self.input_folder_var = tk.StringVar()
        self.output_folder_var = tk.StringVar()
        self.output_file_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")
        self.foam_var = tk.StringVar()  # local selection (can differ until committed)

    def show(self):
        if self.window is not None and self.window.winfo_exists():
            self.window.lift()
            self.window.focus_force()
            return

        self.window = tk.Toplevel(self.parent)
        self.window.title("Combine SEM Results")
        self.window.geometry("750x600")
        self.window.transient(self.parent)
        self.window.grab_set()
        self.window.resizable(True, True)

        self.window.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() // 2) - (self.window.winfo_width() // 2)
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() // 2) - (self.window.winfo_height() // 2)
        self.window.geometry(f"+{x}+{y}")

        self.setup_gui()
        self.load_default_paths()
        self.window.lift()
        self.window.focus_force()

    def setup_gui(self):
        main_frame = ttk.Frame(self.window, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="Combine SEM Results", font=("Arial", 16, "bold")).pack(pady=(0, 20))

        # Foam type selector (allows switching without cerrar ventana)
        foam_frame = ttk.Frame(main_frame)
        foam_frame.pack(fill=tk.X, pady=(0, 15))
        ttk.Label(foam_frame, text="Foam Type:").pack(side=tk.LEFT)
        # Populate values
        current_paper = self.foam_manager.get_current_paper()
        foam_types = self.foam_manager.get_foam_types_for_paper(current_paper)
        current_ft = self.foam_manager.get_current_foam_type()
        if current_ft not in foam_types and foam_types:
            current_ft = foam_types[0]
        self.foam_var.set(current_ft)
        foam_combo = ttk.Combobox(foam_frame, state="readonly", values=foam_types, textvariable=self.foam_var, width=20)
        foam_combo.pack(side=tk.LEFT, padx=(8, 0))
        foam_combo.bind("<<ComboboxSelected>>", self.on_foam_type_changed)

        desc = ("This tool combines SEM results (histogram_*.xlsx) into a single Excel file.\n\n"
                "Instructions:\n"
                "1. Select the folder containing histogram_*.xlsx files (SEM/Output)\n"
                "2. Choose where to save the combined SEM results\n"
                "3. Click 'Combine'")
        ttk.Label(main_frame, text=desc, wraplength=700, justify=tk.LEFT).pack(pady=(0, 20))

        in_frame = ttk.LabelFrame(main_frame, text="Input Folder (SEM Results)", padding=15)
        in_frame.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(in_frame, text="Select folder containing histogram_*.xlsx files", foreground="gray").pack(anchor=tk.W, pady=(0, 6))
        row = ttk.Frame(in_frame)
        row.pack(fill=tk.X)
        ttk.Entry(row, textvariable=self.input_folder_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row, text="Browse", command=self.browse_input_folder).pack(side=tk.LEFT)

        out_frame = ttk.LabelFrame(main_frame, text="Output (SEM Results)", padding=15)
        out_frame.pack(fill=tk.X, pady=(0, 12))
        # Output folder
        ttk.Label(out_frame, text="Output Folder", foreground="gray").pack(anchor=tk.W)
        row2a = ttk.Frame(out_frame)
        row2a.pack(fill=tk.X, pady=(0, 6))
        ttk.Entry(row2a, textvariable=self.output_folder_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row2a, text="Browse", command=self.browse_output_folder).pack(side=tk.LEFT)
        # Output file (computed)
        ttk.Label(out_frame, text="Output File (auto)", foreground="gray").pack(anchor=tk.W)
        row2b = ttk.Frame(out_frame)
        row2b.pack(fill=tk.X)
        ttk.Entry(row2b, textvariable=self.output_file_var, state='readonly').pack(side=tk.LEFT, fill=tk.X, expand=True)

        btns = ttk.Frame(main_frame)
        btns.pack(pady=(10, 0))
        ttk.Button(btns, text="Combine Histograms", command=self.combine_histograms).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btns, text="Close", command=self.window.destroy).pack(side=tk.LEFT)

        status = ttk.Frame(main_frame)
        status.pack(fill=tk.X, pady=(20, 0))
        ttk.Separator(status, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(0, 8))
        container = ttk.Frame(status)
        container.pack(fill=tk.X)
        ttk.Label(container, text="Status:", font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        ttk.Label(container, textvariable=self.status_var).pack(side=tk.LEFT, padx=(8, 0))

    def load_default_paths(self):
        try:
            # Prefer local combobox selection if present
            current_foam = self.foam_var.get() or self.foam_manager.get_current_foam_type()
            paper = self.foam_manager.get_current_paper()
            if not (current_foam and paper):
                return
            # Prefer suggested SEM paths (paper root + foam/SEM/Output) if not explicitly saved
            sem_paths = self.foam_manager.get_suggested_paths("SEM", foam_type=current_foam, paper=paper)
            base = None
            if sem_paths and 'output_folder' in sem_paths:
                base = sem_paths['output_folder']
            if base and os.path.isdir(base):
                # Default both paths to the same SEM/Output folder
                self.input_folder_var.set(base)
                self.output_folder_var.set(base)
                default_out = os.path.join(base, f'SEM_Results_{current_foam}.xlsx')
                self.output_file_var.set(default_out)
        except Exception as e:
            self.status_var.set(f"Error loading paths: {e}")

    def on_foam_type_changed(self, event=None):
        """When user changes foam type in this dialog."""
        new_ft = self.foam_var.get().strip()
        if not new_ft:
            return
        # Persist selection globally so otros módulos lo vean
        try:
            self.foam_manager.set_current_foam_type(new_ft)
        except Exception:
            pass
        # Recalcular rutas solo si estaban vacías o pertenecen a otra espuma
        prev_out = self.output_file_var.get()
        if (not prev_out) or (f'SEM_Results_' in prev_out and not prev_out.endswith(f'{new_ft}.xlsx')):
            self.load_default_paths()
        else:
            # Solo actualizar nombre archivo si se cambió espuma pero usuario mantiene carpeta custom
            out_dir = self.output_folder_var.get() or os.path.dirname(prev_out)
            if out_dir:
                self.output_file_var.set(os.path.join(out_dir, f'SEM_Results_{new_ft}.xlsx'))
        self.status_var.set(f"Foam changed to {new_ft}")

    def browse_input_folder(self):
        folder = filedialog.askdirectory(title="Select folder containing histogram files",
                                         initialdir=self.input_folder_var.get() or None)
        if folder:
            self.input_folder_var.set(folder)

    def browse_output_folder(self):
        folder = filedialog.askdirectory(title="Select output folder",
                                         initialdir=self.output_folder_var.get() or self.input_folder_var.get() or None)
        if folder:
            self.output_folder_var.set(folder)
            # Recompute output file path based on foam type
            foam = self.foam_var.get() or self.foam_manager.get_current_foam_type()
            self.output_file_var.set(os.path.join(folder, f'SEM_Results_{foam}.xlsx'))

    def find_and_extract_histogram_data(self, label: str, input_folder: str):
        file_path = os.path.join(input_folder, f"histogram_{label}.xlsx")
        if not os.path.exists(file_path):
            return False, None
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            sheet_names = wb.sheetnames
            if not sheet_names:
                return False, None
            expected = f"histogram_{label}"
            expected_norm = canonical_histogram_key(expected)
            label_norm = canonical_label_key(label)
            label_token = primary_label_token(label)
            target = None
            loose_candidate = None
            label_only_candidate = None
            # Prefer the last matching sheet (others may have similar names)
            for sn in sheet_names:
                key = canonical_histogram_key(sn)
                hist_idx = key.rfind("HISTOGRAM")
                suffix = key[hist_idx + len("HISTOGRAM"):] if hist_idx != -1 else key
                sn_upper = str(sn).upper()
                hist_idx_raw = sn_upper.rfind("HISTOGRAM")
                suffix_raw = sn[hist_idx_raw + len("HISTOGRAM"):] if hist_idx_raw != -1 else sn
                if key == expected_norm:
                    target = sn  # keep last exact match
                    continue
                if suffix == label_norm:
                    target = sn
                    continue
                if suffix.startswith(label_norm) and label_norm:
                    remainder = suffix[len(label_norm):]
                    # Avoid matching "-2" style suffixes when looking for base label
                    if remainder.startswith('-') and '-' not in label:
                        remainder_digits = remainder[1:].replace('-', '')
                        if remainder_digits.isdigit():
                            continue
                    if remainder and remainder[0].isdigit() and '-' not in label:
                        # Skip if remainder starts with digits (likely a different sample)
                        continue
                    loose_candidate = sn
                    continue
                suffix_tokens = extract_label_tokens(suffix_raw)
                if label_token and label_token in suffix_tokens:
                    target = sn
                    continue
                sheet_tokens = extract_label_tokens(sn)
                if label_token and label_token in sheet_tokens:
                    loose_candidate = sn
                    continue
                sheet_label = canonical_label_key(sn)
                if not label_only_candidate and sheet_label == label_norm:
                    label_only_candidate = sn
            if not target:
                target = loose_candidate or label_only_candidate
            if not target:
                return False, None
            sh = wb[target]
            def val(a1):
                r, c = excel_cell_to_indices(a1)
                return sh.cell(row=r, column=c).value
            # Extract from last histogram_<label> sheet
            v_mean_o = pd.to_numeric(val('L3'), errors='coerce')
            v_std_o  = pd.to_numeric(val('M3'), errors='coerce')
            v_rsd_o  = pd.to_numeric(val('N3'), errors='coerce')
            v_nv     = pd.to_numeric(val('AG3'), errors='coerce')
            v_std_nv = pd.to_numeric(val('AH3'), errors='coerce')

            out = {}
            if not pd.isna(v_mean_o):
                out['\u00F8 (\u00B5m)'] = float(v_mean_o)
            if not pd.isna(v_std_o):
                out['Desvest \u00F8 (\u00B5m)'] = float(v_std_o)
            if not pd.isna(v_rsd_o):
                out['RSD \u00F8 (%)'] = float(v_rsd_o) * 100.0
            if not pd.isna(v_nv):
                out['N\u1D65 (cells\u00B7cm^3)'] = float(v_nv)
            if not pd.isna(v_std_nv):
                out['Desvest N\u1D65 (cells\u00B7cm^3)'] = float(v_std_nv)
            if not pd.isna(v_std_nv) and not pd.isna(v_nv) and float(v_nv) != 0.0:
                out['RSD N\u1D65 (%)'] = float(v_std_nv) / float(v_nv) * 100.0
            # n SEM images = number of sheets - 2
            out['n SEM images'] = max(0, len(sheet_names) - 2)
            return True, out
        except Exception as e:
            self.status_var.set(f"Error processing {os.path.basename(file_path)}: {e}")
            return False, None
        finally:
            if wb:
                wb.close()

    def combine_histograms(self):
        try:
            input_folder = self.input_folder_var.get()
            output_folder = self.output_folder_var.get() or self.input_folder_var.get()
            foam = self.foam_var.get() or self.foam_manager.get_current_foam_type()
            # Ensure global state reflects current selection before naming output
            if foam and foam != self.foam_manager.get_current_foam_type():
                try:
                    self.foam_manager.set_current_foam_type(foam)
                except Exception:
                    pass
            output_file = os.path.join(output_folder, f'SEM_Results_{foam}.xlsx') if output_folder else self.output_file_var.get()
            if not input_folder or not os.path.isdir(input_folder):
                messagebox.showerror("Error", "Please select a valid input folder")
                return
            if not output_folder or not os.path.isdir(output_folder):
                messagebox.showerror("Error", "Please select a valid output folder")
                return
            files = glob.glob(os.path.join(input_folder, 'histogram_*.xlsx'))
            if not files:
                messagebox.showwarning("Warning", "No histogram files found in the selected folder")
                return
            labels = []
            for fp in files:
                m = re.search(r'histogram_(.+)\.xlsx', os.path.basename(fp), re.IGNORECASE)
                if m:
                    labels.append(m.group(1))
            if not labels:
                messagebox.showerror("Error", "No valid histogram files found")
                return
            rows = []
            for label in labels:
                self.status_var.set(f"Processing sample {label}...")
                self.window.update()
                ok, data = self.find_and_extract_histogram_data(label, input_folder)
                if ok and data:
                    row = {'Label': label}
                    row.update(data)
                    rows.append(row)
            if not rows:
                messagebox.showerror("Error", "No data could be extracted from histogram files")
                return
            df = pd.DataFrame(rows)
            # Ensure columns and order
            cols = ['Label', 'n SEM images', '\u00F8 (\u00B5m)', 'Desvest \u00F8 (\u00B5m)', 'RSD \u00F8 (%)', 'N\u1D65 (cells\u00B7cm^3)', 'Desvest N\u1D65 (cells\u00B7cm^3)', 'RSD N\u1D65 (%)']
            for c in cols:
                if c not in df.columns:
                    df[c] = pd.NA
            df = df[cols]
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Combined_Results', index=False)
            self.status_var.set(f"Successfully combined {len(rows)} samples")
            messagebox.showinfo("Success", f"Histogram data combined successfully!\nProcessed: {len(rows)} samples\nOutput: {output_file}")
        except Exception as e:
            self.status_var.set(str(e))
            messagebox.showerror("Error", str(e))
