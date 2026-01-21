import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import re
import glob
import warnings
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.widgets import SpanSelector
from scipy.stats import linregress
import numpy as np
from .foam_type_manager import FoamTypeManager, FoamTypeSelector

# Suppress warnings for better user experience
warnings.filterwarnings("ignore", category=FutureWarning, module="pandas")

FOAM_CLASS_FLEX = "Flexible"
FOAM_CLASS_RIGID = "Rigid"


class OCModule:
    def __init__(self, root, paper_path=None, foam_type=None):
        self.root = root
        self.root.title("Open-Cell Content Analysis")
        self.root.geometry("900x700")

        # Store context
        self.paper_path = paper_path
        self.foam_manager = FoamTypeManager()

        # Current foam type
        if foam_type and foam_type in self.foam_manager.get_foam_types():
            self.current_foam_type = foam_type
        else:
            self.current_foam_type = self.foam_manager.get_current_foam_type()

        # Foam class selection (Flexible / Rigid)
        self.instrument_options = [FOAM_CLASS_FLEX, FOAM_CLASS_RIGID]
        self.instrument_var = tk.StringVar(value=FOAM_CLASS_FLEX)
        self.current_instrument = self.instrument_var.get()

        # Paths / state
        self.picnometry_folder = None
        self.selected_picnometry_files = []
        self.density_file = None
        self.output_folder = None
        
        # Review data after processing
        self.processed_data = []  # List of dicts with processing results
        self.review_window = None
        self._last_density_df = None

        # Suggest density at paper root
        if self.paper_path:
            maybe_density = os.path.join(self.paper_path, "Density.xlsx")
            if os.path.exists(maybe_density):
                self.density_file = maybe_density

        self.create_widgets()

    def create_widgets(self):
        main = ttk.Frame(self.root, padding="10")
        main.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(3, weight=1)

        ttk.Label(main, text="Open-Cell Content Analysis", font=("Arial", 16, "bold")).grid(row=0, column=0, pady=(0, 16))

        # Foam type selector
        self.foam_selector = FoamTypeSelector(main, self.foam_manager, self.on_foam_type_changed)

        instrument_frame = ttk.LabelFrame(main, text="Foam Type", padding="10")
        instrument_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(4, 10))
        ttk.Label(instrument_frame, text="Select Foam Type:").grid(row=0, column=0, sticky=tk.W)
        instrument_combo = ttk.Combobox(
            instrument_frame,
            textvariable=self.instrument_var,
            values=self.instrument_options,
            state="readonly",
            width=28,
        )
        instrument_combo.grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        instrument_combo.bind("<<ComboboxSelected>>", lambda _e: self.on_instrument_changed())

        setup = ttk.LabelFrame(main, text="Setup Files and Folders", padding="10")
        setup.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(4, 10))
        setup.columnconfigure(1, weight=1)

        # Input folder like PDR
        ttk.Label(setup, text="Input Folder:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.picnometry_folder_var = tk.StringVar()
        ttk.Entry(setup, textvariable=self.picnometry_folder_var, state="readonly", width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5))
        ttk.Button(setup, text="Browse", command=self.browse_picnometry_folder).grid(row=0, column=2)

        # Density file (paper-level)
        ttk.Label(setup, text="Density File:").grid(row=1, column=0, sticky=tk.W, pady=(5, 5))
        self.density_file_var = tk.StringVar()
        ttk.Entry(setup, textvariable=self.density_file_var, state="readonly", width=60).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5))
        ttk.Button(setup, text="Browse", command=self.browse_density_file).grid(row=1, column=2)

        # Output folder
        ttk.Label(setup, text="Output Folder:").grid(row=2, column=0, sticky=tk.W, pady=(5, 5))
        self.output_folder_var = tk.StringVar()
        ttk.Entry(setup, textvariable=self.output_folder_var, state="readonly", width=60).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 5))
        ttk.Button(setup, text="Browse", command=self.browse_output_folder).grid(row=2, column=2)

        # Results file
        ttk.Label(setup, text="OC Results:").grid(row=3, column=0, sticky=tk.W, pady=(5, 0))
        self.results_file_var = tk.StringVar()
        results_row = ttk.Frame(setup)
        results_row.grid(row=3, column=1, columnspan=2, sticky=(tk.W, tk.E))
        results_row.columnconfigure(0, weight=1)
        ttk.Entry(results_row, textvariable=self.results_file_var, state="readonly", width=60).grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(10, 5))
        ttk.Button(results_row, text="Open/Create", command=self.open_or_create_results_file).grid(row=0, column=1)

        # Suggested paths
        self.load_suggested_paths()

        # Scan button
        ttk.Button(setup, text="🔍 Scan Folder", command=self.scan_picnometry_files).grid(row=4, column=0, pady=(10, 5))

        # Files table
        file_frame = ttk.LabelFrame(main, text="Selected Picnometry Files", padding="10")
        file_frame.grid(row=3, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        file_frame.columnconfigure(0, weight=1)
        file_frame.rowconfigure(1, weight=1)

        # Selection buttons
        sel_frame = ttk.Frame(file_frame)
        sel_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        ttk.Button(sel_frame, text="Select All", command=self.select_all_files).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(sel_frame, text="Select None", command=self.select_none_files).grid(row=0, column=1, padx=(0, 5))
        ttk.Label(sel_frame, text="(Ctrl+Click to select/deselect individual files)").grid(row=0, column=2, padx=(10, 0))

        cols = ("filename", "size", "status")
        self.file_tree = ttk.Treeview(file_frame, columns=cols, show="headings", height=8, selectmode="extended")
        self.file_tree.heading("filename", text="Picnometry File Name")
        self.file_tree.heading("size", text="Size (KB)")
        self.file_tree.heading("status", text="Status")
        self.file_tree.column("filename", width=420)
        self.file_tree.column("size", width=100, anchor="center")
        self.file_tree.column("status", width=160, anchor="center")
        self.file_tree.grid(row=1, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        vbar = ttk.Scrollbar(file_frame, orient=tk.VERTICAL, command=self.file_tree.yview)
        vbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.file_tree.configure(yscrollcommand=vbar.set)
        
        # Bind Ctrl+Click for selection (remove old click binding)
        self.file_tree.bind("<Control-Button-1>", self.on_ctrl_click)

        # Buttons
        btns = ttk.Frame(main)
        btns.grid(row=4, column=0, pady=10)
        ttk.Button(btns, text=" Process Selected", command=self.process_files).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(btns, text="📊 Open/Create Results", command=self.open_or_create_results_file).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(btns, text="💾 Save Paths", command=self.save_current_paths).grid(row=0, column=2, padx=(0, 10))
        ttk.Button(btns, text="❌ Close", command=self.root.destroy).grid(row=0, column=3)

        # Status + progress
        self.status_var = tk.StringVar(value="Ready - Select foam type and configure paths")
        ttk.Label(main, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W).grid(row=5, column=0, sticky=(tk.W, tk.E))
        self.progress = ttk.Progressbar(main, mode="determinate")
        self.progress.grid(row=6, column=0, sticky=(tk.W, tk.E), pady=(5, 0))

        # Pre-fill shown values if known
        if self.density_file:
            self.density_file_var.set(self.density_file)

    def load_suggested_paths(self):
        sug = self.foam_manager.get_suggested_paths("OC", self.current_foam_type)
        if not sug:
            return
        # Input (suggested path even if it doesn't exist yet)
        if sug.get("input_folder"):
            if os.path.exists(sug["input_folder"]) or not self.picnometry_folder:
                self.picnometry_folder = sug["input_folder"]
                self.picnometry_folder_var.set(self.picnometry_folder)
        # Output (suggested path even if it doesn't exist yet)
        if sug.get("output_folder"):
            if os.path.exists(sug["output_folder"]) or not self.output_folder:
                self.output_folder = sug["output_folder"]
                self.output_folder_var.set(self.output_folder)
        # Density (suggested path even if it doesn't exist yet)
        if sug.get("density_file"):
            if os.path.exists(sug["density_file"]) or not self.density_file:
                self.density_file = sug["density_file"]
                self.density_file_var.set(self.density_file)
        # Results file path suggestion
        if sug.get("results_file"):
            self.results_file_var.set(sug["results_file"])

    def on_instrument_changed(self):
        self.current_instrument = self.instrument_var.get()
        if hasattr(self, "status_var"):
            self.status_var.set(f"Instrument set to {self.current_instrument}")

    def on_foam_type_changed(self, new_type):
        # Save current paths before switching
        if self.output_folder or self.density_file or self.picnometry_folder:
            self.save_current_paths()
        self.current_foam_type = new_type
        # Clear files table
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        self.selected_picnometry_files = []
        # Reload suggestions
        self.load_suggested_paths()
        self.status_var.set(f"Switched to {new_type} - paths updated")

    def save_current_paths(self):
        data = {}
        if self.picnometry_folder:
            data["input_folder"] = self.picnometry_folder
        if self.output_folder:
            data["output_folder"] = self.output_folder
        if self.density_file:
            data["density_file"] = self.density_file
        if self.results_file_var.get():
            data["results_file"] = self.results_file_var.get()
        if data:
            self.foam_manager.save_module_paths("OC", self.current_foam_type, data)
            self.status_var.set(f"Paths saved for {self.current_foam_type}")

    def on_ctrl_click(self, event):
        """Handle Ctrl+Click for file selection (similar to PDR)"""
        pass  # Let default treeview selection handle this

    def select_all_files(self):
        """Select all files in the tree"""
        for item in self.file_tree.get_children():
            self.file_tree.selection_add(item)

    def select_none_files(self):
        """Deselect all files in the tree"""
        self.file_tree.selection_remove(self.file_tree.selection())

    def get_selected_files(self):
        """Get list of files that are selected in the treeview"""
        selected = []
        selected_items = self.file_tree.selection()
        for item in selected_items:
            values = self.file_tree.item(item, "values")
            if values:
                filename = values[0]  # filename is now in first column
                # Find full path
                for full_path in self.selected_picnometry_files:
                    if os.path.basename(full_path) == filename:
                        selected.append(full_path)
                        break
        return selected

    def browse_picnometry_folder(self):
        # Try suggested as initial dir
        initial = None
        sug = self.foam_manager.get_suggested_paths("OC", self.current_foam_type)
        if sug and sug.get("input_folder"):
            initial = sug["input_folder"]
        folder = filedialog.askdirectory(title="Select Picnometry Folder", initialdir=initial)
        if folder:
            self.picnometry_folder = folder
            self.picnometry_folder_var.set(folder)
            self.scan_picnometry_files()

    def scan_picnometry_files(self):
        self.selected_picnometry_files = []
        folder = self.picnometry_folder or self.picnometry_folder_var.get()
        if not folder:
            messagebox.showwarning("No folder", "Please select an Input Folder first.")
            return
        excel_files = []
        for ext in ("*.xlsx", "*.xls"):
            excel_files.extend(glob.glob(os.path.join(folder, ext)))
        excel_files = sorted(excel_files)
        if not excel_files:
            messagebox.showinfo("Info", "No Excel files found in the selected folder")
            self.update_file_tree([])
            return
        self.selected_picnometry_files = excel_files
        self.update_file_tree(excel_files)
        self.status_var.set(f"Found {len(excel_files)} files in folder")

    def update_file_tree(self, files):
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        for p in files:
            name = os.path.basename(p)
            try:
                size_kb = round(os.path.getsize(p) / 1024, 1)
            except Exception:
                size_kb = "?"
            item = self.file_tree.insert("", "end", values=(name, size_kb, "Ready"))
            # Select all by default
            self.file_tree.selection_add(item)

    def browse_density_file(self):
        initial = None
        sug = self.foam_manager.get_suggested_paths("OC", self.current_foam_type)
        if sug and sug.get("density_file"):
            initial = os.path.dirname(sug["density_file"]) or None
        filename = filedialog.askopenfilename(
            title="Select Density.xlsx file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="Density.xlsx",
            initialdir=initial,
        )
        if filename:
            self.density_file = filename
            self.density_file_var.set(filename)
            self.status_var.set(f"Density file set: {os.path.basename(filename)}")

    def browse_output_folder(self):
        initial = None
        sug = self.foam_manager.get_suggested_paths("OC", self.current_foam_type)
        if sug and sug.get("output_folder"):
            initial = os.path.dirname(sug["output_folder"]) or None
        folder = filedialog.askdirectory(title="Select output folder", initialdir=initial)
        if folder:
            self.output_folder = folder
            self.output_folder_var.set(folder)
            # Suggest results file path if not set
            results = os.path.join(folder, self.get_output_filename())
            self.results_file_var.set(results)
            self.status_var.set(f"Output folder set: {os.path.basename(folder)}")

    def get_output_filename(self):
        """Generate dynamic filename based on foam type"""
        if not self.current_foam_type:
            print(f"Warning: current_foam_type is empty or None: '{self.current_foam_type}'")
            return "OC_Results.xlsx"  # Fallback
        return f"OC_Results_{self.current_foam_type}.xlsx"

    def _load_density_dataframe(self):
        if not self.density_file or not os.path.exists(self.density_file):
            raise ValueError("Density.xlsx not set or not found.")
        try:
            density_df = pd.read_excel(self.density_file, sheet_name=self.current_foam_type, engine="openpyxl")
            density_df["Label"] = density_df["Label"].astype(str)
            density_df.set_index("Label", inplace=True)
        except Exception as e:
            raise ValueError(f"Could not read density sheet for '{self.current_foam_type}': {e}")
        try:
            self._density_pos = self._excel_letter_to_iloc_pos('F', label_letter='B')
            self._rho_r_pos = self._excel_letter_to_iloc_pos('I', label_letter='B')
            ncols_check = density_df.shape[1]
            if not (0 <= self._density_pos < ncols_check and 0 <= self._rho_r_pos < ncols_check):
                raise IndexError(
                    f"Positions out of range (ncols={ncols_check}): F->{self._density_pos}, I->{self._rho_r_pos}"
                )
        except Exception as e:
            raise ValueError(f"Failed to map Excel columns F/I: {e}")
        self._last_density_df = density_df
        return density_df

    def _ensure_density_dataframe(self):
        if self._last_density_df is not None:
            return self._last_density_df
        return self._load_density_dataframe()

    def open_or_create_results_file(self):
        target = self.results_file_var.get()
        if not target:
            sug = self.foam_manager.get_suggested_paths("OC", self.current_foam_type)
            target = sug.get("results_file") if sug else None
            if not target:
                messagebox.showerror("Error", "No results file path configured. Set an Output Folder first.")
                return
            self.results_file_var.set(target)
        
        os.makedirs(os.path.dirname(target), exist_ok=True)
        
        if os.path.exists(target):
            # File exists, open it directly
            try:
                os.startfile(target)
                self.status_var.set(f"Opened {os.path.basename(target)}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open results file: {e}")
        else:
            # File doesn't exist, ask to create
            response = messagebox.askyesno(
                "Create OC Results File", 
                "No OC results file found. Would you like to create a new one?"
            )
            
            if response:
                try:
                    cols = [
                        "Label",
                        "Density (g/cm3)",
                        "m (g)",
                        "Vext (cm3)",
                        "Vpyc unfixed (cm3)",
                        "Vpyc (cm3)",
                        "ρr",
                        "R2",
                        "Vext - Vpyc (cm3)",
                        "1-ρr",
                        "Vext(1-ρr) (cm3)",
                        "%OC",
                        "Comment Analysis",
                    ]
                    pd.DataFrame(columns=cols).to_excel(target, index=False, engine="openpyxl")
                    # Overwrite with canonical headers to avoid duplicates/mojibake
                    try:
                        _canon_cols = [
                            "Label",
                            "\u03C1 foam (g/cm^3)",
                            "m (g)",
                            "Vext (cm3)",
                            "Vpyc unfixed (cm3)",
                            "Vpyc (cm3)",
                            "\u03C1r",
                            "R2",
                            "Vext - Vpyc (cm3)",
                            "1-\u03C1r",
                            "Vext(1-\u03C1r) (cm3)",
                            "%OC",
                            "Comment Analysis",
                        ]
                        pd.DataFrame(columns=_canon_cols).to_excel(target, index=False, engine="openpyxl")
                    except Exception:
                        pass
                    self.status_var.set(f"Created {os.path.basename(target)}")
                    
                    # Open the newly created file
                    os.startfile(target)
                    messagebox.showinfo("Success", f"New OC results file created and opened:\\n{target}")
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create OC results file: {str(e)}")
                    self.status_var.set("Error creating OC results file")

    # ---------- Processing ----------
    def process_files(self):
        selected_files = self.get_selected_files()
        if not selected_files:
            messagebox.showerror("Error", "No files selected for processing. Check files in the list.")
            return
        if not self.output_folder:
            messagebox.showerror("Error", "Output folder not selected.")
            return
        os.makedirs(self.output_folder, exist_ok=True)

        try:
            density_df = self._load_density_dataframe()
        except Exception as e:
            density_df = None
            messagebox.showwarning("Warning", f"Density file not loaded: {e}")

        foam_class = self.instrument_var.get()

        self.progress["maximum"] = len(selected_files)
        self.progress["value"] = 0
        self.processed_data = []
        errors = 0

        for i, path in enumerate(selected_files):
            fname = os.path.basename(path)
            self.status_var.set(f"Processing {fname}…")
            self.root.update()
            ok, data = self.process_single_file(path, density_df, foam_class=foam_class)
            if ok and data:
                # Add extra info for review
                data['filename'] = fname
                data['filepath'] = path
                data['foam_class'] = data.get('foam_class', foam_class)
                self.processed_data.append(data)
                self._set_tree_status(fname, "✅ Processed")
            else:
                errors += 1
                self._set_tree_status(fname, "❌ Error")
            self.progress["value"] = i + 1
            self.root.update()

        if self.processed_data:
            self.save_results(self.processed_data)
            self.status_var.set(f"Processed and saved {len(self.processed_data)} files")
        else:
            self.status_var.set("No files were successfully processed")

        self.progress["value"] = 0

    def _set_tree_status(self, filename, status):
        for item in self.file_tree.get_children():
            vals = self.file_tree.item(item)["values"]
            if vals and len(vals) > 0 and vals[0] == filename:  # filename is now in column 0
                self.file_tree.set(item, "status", status)
                return

    def process_single_file(self, file_path, density_df, foam_class):
        fname = os.path.basename(file_path)
        try:
            label = os.path.splitext(fname)[0]
            for sep in (" ", "-", "_"):
                prefix = f"{self.current_foam_type}{sep}"
                if label.startswith(prefix):
                    label = label[len(prefix):]
                    break

            mass, df = self._parse_picnometry_file(file_path)
            if mass is None or df is None or df.empty:
                raise ValueError("Could not extract mass or data table")

            validated = self._show_validation_window(label, df, foam_class, file_path)
            if not validated:
                return False, None

            vpyc = validated["Vpyc (cm3)"]
            density = pd.NA
            rho_r = pd.NA
            if density_df is not None:
                target_label = str(label)
                if target_label in density_df.index:
                    row = density_df.loc[target_label]
                else:
                    row = None
                if row is not None:
                    try:
                        density = row.iloc[self._density_pos]
                    except Exception:
                        density = row.iloc[0] if len(row) > 0 else pd.NA
                    try:
                        rho_r = row.iloc[self._rho_r_pos]
                    except Exception:
                        rho_r = row.iloc[1] if len(row) > 1 else pd.NA

            result = {
                "Label": label,
                "Density (g/cm3)": density,
                "m (g)": mass,
                "Vpyc unfixed (cm3)": vpyc,
                "Vpyc (cm3)": vpyc,
                "ρr": rho_r,
                "Comment Analysis": validated.get("comment", ""),
                "R2": validated.get("R2"),
                "foam_class": foam_class,
            }
            return True, result
        except Exception as e:
            print(f"ERROR processing {fname}: {e}")
            return False, None

    def _parse_picnometry_file(self, file_path):
        """Extract mass and measurement table from a picnometry Excel/CSV file."""
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        mass = self._extract_mass(ws)
        df = self._extract_table(ws)
        wb.close()
        return mass, df

    def _extract_mass(self, ws):
        """Find 'Sample mass:' row and parse numeric value."""
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            first = row[0]
            if isinstance(first, str) and "sample mass" in first.lower():
                val = row[1] if len(row) > 1 else None
                parsed = self._to_float(val)
                if parsed is not None:
                    return parsed
        # Fallback to B14 if not found
        try:
            fallback = ws["B14"].value
            return self._to_float(fallback)
        except Exception:
            return None

    def _extract_table(self, ws):
        """Read measurement table starting at row 35 until an empty row."""
        records = []
        for row in ws.iter_rows(min_row=35, values_only=True):
            cells = list(row[:4])
            if all((c is None) or (isinstance(c, str) and not c.strip()) for c in cells):
                break
            cycle, p1, p2, vol = cells if len(cells) == 4 else (None, None, None, None)
            records.append({
                "Cycle #": self._to_float(cycle),
                "P1 Pressure (psig)": self._to_float(p1),
                "P2 Pressure (psig)": self._to_float(p2),
                "Volume (cm3)": self._to_float(vol),
            })
        df = pd.DataFrame(records)
        return df.dropna(subset=["P1 Pressure (psig)", "Volume (cm3)"])

    def _to_float(self, value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            txt = value.strip().replace("g", "").replace("cm3", "").replace("cm³", "").replace(" ", "")
            txt = txt.replace(",", ".")
            m = re.search(r"[-+]?\d*\.?\d+", txt)
            if m:
                try:
                    return float(m.group(0))
                except Exception:
                    return None
        return None

    def _show_validation_window(self, label, df, foam_class, file_path=None):
        """Interactive selection and regression/average for pycnometry."""
        result_holder = {}

        top = tk.Toplevel(self.root)
        top.title(f"Validate Pycnometry - {label}")
        top.transient(self.root)
        top.grab_set()

        frame = ttk.Frame(top, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        fig = Figure(figsize=(7, 4), dpi=100)
        ax = fig.add_subplot(111)
        ax.set_xlabel("P1 Pressure (psig)")
        ax.set_ylabel("Pycnometric volume (cm³)")

        p1 = df["P1 Pressure (psig)"].to_numpy(dtype=float)
        vol = df["Volume (cm3)"].to_numpy(dtype=float)
        colors = ["#000000"] * len(p1)
        scatter = ax.scatter(p1, vol, c=colors, s=35, alpha=0.8)
        line_plot, = ax.plot([], [], color="#E69F00", linewidth=2, alpha=0.8)
        ax.grid(True, linestyle="--", alpha=0.3)

        stats_var = tk.StringVar(value="Select points to compute Vpyc")

        def update_selection(mask):
            sel_colors = ["#E69F00" if m else "#000000" for m in mask]
            scatter.set_color(sel_colors)
            selected_p1 = p1[mask]
            selected_vol = vol[mask]
            vpyc = None
            r2 = None
            if len(selected_p1) >= 1:
                if foam_class == FOAM_CLASS_FLEX and len(selected_p1) >= 2:
                    res = linregress(selected_p1, selected_vol)
                    vpyc = res.intercept
                    r2 = res.rvalue ** 2
                    xs = np.linspace(selected_p1.min(), selected_p1.max(), 100)
                    ys = res.slope * xs + res.intercept
                    line_plot.set_data(xs, ys)
                else:
                    vpyc = float(np.nanmean(selected_vol))
                    line_plot.set_data([selected_p1.min(), selected_p1.max()], [vpyc, vpyc])
            fig.canvas.draw_idle()
            if vpyc is None or np.isnan(vpyc):
                stats_var.set("Select at least 2 points for Flexible or 1 for Rigid")
            else:
                r2_txt = f" | R²={r2:.4f}" if r2 is not None else ""
                stats_var.set(f"Vpyc={vpyc:.4f} cm³{r2_txt}")
            result_holder["Vpyc (cm3)"] = vpyc
            result_holder["R2"] = r2
            result_holder["mask"] = mask

        initial_mask = np.ones_like(p1, dtype=bool)
        update_selection(initial_mask)

        def on_span(xmin, xmax):
            lo, hi = sorted([xmin, xmax])
            mask = (p1 >= lo) & (p1 <= hi)
            if not mask.any():
                return
            update_selection(mask)

        SpanSelector(ax, on_span, "horizontal", useblit=True, props=dict(alpha=0.15, facecolor="#E69F00"))

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        NavigationToolbar2Tk(canvas, frame)

        info = ttk.Label(frame, textvariable=stats_var, font=("Arial", 10, "bold"))
        info.pack(pady=(6, 6))

        btns = ttk.Frame(frame)
        btns.pack(pady=(4, 0))

        def confirm():
            vpyc_val = result_holder.get("Vpyc (cm3)")
            if vpyc_val is None or np.isnan(vpyc_val):
                messagebox.showerror("Error", "Please select points to compute Vpyc.")
                return
            selected_mask = result_holder.get("mask")
            n_points = int(selected_mask.sum()) if selected_mask is not None else len(p1)
            result_holder["comment"] = f"Validated ({foam_class}) with {n_points} points"
            png_path = os.path.join(self.output_folder or os.path.dirname(file_path), f"{label}_pycnometry.png")
            os.makedirs(os.path.dirname(png_path), exist_ok=True)
            fig.savefig(png_path, dpi=150, bbox_inches="tight")
            top.destroy()

        def cancel():
            result_holder.clear()
            top.destroy()

        ttk.Button(btns, text="Validate/Confirm", command=confirm).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btns, text="Cancel", command=cancel).pack(side=tk.LEFT)

        top.wait_window()
        return result_holder if result_holder else None

    # ---------- Helpers: Excel column letters to iloc position ----------
    def _excel_col_to_number(self, letters):
        """Convert Excel column letters (e.g., 'A', 'F', 'AA') to 1-based number."""
        if not letters:
            raise ValueError("Empty Excel column letters")
        letters = str(letters).strip().upper()
        num = 0
        for ch in letters:
            if not ('A' <= ch <= 'Z'):
                continue
            num = num * 26 + (ord(ch) - ord('A') + 1)
        if num <= 0:
            raise ValueError(f"Invalid Excel column letters: {letters}")
        return num

    def _excel_letter_to_iloc_pos(self, letters, label_letter='B'):
        """Return 0-based iloc position for Excel column `letters` after set_index('Label').

        Uses the actual Label column letter to adjust the removal offset:
        pos = (E-1) - 1 if E > L else (E-1), where E,L are 1-based Excel indices.
        """
        excel_num = self._excel_col_to_number(letters)
        label_num = self._excel_col_to_number(label_letter) if label_letter else 1
        base = excel_num - 1
        pos = base - (1 if excel_num > label_num else 0)
        if pos < 0:
            raise IndexError(f"Computed negative iloc position for Excel column {letters}: {pos}")
        return pos

    def show_review_window(self):
        """Show a window to review and edit processing results before saving"""
        if self.review_window and self.review_window.winfo_exists():
            self.review_window.destroy()
        
        self.review_window = tk.Toplevel(self.root)
        self.review_window.title("Review Processing Results")
        self.review_window.geometry("1000x600")
        self.review_window.transient(self.root)
        self.review_window.grab_set()
        
        main_frame = ttk.Frame(self.review_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        ttk.Label(main_frame, text="Review and Edit Processing Results", 
                 font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        # Table frame
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create treeview for results
        columns = ("filename", "instrument", "comment", "volume")
        self.review_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        self.review_tree.heading("filename", text="File Name")
        self.review_tree.heading("instrument", text="Instrument")
        self.review_tree.heading("comment", text="Comment Analysis")
        self.review_tree.heading("volume", text="Balls Volume (cm³)")
        
        self.review_tree.column("filename", width=250)
        self.review_tree.column("instrument", width=180, anchor="center")
        self.review_tree.column("comment", width=400)
        self.review_tree.column("volume", width=150, anchor="center")
        
        # Scrollbars
        v_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.review_tree.yview)
        h_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.review_tree.xview)
        self.review_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        self.review_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Populate table
        self.review_data_vars = {}  # Store variables for editing
        for i, data in enumerate(self.processed_data):
            filename = data.get('filename', 'Unknown')
            comment = data.get('raw_comment', '')
            detected_volume = data.get('detected_balls_volume', 0.0)
            instrument = data.get('instrument', INSTRUMENT_PRESSTECH)
            
            item_id = self.review_tree.insert("", "end", values=(
                filename,
                instrument,
                comment[:50] + "..." if len(comment) > 50 else comment,
                f"{detected_volume:.6f}"
            ))
            
            # Store variables for this row
            self.review_data_vars[item_id] = {
                'volume': tk.DoubleVar(value=detected_volume),
                'instrument': tk.StringVar(value=instrument),
                'data_index': i
            }
        
        # Edit frame
        edit_frame = ttk.LabelFrame(main_frame, text="Edit Selected Row", padding="10")
        edit_frame.pack(fill=tk.X, pady=(10, 0))
        
        edit_grid = ttk.Frame(edit_frame)
        edit_grid.pack(fill=tk.X)
        
        ttk.Label(edit_grid, text="Instrument:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.edit_instrument_var = tk.StringVar(value=self.instrument_var.get())
        self.edit_instrument_combo = ttk.Combobox(
            edit_grid,
            textvariable=self.edit_instrument_var,
            values=self.instrument_options,
            state="readonly",
            width=25,
        )
        self.edit_instrument_combo.grid(row=0, column=1, padx=(0, 20), sticky=tk.W)
        self.edit_instrument_combo.bind("<<ComboboxSelected>>", lambda _e: self._update_edit_controls(self.edit_instrument_var.get()))

        ttk.Label(edit_grid, text="Balls volume (cm³):").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(6, 0))
        self.edit_volume_var = tk.DoubleVar(value=0.0)
        volume_entry = ttk.Entry(edit_grid, textvariable=self.edit_volume_var, width=15)
        volume_entry.grid(row=1, column=1, padx=(0, 20), pady=(6, 0))
        self.edit_volume_entry = volume_entry
        self._update_edit_controls(self.edit_instrument_var.get())

        ttk.Button(edit_grid, text="Apply to Selected", 
                  command=self.apply_edit_to_selected).grid(row=1, column=2, padx=(10, 0), pady=(6, 0))
        
        # Bind selection event
        self.review_tree.bind("<<TreeviewSelect>>", self.on_review_selection_change)
        
        # Buttons frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=(10, 0))
        
        ttk.Button(btn_frame, text="Save to Excel", 
                  command=self.save_reviewed_results).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="Cancel", 
                  command=self.review_window.destroy).pack(side=tk.LEFT)

    def on_review_selection_change(self, event):
        """Update edit fields when selection changes"""
        selection = self.review_tree.selection()
        if selection:
            item_id = selection[0]
            if item_id in self.review_data_vars:
                vars_dict = self.review_data_vars[item_id]
                self.edit_volume_var.set(vars_dict['volume'].get())
                instrument = vars_dict['instrument'].get()
                self.edit_instrument_var.set(instrument)
                self._update_edit_controls(instrument)

    def apply_edit_to_selected(self):
        """Apply edit values to selected row"""
        selection = self.review_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a row to edit.")
            return

        item_id = selection[0]
        if item_id in self.review_data_vars:
            vars_dict = self.review_data_vars[item_id]
            self._apply_review_update(item_id, vars_dict)

    def _apply_review_update(self, item_id, vars_dict):
        instrument = self.edit_instrument_var.get()
        vars_dict['instrument'].set(instrument)
        self._update_edit_controls(instrument)

        try:
            density_df = self._ensure_density_dataframe()
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return

        data_index = vars_dict['data_index']
        base_data = self.processed_data[data_index]
        filepath = base_data.get('filepath')
        filename = base_data.get('filename', 'Unknown')
        if not filepath:
            messagebox.showerror("Error", "Missing original file path for reprocessing.")
            return

        ok, new_data = self.process_single_file(filepath, density_df, instrument=instrument)
        if not ok or not new_data:
            messagebox.showerror("Error", f"Failed to reprocess {filename} with {instrument}.")
            return

        new_data['filename'] = filename
        new_data['filepath'] = filepath
        new_data['instrument'] = instrument

        if instrument == INSTRUMENT_CELLMAT:
            manual_volume = self.edit_volume_var.get()
            detected_volume = new_data.get('detected_balls_volume', 0.0)
            if abs(manual_volume - detected_volume) > 1e-9:
                vpyc_original = new_data.get("Vpyc unfixed (cm3)")
                if vpyc_original is not None:
                    new_data["Vpyc (cm3)"] = vpyc_original - manual_volume
                new_data["detected_balls_volume"] = manual_volume
                raw_comment = new_data.get("raw_comment", "")
                new_data["Comment Analysis"] = (
                    f"Original: '{raw_comment}' | Manual: balls volume = {manual_volume:.6f} cm³"
                )
            vars_dict['volume'].set(new_data.get("detected_balls_volume", 0.0))
        else:
            new_data["detected_balls_volume"] = 0.0
            new_data["Vpyc (cm3)"] = new_data.get("Vpyc unfixed (cm3)")
            new_data["Comment Analysis"] = INSTRUMENT_PRESSTECH
            vars_dict['volume'].set(0.0)
            self.edit_volume_var.set(0.0)

        self.processed_data[data_index] = new_data

        comment = new_data.get("Comment Analysis", "")
        comment_display = comment if len(comment) <= 50 else comment[:50] + "..."
        volume_display = f"{vars_dict['volume'].get():.6f}"
        self.review_tree.item(item_id, values=(filename, instrument, comment_display, volume_display))

    def _update_edit_controls(self, instrument):
        if instrument == INSTRUMENT_PRESSTECH:
            self.edit_volume_var.set(0.0)
            self.edit_volume_entry.configure(state="disabled")
        else:
            self.edit_volume_entry.configure(state="normal")

    def save_reviewed_results(self):
        """Save the reviewed results to Excel"""
        # Update processed_data with edited values
        for item_id, vars_dict in self.review_data_vars.items():
            data_index = vars_dict['data_index']
            instrument_choice = vars_dict['instrument'].get()
            data = self.processed_data[data_index]
            data['instrument'] = instrument_choice

            if instrument_choice == INSTRUMENT_PRESSTECH:
                data["detected_balls_volume"] = 0.0
                vpyc_original = data.get("Vpyc unfixed (cm3)")
                if vpyc_original is not None:
                    data["Vpyc (cm3)"] = vpyc_original
                data["Comment Analysis"] = INSTRUMENT_PRESSTECH
                continue

            new_balls_volume = vars_dict['volume'].get()
            vpyc_original = data["Vpyc unfixed (cm3)"]
            vpyc_corrected = vpyc_original - new_balls_volume

            detected_volume = data.get("detected_balls_volume", 0.0)
            changed = abs(new_balls_volume - detected_volume) > 1e-9

            data["Vpyc (cm3)"] = vpyc_corrected
            data["detected_balls_volume"] = new_balls_volume

            if changed:
                raw_comment = data.get("raw_comment", "")
                data["Comment Analysis"] = (
                    f"Original: '{raw_comment}' | Manual: balls volume = {new_balls_volume:.6f} cm³"
                )
        
        # Remove extra fields not needed for Excel
        excel_data = []
        for data in self.processed_data:
            excel_row = {k: v for k, v in data.items() 
                        if k not in ['filename', 'filepath', 'raw_comment', 'detected_balls_volume', 'instrument']}
            excel_data.append(excel_row)
        
        try:
            self.save_results(excel_data)
            self.status_var.set(f"Saved {len(excel_data)} results to Excel")
            self.review_window.destroy()
            # Get the output file name for the success message
            target_file = self.results_file_var.get() or os.path.join(self.output_folder or "", self.get_output_filename())
            messagebox.showinfo("Success", f"Saved {len(excel_data)} results to {os.path.basename(target_file)}")
        except Exception as e:
            messagebox.showerror("Error", f"Error saving results: {e}")

    def save_results(self, results_list):
        # Determine results file path
        output_file = self.results_file_var.get() or os.path.join(self.output_folder or "", self.get_output_filename())
        new_df = pd.DataFrame(results_list)
        if os.path.exists(output_file):
            try:
                existing = pd.read_excel(output_file, engine="openpyxl")
                base_cols = [
                    "Label",
                    "Density (g/cm3)",
                    "m (g)",
                    "Vext (cm3)",
                    "Vpyc unfixed (cm3)",
                    "Vpyc (cm3)",
                    "ρr",
                    "Vext - Vpyc (cm3)",
                    "1-ρr",
                    "Vext(1-ρr) (cm3)",
                    "%OC",
                    "Comment Analysis",
                ]
                # Keep only existing columns that we recognize
                existing_cols = [col for col in base_cols if col in existing.columns]
                if existing_cols:
                    existing = existing[existing_cols]
                else:
                    # minimal fallback
                    existing = existing[["Label", "Density (g/cm3)", "m (g)", "Vpyc (cm3)", "ρr"]]
                # Normalize and clean labels
                existing['Label'] = existing['Label'].astype(str)
                existing = existing.dropna(subset=['Label'])
                new_df = new_df.dropna(subset=['Label'])

                # Incremental append-only: add only samples not already present
                existing_labels = set(existing['Label'].astype(str))
                new_only = new_df[~new_df['Label'].astype(str).isin(existing_labels)]

                final_df = (
                    pd.concat([existing, new_only], ignore_index=True)
                      .sort_values(by='Label')
                      .reset_index(drop=True)
                )
            except Exception:
                final_df = new_df.sort_values(by='Label').reset_index(drop=True)
        else:
            final_df = new_df.sort_values(by='Label').reset_index(drop=True)

        # === NUEVO: Calcular las columnas derivadas en Python (sin fórmulas Excel) ===
        df = final_df.copy()

        # Asegurar tipos numéricos
        num_cols = [
            "Density (g/cm3)", "m (g)", "Vpyc unfixed (cm3)", "Vpyc (cm3)", "ρr"
        ]
        for c in num_cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce')

        # Crear columnas si faltan
        for col in ["Vext (cm3)", "Vext - Vpyc (cm3)", "1-ρr", "Vext(1-ρr) (cm3)", "%OC"]:
            if col not in df.columns:
                df[col] = pd.NA

        # Calcular Vext = m / densidad
        if all(c in df.columns for c in ["m (g)", "Density (g/cm3)"]):
            df["Vext (cm3)"] = df.apply(
                lambda r: (r["m (g)"] / r["Density (g/cm3)"])
                if pd.notna(r["m (g)"]) and pd.notna(r["Density (g/cm3)"]) and r["Density (g/cm3)"] not in (0, 0.0)
                else pd.NA,
                axis=1,
            )

        # Calcular Vext - Vpyc
        if all(c in df.columns for c in ["Vext (cm3)", "Vpyc (cm3)"]):
            df["Vext - Vpyc (cm3)"] = df.apply(
                lambda r: (r["Vext (cm3)"] - r["Vpyc (cm3)"])
                if pd.notna(r["Vext (cm3)"]) and pd.notna(r["Vpyc (cm3)"])
                else pd.NA,
                axis=1,
            )

        # Calcular 1 - ρr
        if "ρr" in df.columns:
            df["1-ρr"] = df["ρr"].apply(lambda v: 1 - v if pd.notna(v) else pd.NA)

        # Calcular Vext(1-ρr)
        if all(c in df.columns for c in ["Vext (cm3)", "1-ρr"]):
            df["Vext(1-ρr) (cm3)"] = df.apply(
                lambda r: (r["Vext (cm3)"] * r["1-ρr"])
                if pd.notna(r["Vext (cm3)"]) and pd.notna(r["1-ρr"])
                else pd.NA,
                axis=1,
            )

        # Calcular %OC = (Vext - Vpyc) / (Vext*(1-ρr)) * 100
        if all(c in df.columns for c in ["Vext - Vpyc (cm3)", "Vext(1-ρr) (cm3)"]):
            def _calc_oc(r):
                num = r["Vext - Vpyc (cm3)"]
                den = r["Vext(1-ρr) (cm3)"]
                if pd.notna(num) and pd.notna(den) and den not in (0, 0.0):
                    return (num / den) * 100
                return pd.NA
            df["%OC"] = df.apply(_calc_oc, axis=1)

        # Debug detallado
        print("DEBUG OC PY: Calculadas columnas derivadas")
        for col in ["Vext (cm3)", "Vext - Vpyc (cm3)", "1-ρr", "Vext(1-ρr) (cm3)", "%OC"]:
            if col in df.columns:
                vals = df[col].head(5).tolist()
                print(f"DEBUG OC PY: {col} -> {vals}")

        order = [
            "Label", "Density (g/cm3)", "m (g)", "Vext (cm3)", "Vpyc unfixed (cm3)", "Vpyc (cm3)",
            "ρr", "Vext - Vpyc (cm3)", "1-ρr", "Vext(1-ρr) (cm3)", "%OC", "R2", "Comment Analysis",
        ]
        existing_order = [c for c in order if c in df.columns]
        df = df[existing_order]

        # Guardar valores numéricos (sin fórmulas) para que pandas los lea directamente
        df.to_excel(output_file, index=False, engine="openpyxl")
        print(f"Saved results (numeric) to {output_file}")


def main():
    root = tk.Tk()
    app = OCModule(root)
    root.mainloop()


if __name__ == "__main__":
    main()

def _clean_presstech_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        cleaned = cleaned.replace("cm³", "").replace("cm3", "").replace("g", "")
        cleaned = cleaned.replace(" ", "")
        cleaned = cleaned.replace(",", ".")
        cleaned = re.sub(r"[^\d\.\-+eE]", "", cleaned)
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _resolve_density_rho(self, density_df, label):
    density_row = density_df.loc[label]
    try:
        density = density_row.iloc[self._density_pos]
    except Exception:
        density = None
    try:
        rho_r = density_row.iloc[self._rho_r_pos]
    except Exception:
        rho_r = None

    if rho_r is None or (isinstance(rho_r, float) and pd.isna(rho_r)):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.density_file, data_only=True, read_only=True)
            if self.current_foam_type in wb.sheetnames:
                ws = wb[self.current_foam_type]
                excel_row = int(density_df.index.get_loc(label)) + 2  # header row is 1
                cell_val = ws[f"I{excel_row}"].value
                if cell_val is not None:
                    rho_r = float(cell_val)
        except Exception:
            pass
    return density, rho_r


