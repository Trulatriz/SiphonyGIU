import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import re
import glob
import shutil
import string  # stray block below is disabled
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from .foam_type_manager import FoamTypeManager, FoamTypeSelector

# Canonical column names used throughout Combine + downstream modules
RHO_FOAM_G = "\u03C1 foam (g/cm^3)"
RHO_FOAM_KG = "\u03C1 foam (kg/m^3)"
DESV_RHO_FOAM_G = "Desvest \u03C1 foam (g/cm^3)"
DESV_RHO_FOAM_KG = "Desvest \u03C1 foam (kg/m^3)"
PDER_RHO_FOAM = "%DER \u03C1 foam (g/cm^3)"
RHO_REL = "\u03C1\u1D63"  # ρ with subscript r
EXPANSION_COL = "X"

BASE_NEW_COLUMN_ORDER = [
    'Polymer', 'Additive', 'Additive %', 'Label',
    'm(g)', 'Water (g)', 'T (\u00B0C)', 'P CO2 (bar)', 'Psat (MPa)', 't (min)',
    'Pi (MPa)', 'Pf (MPa)', 'PDR (MPa/s)',
    'n SEM images', '\u00F8 (\u00B5m)', 'Desvest \u00F8 (\u00B5m)', 'RSD \u00F8 (%)',
    'N\u1D65 (cells\u00B7cm^3)', 'Desvest N\u1D65 (cells\u00B7cm^3)', 'RSD N\u1D65 (%)',
    RHO_FOAM_G, RHO_FOAM_KG, DESV_RHO_FOAM_G, DESV_RHO_FOAM_KG, PDER_RHO_FOAM, RHO_REL, EXPANSION_COL,
    'OC (%)',
    'DSC Tm (\u00B0C)', 'DSC Xc (%)', 'DSC Tg (\u00B0C)'
]

DENSITY_DATA_COLUMNS = [RHO_FOAM_G, RHO_FOAM_KG, DESV_RHO_FOAM_G, DESV_RHO_FOAM_KG, PDER_RHO_FOAM, RHO_REL, EXPANSION_COL]

DOE_HEADER_CANDIDATES = {
    'Polymer': ['Polymer', 'Pol\u00EDmero'],
    'Additive': ['Additive', 'Aditivo'],
    'Additive %': ['Additive %', '% Additive', 'Additive%'],
    'Label': ['Label', 'Sample', 'Muestra'],
    'm(g)': ['m(g)', 'Mass (g)', 'm g'],
    'Water (g)': ['Water (g)', 'Water', 'Agua (g)'],
    'T (\u00B0C)': ['T (\u00B0C)', 'Temperature (\u00B0C)', 'T C'],
    'P CO2 (bar)': ['P CO2 (bar)', 'PCO2 (bar)', 'P CO2'],
    'Psat (MPa)': ['Psat (MPa)', 'Psat'],
    't (min)': ['t (min)', 'Time (min)', 't min'],
}

DOE_FALLBACK_LETTERS = {
    'Label': 'A',
    'Additive': 'B',
    'm(g)': 'C',
    'Water (g)': 'D',
    'T (\u00B0C)': 'E',
    'P CO2 (bar)': 'F',
    't (min)': 'G',
}

def _cm_clean_header(name):
    if name is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(name).lower())

def _cm_find_header_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized = {_cm_clean_header(col): col for col in df.columns}
    for cand in candidates:
        key = _cm_clean_header(cand)
        if key in normalized:
            return normalized[key]
    return None

def _cm_get_doe_series(df: pd.DataFrame, target: str) -> pd.Series:
    col_name = _cm_find_header_column(df, DOE_HEADER_CANDIDATES.get(target, []))
    if col_name:
        return df[col_name]
    letter = DOE_FALLBACK_LETTERS.get(target)
    if letter:
        return _cm_col(df, letter)
    return pd.Series([pd.NA] * len(df))


def _normalize_numeric_series(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series(dtype=float)
    cleaned = series.astype(str).str.replace(r"\s", "", regex=True).str.replace(',', '.', regex=False)
    return pd.to_numeric(cleaned, errors='coerce')


def _ensure_psat_column(df: pd.DataFrame) -> pd.DataFrame:
    """Create Psat (MPa) derived from P CO2 (bar) when available."""
    if df is None or df.empty:
        return df
    psat_empty = "Psat (MPa)" in df.columns and df["Psat (MPa)"].isna().all()
    if ("Psat (MPa)" not in df.columns or psat_empty) and "P CO2 (bar)" in df.columns:
        psat = pd.to_numeric(df["P CO2 (bar)"], errors="coerce") / 10
        df = df.copy()
        df["Psat (MPa)"] = psat
    return df

class CombineModule:
    def __init__(self, root, paper_path=None):
        self.root = root
        self.root.title("Combine Results")
        self.root.geometry("800x600")

        # Store paper path for potential use
        self.paper_path = paper_path

        # Initialize foam type manager
        self.foam_manager = FoamTypeManager()

        # If caller provided an explicit paper_path (from New Paper workflow), register it
        try:
            current_paper = self.foam_manager.get_current_paper()
            if paper_path and os.path.exists(paper_path):
                # If the provided path already ends with the paper name, use it directly
                if os.path.basename(os.path.normpath(paper_path)) == current_paper:
                    self.foam_manager.set_paper_root_path(current_paper, paper_path)
                else:
                    # If a parent contains the paper folder, try to locate it
                    candidate = os.path.join(paper_path, current_paper)
                    if os.path.exists(candidate):
                        self.foam_manager.set_paper_root_path(current_paper, candidate)
        except Exception:
            pass

        # Variables for specific input files instead of generic input folder
        self.doe_file = None
        self.cell_analysis_file = None
        self.density_file = None
        self.pdr_file = None
        self.oc_file = None
        self.dsc_file = None
        self.output_folder = None

        # Get foam types from foam manager instead of fixed polymer list
        self.foam_types = self.foam_manager.get_foam_types()  # Dynamic foam types

        # Column configuration
        self.new_column_order = list(BASE_NEW_COLUMN_ORDER)

        self.create_widgets()

        # Set default paths after widgets are created
        self.load_suggested_paths()
        # Normalize final column order to canonical names (avoid encoding issues)
        self.new_column_order = list(BASE_NEW_COLUMN_ORDER)
        
    def normalize_label(self, s):
        """Return a cleaned label, extracting numeric token when present without adding replicate suffixes."""
        if s is None:
            return ""
        label = str(s).strip()
        label = re.sub(r"\.(xlsx|xls|csv|txt)$", "", label, flags=re.IGNORECASE)
        matches = re.findall(r"(\d{4,}(?:-\d+)?)", label)
        if matches:
            return matches[-1]
        return label

    def create_widgets(self):
        """Create the GUI widgets with tabbed interface for better organization"""
        main = ttk.Frame(self.root, padding="10")
        main.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=1)

        # Title
        ttk.Label(main, text="Combine Results", font=("Arial", 16, "bold")).grid(row=0, column=0, pady=(0, 16))

        # Foam type selector
        self.foam_selector = FoamTypeSelector(main, self.foam_manager, self.on_foam_type_changed)

        # Create Notebook for tabs with increased height
        self.notebook = ttk.Notebook(main, height=200)
        self.notebook.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(4, 10))
        main.rowconfigure(2, weight=1)
        
        # Create tabs
        self.setup_general_tab()
        self.setup_foam_type_tabs()

        # Output folder
        output_frame = ttk.LabelFrame(main, text="Output", padding="10")
        output_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(4, 10))
        output_frame.columnconfigure(1, weight=1)

        ttk.Label(output_frame, text="Output Folder:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.output_folder_var = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.output_folder_var, width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=(0, 5))
        ttk.Button(output_frame, text="Browse", command=self.browse_output_folder).grid(row=0, column=2, pady=(0, 5))

        # Results file path
        ttk.Label(output_frame, text="Results File:").grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        self.results_file_var = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.results_file_var, width=60).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=(5, 0))

        # Process button
        ttk.Button(main, text="Combine All Data", command=self.combine_all_data).grid(row=5, column=0, pady=(10, 5))

        # Progress bar
        self.progress = ttk.Progressbar(main, mode='indeterminate')
        self.progress.grid(row=6, column=0, sticky=(tk.W, tk.E), pady=(5, 5))

        # Status
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(main, textvariable=self.status_var).grid(row=7, column=0, pady=(5, 0))

    def setup_general_tab(self):
        """Setup the General tab with paper-level files"""
        general_frame = ttk.Frame(self.notebook)
        self.notebook.add(general_frame, text="📋 General")
        
        # Configure grid
        general_frame.columnconfigure(1, weight=1)
        
        # Paper-level files
        ttk.Label(general_frame, text="Paper-level files", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=3, pady=(10, 15), sticky=tk.W)
        
        # DoE file
        ttk.Label(general_frame, text="DoE File:").grid(row=1, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
        self.doe_file_var = tk.StringVar()
        ttk.Entry(general_frame, textvariable=self.doe_file_var, width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
        ttk.Button(general_frame, text="Browse", command=self.browse_doe_file).grid(row=1, column=2, pady=(0, 8))
        
        # Density file
        ttk.Label(general_frame, text="Density File:").grid(row=2, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
        self.density_file_var = tk.StringVar()
        ttk.Entry(general_frame, textvariable=self.density_file_var, width=50).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
        ttk.Button(general_frame, text="Browse", command=self.browse_density_file).grid(row=2, column=2, pady=(0, 8))
        
        # Instructions
        instructions = ("These files are shared across all foam types:\n"
                       "• DoE: Design of Experiments data\n"
                       "• Density: Density measurements for all foam types")
        ttk.Label(general_frame, text=instructions, foreground="gray", justify=tk.LEFT).grid(row=3, column=0, columnspan=3, pady=(20, 10), sticky=tk.W)

    def setup_foam_type_tabs(self):
        """Setup tabs for each foam type"""
        current_paper = self.foam_manager.get_current_paper()
        foam_types = self.foam_manager.get_foam_types_for_paper(current_paper)
        
        # Store variables for each foam type
        self.foam_vars = {}
        
        for foam_type in foam_types:
            # Create tab for this foam type
            foam_frame = ttk.Frame(self.notebook)
            self.notebook.add(foam_frame, text=f"{foam_type}")
            
            # Configure grid
            foam_frame.columnconfigure(1, weight=1)
            
            # Initialize variables for this foam type
            self.foam_vars[foam_type] = {
                'sem': tk.StringVar(),
                'pdr': tk.StringVar(),
                'oc': tk.StringVar(),
                'dsc': tk.StringVar()
            }
            
            # Foam-specific files header
            ttk.Label(foam_frame, text=f"Files for {foam_type}", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=3, pady=(10, 15), sticky=tk.W)
            
            # SEM Results File
            ttk.Label(foam_frame, text="SEM Results File:").grid(row=1, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
            ttk.Entry(foam_frame, textvariable=self.foam_vars[foam_type]['sem'], width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
            ttk.Button(foam_frame, text="Browse", command=lambda ft=foam_type: self.browse_sem_file(ft)).grid(row=1, column=2, pady=(0, 8))
            
            # PDR Results File
            ttk.Label(foam_frame, text="PDR Results File:").grid(row=2, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
            ttk.Entry(foam_frame, textvariable=self.foam_vars[foam_type]['pdr'], width=50).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
            ttk.Button(foam_frame, text="Browse", command=lambda ft=foam_type: self.browse_pdr_file(ft)).grid(row=2, column=2, pady=(0, 8))
            
            # OC Results File
            ttk.Label(foam_frame, text="OC Results File:").grid(row=3, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
            ttk.Entry(foam_frame, textvariable=self.foam_vars[foam_type]['oc'], width=50).grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
            ttk.Button(foam_frame, text="Browse", command=lambda ft=foam_type: self.browse_oc_file(ft)).grid(row=3, column=2, pady=(0, 8))
            
            # DSC Results File
            ttk.Label(foam_frame, text="DSC Results File:").grid(row=4, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
            ttk.Entry(foam_frame, textvariable=self.foam_vars[foam_type]['dsc'], width=50).grid(row=4, column=1, sticky=(tk.W, tk.E), padx=(0, 10), pady=(0, 8))
            ttk.Button(foam_frame, text="Browse", command=lambda ft=foam_type: self.browse_dsc_file(ft)).grid(row=4, column=2, pady=(0, 8))
            

    def on_foam_type_changed(self, new_type):
        """Handle foam type selection change"""
        if hasattr(self, 'status_var'):
            self.status_var.set(f"Foam type changed to {new_type}")
        # Reload suggested paths for new foam type
        self.load_suggested_paths()

    def load_suggested_paths(self):
        """Load suggested paths for current foam type and set default file paths in tabs"""
        try:
            current_paper = self.foam_manager.get_current_paper()
            if not current_paper:
                return
                
            base_path = self.get_base_path()
            if not base_path:
                return
                
            # Load paper-level files (General tab)
            doe_path = os.path.join(base_path, "DoE.xlsx")
            if os.path.exists(doe_path):
                self.doe_file_var.set(doe_path)
                
            density_path = os.path.join(base_path, "Density.xlsx")
            if os.path.exists(density_path):
                self.density_file_var.set(density_path)
            
            # Load foam-specific files for each foam type
            foam_types = self.foam_manager.get_foam_types_for_paper(current_paper)
            
            for foam_type in foam_types:
                if foam_type not in getattr(self, 'foam_vars', {}):
                    continue
                    
                foam_path = os.path.join(base_path, foam_type)
                
                # SEM Combined Results
                sem_paths = self.foam_manager.get_module_paths("SEM", foam_type=foam_type, paper=current_paper)
                if sem_paths and 'output_folder' in sem_paths:
                    sem_output_path = sem_paths['output_folder']
                else:
                    sem_output_path = os.path.join(foam_path, "SEM", "Output")
                
                sem_file = os.path.join(sem_output_path, f"SEM_Results_{foam_type}.xlsx")
                self.foam_vars[foam_type]['sem'].set(sem_file)
                
                # PDR Results
                pdr_paths = self.foam_manager.get_module_paths("PDR", foam_type=foam_type, paper=current_paper)
                if pdr_paths and 'output_folder' in pdr_paths:
                    pdr_output_path = pdr_paths['output_folder']
                else:
                    pdr_output_path = os.path.join(foam_path, "PDR", "Output")
                
                pdr_file = os.path.join(pdr_output_path, f"PDR_Results_{foam_type}.xlsx")
                self.foam_vars[foam_type]['pdr'].set(pdr_file)
                
                # OC Results
                oc_paths = self.foam_manager.get_module_paths("OC", foam_type=foam_type, paper=current_paper)
                if oc_paths and 'output_folder' in oc_paths:
                    oc_output_path = oc_paths['output_folder']
                else:
                    oc_output_path = os.path.join(foam_path, "Open-cell content", "Output")
                
                oc_file = os.path.join(oc_output_path, f"OC_Results_{foam_type}.xlsx")
                self.foam_vars[foam_type]['oc'].set(oc_file)
                
                # DSC Results
                dsc_paths = self.foam_manager.get_module_paths("DSC", foam_type=foam_type, paper=current_paper)
                if dsc_paths and 'output_folder' in dsc_paths:
                    dsc_output_path = dsc_paths['output_folder']
                else:
                    dsc_output_path = os.path.join(foam_path, "DSC", "Output")
                
                dsc_file = os.path.join(dsc_output_path, f"DSC_Results_{foam_type}.xlsx")
                self.foam_vars[foam_type]['dsc'].set(dsc_file)

            # Update output paths
            combine_suggested = self.foam_manager.get_suggested_paths("Combine", foam_types[0] if foam_types else None)
            if combine_suggested and "output_folder" in combine_suggested:
                if os.path.exists(combine_suggested["output_folder"]) or not hasattr(self, 'output_folder') or not self.output_folder:
                    self.output_folder_var.set(combine_suggested["output_folder"])
                    self.output_folder = combine_suggested["output_folder"]

            # Update results file path with date
            if hasattr(self, 'results_file_var') and hasattr(self, 'output_folder_var') and self.output_folder_var.get():
                current_date = datetime.now().strftime("%Y%m%d")
                results_filename = f"All_Results_{current_date}.xlsx"
                results_path = os.path.join(self.output_folder_var.get(), results_filename)
                self.results_file_var.set(results_path)

        except Exception as e:
            if hasattr(self, 'status_var'):
                self.status_var.set(f"Error loading paths: {str(e)}")

    def set_default_paths(self):
        """Set default paths when no config is available"""
        try:
            current_paper = self.foam_manager.get_current_paper()
            current_foam_type = self.foam_manager.get_current_foam_type()
            base_path = os.path.join(os.path.expanduser("~/Documents/PressTech"), current_paper)
            
            # Default output folder (Combine specific)
            if not self.output_folder:
                default_output = os.path.join(base_path, "Combine", "Output combine")
                self.output_folder = default_output
                if hasattr(self, 'output_folder_var'):
                    self.output_folder_var.set(default_output)
            
            # Default results file
            if hasattr(self, 'results_file_var') and not self.results_file_var.get():
                default_results = os.path.join(self.output_folder, "All_Results.xlsx")
                self.results_file_var.set(default_results)
        except Exception as e:
            if hasattr(self, 'status_var'):
                self.status_var.set(f"Error loading paths: {str(e)}")

    def get_base_path(self):
        """Get base path from the actual paper configuration"""
        try:
            current_paper = self.foam_manager.get_current_paper()
            if not current_paper:
                return None

            # 1. Preferred: explicit root registered by New Paper
            paper_root = self.foam_manager.get_paper_root_path(current_paper)
            if paper_root and os.path.exists(paper_root):
                return paper_root

            # 2. Derive from any saved module path (legacy support)
            for module in ["OC", "PDR", "DSC", "SEM", "DSC"]:
                try:
                    saved_paths = self.foam_manager.get_module_paths(module, paper=current_paper)
                    if saved_paths and 'input_folder' in saved_paths:
                        parts = os.path.normpath(saved_paths['input_folder']).split(os.sep)
                        if current_paper in parts:
                            idx = parts.index(current_paper)
                            candidate = os.sep.join(parts[:idx+1])
                            if os.path.exists(candidate):
                                # Register for future fast lookup
                                self.foam_manager.set_paper_root_path(current_paper, candidate)
                                return candidate
                except Exception:
                    continue

            # 3. Use provided self.paper_path if plausible
            if self.paper_path and os.path.exists(self.paper_path):
                if os.path.basename(os.path.normpath(self.paper_path)) == current_paper:
                    self.foam_manager.set_paper_root_path(current_paper, self.paper_path)
                    return self.paper_path
                else:
                    nested = os.path.join(self.paper_path, current_paper)
                    if os.path.exists(nested):
                        self.foam_manager.set_paper_root_path(current_paper, nested)
                        return nested

            # 4. Heuristic search upward from cwd for a folder named current_paper (no Documents fallback)
            try:
                cwd_parts = os.path.normpath(os.getcwd()).split(os.sep)
                for i in range(len(cwd_parts), 0, -1):
                    candidate = os.sep.join(cwd_parts[:i] + [current_paper])
                    if os.path.exists(candidate):
                        self.foam_manager.set_paper_root_path(current_paper, candidate)
                        return candidate
            except Exception:
                pass

            # If not found, return None (forces user to open via other modules first)
            return None
        except Exception:
            return None

    def browse_doe_file(self):
        """Browse for DoE file"""
        # Set initial directory to paper level
        initial_dir = None
        base_path = self.get_base_path()
        if base_path:
            initial_dir = base_path
            
        file_path = filedialog.askopenfilename(
            title="Select DoE File",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.doe_file = file_path
            self.doe_file_var.set(file_path)

    def browse_sem_file(self, foam_type):
        """Browse for SEM Combined Results file for specific foam type"""
        initial_dir = None
        base_path = self.get_base_path()
        if base_path:
            initial_dir = os.path.join(base_path, foam_type, "SEM", "Output")
            
        file_path = filedialog.askopenfilename(
            title=f"Select SEM Combined Results File for {foam_type}",
            initialdir=initial_dir,
            filetypes=[
                ("Combined SEM Results", "SEM_Results*.xlsx"),
                ("Excel files", "*.xlsx *.xls"), 
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.foam_vars[foam_type]['sem'].set(file_path)

    def browse_pdr_file(self, foam_type):
        """Browse for PDR Results file for specific foam type"""
        initial_dir = None
        base_path = self.get_base_path()
        if base_path:
            initial_dir = os.path.join(base_path, foam_type, "PDR", "Output")
            
        file_path = filedialog.askopenfilename(
            title=f"Select PDR Results File for {foam_type}",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.foam_vars[foam_type]['pdr'].set(file_path)

    def browse_oc_file(self, foam_type):
        """Browse for OC Results file for specific foam type"""
        initial_dir = None
        base_path = self.get_base_path()
        if base_path:
            initial_dir = os.path.join(base_path, foam_type, "Open-cell content", "Output")
            
        file_path = filedialog.askopenfilename(
            title=f"Select Open Cell Content Results File for {foam_type}",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.foam_vars[foam_type]['oc'].set(file_path)

    def browse_dsc_file(self, foam_type):
        """Browse for DSC Results file for specific foam type"""
        initial_dir = None
        base_path = self.get_base_path()
        if base_path:
            initial_dir = os.path.join(base_path, foam_type, "DSC", "Output")
            
        file_path = filedialog.askopenfilename(
            title=f"Select DSC Results File for {foam_type}",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.foam_vars[foam_type]['dsc'].set(file_path)

    def browse_density_file(self):
        """Browse for Density file"""
        # Set initial directory to paper level
        initial_dir = None
        base_path = self.get_base_path()
        if base_path:
            initial_dir = base_path
            
        file_path = filedialog.askopenfilename(
            title="Select Density File",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.density_file = file_path
            self.density_file_var.set(file_path)



    def browse_output_folder(self):
        """Browse for output folder"""
        # Set initial directory to foam type's Combine folder
        initial_dir = None
        base_path = self.get_base_path()
        current_foam_type = self.foam_manager.get_current_foam_type()
        if base_path and current_foam_type:
            initial_dir = os.path.join(base_path, current_foam_type, "Combine")
            
        folder = filedialog.askdirectory(
            title="Select Output Folder",
            initialdir=initial_dir
        )
        if folder:
            self.output_folder = folder
            self.output_folder_var.set(folder)

    def manage_previous_results(self, output_path):
        """Manage previous results: move current to Previous Results folder"""
        try:
            output_dir = os.path.dirname(output_path)
            previous_dir = os.path.join(output_dir, "Previous results")
            
            # Create Previous results directory if it doesn't exist
            if not os.path.exists(previous_dir):
                os.makedirs(previous_dir)
            
            # Look for existing All_Results files in the main Combine folder
            existing_files = glob.glob(os.path.join(output_dir, "All_Results_*.xlsx"))
            
            # Move existing files to Previous results
            for file_path in existing_files:
                if os.path.basename(file_path) != os.path.basename(output_path):  # Don't move the new file
                    filename = os.path.basename(file_path)
                    dest_path = os.path.join(previous_dir, filename)
                    
                    # If file already exists in Previous results, add timestamp
                    if os.path.exists(dest_path):
                        name, ext = os.path.splitext(filename)
                        timestamp = datetime.now().strftime("%H%M%S")
                        new_filename = f"{name}_{timestamp}{ext}"
                        dest_path = os.path.join(previous_dir, new_filename)
                    
                    # Move the file
                    shutil.move(file_path, dest_path)
                    self.status_var.set(f"Moved {filename} to Previous results")
            
        except Exception as e:
            print(f"Error managing previous results: {e}")

    def copy_to_previous_results(self, output_path):
        """Copy the new result file to Previous Results folder"""
        try:
            output_dir = os.path.dirname(output_path)
            previous_dir = os.path.join(output_dir, "Previous results")
            
            # Create Previous results directory if it doesn't exist
            if not os.path.exists(previous_dir):
                os.makedirs(previous_dir)
            
            # Copy the new file to Previous results
            filename = os.path.basename(output_path)
            dest_path = os.path.join(previous_dir, filename)
            
            if os.path.exists(output_path):
                shutil.copy2(output_path, dest_path)
                
        except Exception as e:
            print(f"Error copying to previous results: {e}")

    def combine_all_data(self):
        """Main combine function using tabbed interface input files"""
        try:
            # Validate that we have foam types and variables set up
            if not hasattr(self, 'foam_vars') or not self.foam_vars:
                messagebox.showerror("Error", "No foam types configured. Please check your setup.")
                return
                
            # Build input sets from the tabbed interface
            input_sets = {}
            current_paper = self.foam_manager.get_current_paper()
            foam_types = self.foam_manager.get_foam_types_for_paper(current_paper)
            
            # Get paper-level files from General tab
            doe_file = self.doe_file_var.get().strip() if self.doe_file_var.get() else None
            density_file = self.density_file_var.get().strip() if self.density_file_var.get() else None
            
            # Check for at least some files
            has_files = False
            if doe_file and os.path.exists(doe_file):
                has_files = True
            if density_file and os.path.exists(density_file):
                has_files = True
            
            # Get foam-specific files from each foam type tab
            for foam_type in foam_types:
                if foam_type not in self.foam_vars:
                    continue
                    
                sem_file = self.foam_vars[foam_type]['sem'].get().strip() if self.foam_vars[foam_type]['sem'].get() else None
                pdr_file = self.foam_vars[foam_type]['pdr'].get().strip() if self.foam_vars[foam_type]['pdr'].get() else None
                oc_file = self.foam_vars[foam_type]['oc'].get().strip() if self.foam_vars[foam_type]['oc'].get() else None
                dsc_file = self.foam_vars[foam_type]['dsc'].get().strip() if self.foam_vars[foam_type]['dsc'].get() else None
                
                # Check if any of the foam-specific files exist
                foam_has_files = False
                for file_path in [sem_file, pdr_file, oc_file, dsc_file]:
                    if file_path and os.path.exists(file_path):
                        foam_has_files = True
                        has_files = True
                        break
                
                # Only add to input_sets if this foam type has at least one file
                if foam_has_files or doe_file or density_file:
                    input_sets[foam_type] = {
                        'doe': doe_file,
                        'density': density_file, 
                        'sem': sem_file,
                        'pdr': pdr_file,
                        'oc': oc_file,
                        'dsc': dsc_file,
                    }

            if not has_files:
                messagebox.showerror("Error", "Please select at least one input file in the tabs")
                return
                
            if not self.output_folder_var.get():
                messagebox.showerror("Error", "Please select an output folder")  
                return
            
            self.status_var.set("Starting combine process...")
            self.progress.start()
            
            # Get output file path with current date
            current_date = datetime.now().strftime("%Y%m%d")
            results_filename = f"All_Results_{current_date}.xlsx"
            output_path = os.path.join(self.output_folder_var.get(), results_filename)
            
            # Manage previous results before creating new one
            self.manage_previous_results(output_path)
            
            # Create combined Excel file
            self.create_combined_excel(output_path, input_sets)
            
            # Copy the new file to Previous results
            self.copy_to_previous_results(output_path)
            
            self.status_var.set("Combine completed successfully!")
            messagebox.showinfo("Success", f"Data combined successfully!\nFile saved: {results_filename}\nFoam types processed: {len(input_sets)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Combine failed: {str(e)}")
            self.status_var.set("Combine failed")
        finally:
            self.progress.stop()

    def extract_sample_labels(self, file_path, sheet_name=None):
        """Extract sample labels from an Excel file using original Combine.py logic"""
        try:
            if not file_path or not os.path.exists(file_path):
                return []
            
            # Open workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            labels = []
            
            # If specific sheet name provided, use it
            if sheet_name and sheet_name in workbook.sheetnames:
                sheets_to_process = [sheet_name]
            else:
                # Look for foam type sheets (similar to original polymer sheets)
                foam_types = self.foam_manager.get_foam_types()
                sheets_to_process = [s for s in workbook.sheetnames if s in foam_types]
                
                # If no foam type sheets, try common sheet names
                if not sheets_to_process:
                    common_sheets = ['Sheet1', 'Data', 'Results', 'Registros']
                    sheets_to_process = [s for s in workbook.sheetnames if s in common_sheets]
                    if not sheets_to_process and workbook.sheetnames:
                        sheets_to_process = [workbook.sheetnames[0]]  # Use first sheet
            
            # Process each sheet
            for sheet_name in sheets_to_process:
                try:
                    sheet = workbook[sheet_name]
                    
                    # Determine label column based on file type (following original Combine.py logic)
                    label_col_index = 1  # Default to column A
                    
                    filename_upper = os.path.basename(file_path).upper()
                    if 'DOE' in filename_upper:
                        label_col_index = string.ascii_uppercase.index('A') + 1  # Column A
                    elif 'DENSITY' in filename_upper:
                        label_col_index = string.ascii_uppercase.index('B') + 1  # Column B
                    elif 'PDR' in filename_upper or 'REGISTROS' in filename_upper:
                        label_col_index = string.ascii_uppercase.index('A') + 1  # Column A
                    elif 'DSC' in filename_upper:
                        # For DSC, try to find Label column in header
                        for col in range(1, min(10, sheet.max_column + 1)):
                            header = sheet.cell(row=1, column=col).value
                            if header and str(header).lower().strip() in ['label', 'sample', 'muestra', 'etiqueta']:
                                label_col_index = col
                                break
                    elif 'OC' in filename_upper or 'OPEN' in filename_upper:
                        # For OC, try to find Label column in header  
                        for col in range(1, min(10, sheet.max_column + 1)):
                            header = sheet.cell(row=1, column=col).value
                            if header and str(header).lower().strip() in ['label', 'sample', 'muestra', 'etiqueta']:
                                label_col_index = col
                                break
                    elif 'HISTOGRAM' in filename_upper or 'SEM' in filename_upper:
                        # For SEM/histogram files, try to find Label column in header
                        for col in range(1, min(10, sheet.max_column + 1)):
                            header = sheet.cell(row=1, column=col).value
                            if header and str(header).lower().strip() in ['label', 'sample', 'muestra', 'etiqueta']:
                                label_col_index = col
                                break
                    
                    print(f"Using column {string.ascii_uppercase[label_col_index-1]} for labels in {sheet_name}")  # Debug info
                    
                    # Extract labels from column
                    for row_num in range(2, sheet.max_row + 1):  # Skip header row
                        label_raw = sheet.cell(row=row_num, column=label_col_index).value
                        if label_raw is None or str(label_raw).strip() == "":
                            continue
                        label = self.normalize_label(label_raw)
                        if label and label not in labels:
                            labels.append(label)
                            
                except Exception as e:
                    print(f"Error processing sheet {sheet_name}: {e}")
                    continue
            
            workbook.close()
            return labels
            
        except Exception as e:
            print(f"Error extracting labels from {file_path}: {e}")
            return []

    def find_common_labels(self, input_files):
        """Find labels that are common across multiple input files"""
        try:
            all_labels = {}
            file_labels = {}
            
            # Extract labels from each file
            for file_path in input_files:
                if file_path and os.path.exists(file_path):
                    filename = os.path.basename(file_path)
                    labels = self.extract_sample_labels(file_path)
                    
                    if labels:
                        file_labels[filename] = labels
                        print(f"Found {len(labels)} labels in {filename}: {labels[:5]}...")  # Debug info
                        
                        for label in labels:
                            if label not in all_labels:
                                all_labels[label] = []
                            all_labels[label].append(filename)
                    else:
                        print(f"No labels found in {filename}")
            
            # Find labels that appear in at least 2 files
            common_labels = []
            for label, files in all_labels.items():
                if len(files) >= 2:  # Label appears in at least 2 files
                    common_labels.append(label)
            
            print(f"Common labels found: {common_labels}")  # Debug info
            self.status_var.set(f"Found {len(common_labels)} common sample labels")
            return sorted(common_labels), file_labels
            
        except Exception as e:
            print(f"Error finding common labels: {e}")
            return [], {}

    def create_combined_excel(self, output_path, input_sets):
        """Create Excel file with All_Results sheet and individual foam type sheets.
        input_sets: dict[foam_type] -> {'doe','sem','density','pdr','oc','dsc'} file paths (may be None)
        """
        try:
            any_labels_total = 0
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Create main sheet with all results (combining all foam types)
                all_data = pd.DataFrame()
                
                # For each foam type with inputs, read data and combine
                for foam_type, files_map in input_sets.items():
                    foam_data = self.merge_for_foam(foam_type, files_map)
                    if not foam_data.empty:
                        any_labels_total += len(foam_data)
                    if not foam_data.empty:
                        # Add foam type column
                        foam_data['Polymer'] = foam_type
                        # Append to all_data
                        if all_data.empty:
                            all_data = foam_data.copy()
                        else:
                            all_data = pd.concat([all_data, foam_data], ignore_index=True)
                        
                        # Create individual sheet for this foam type
                        sheet_name = foam_type.replace('/', '_')[:31]  # Excel sheet name limit
                        foam_data.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Write the combined data to All_Results sheet
                # Ensure all columns exist and order them
                if RHO_FOAM_G in all_data.columns:
                    all_data[RHO_FOAM_KG] = pd.to_numeric(all_data[RHO_FOAM_G], errors='coerce') * 1000
                if DESV_RHO_FOAM_G in all_data.columns:
                    all_data[DESV_RHO_FOAM_KG] = pd.to_numeric(all_data[DESV_RHO_FOAM_G], errors='coerce') * 1000
                all_data = _ensure_psat_column(all_data)
                for col in self.new_column_order:
                    if col not in all_data.columns:
                        all_data[col] = pd.NA
                all_data = all_data[self.new_column_order]
                all_data.to_excel(writer, sheet_name='All_Results', index=False)
            
            self.status_var.set(f"Excel file created. Total samples: {any_labels_total}")
            
        except Exception as e:
            raise Exception(f"Error creating Excel file: {e}")

    def process_foam_type_data(self, foam_type, input_files, common_labels):
        """Process data for a specific foam type"""
        try:
            # This is a simplified version - would need more sophisticated data merging
            # For now, return a basic DataFrame with the common labels
            data = {
                'Label': common_labels,
                'Polymer': [foam_type] * len(common_labels)
            }
            
            # Here you would add logic to:
            # 1. Read data from each input file for this foam type
            # 2. Match by sample labels
            # 3. Combine the data into a single row per sample
            
            return pd.DataFrame(data)
            
        except Exception as e:
            print(f"Error processing data for {foam_type}: {e}")
            return pd.DataFrame()

    # ----- Reading helpers -----

    def _read_doe_pos(self, path, foam):
        columns = ['Polymer', 'Additive', 'Additive %', 'Label', 'm(g)', 'Water (g)', 'T (\u00B0C)', 'P CO2 (bar)', 'Psat (MPa)', 't (min)']
        if not path or not os.path.exists(path):
            return pd.DataFrame(columns=columns)
        try:
            df = pd.read_excel(path, sheet_name=foam, engine='openpyxl')
            out = pd.DataFrame({
                'Label': _cm_col(df, 'A').map(self.normalize_label),
                'Additive': _cm_col(df, 'B'),
                'Additive %': pd.to_numeric(_cm_col(df, 'C'), errors='coerce'),
                'm(g)': _cm_col(df, 'D'),
                'Water (g)': _cm_col(df, 'E'),
                'T (\u00B0C)': _cm_col(df, 'F'),
                'P CO2 (bar)': _cm_col(df, 'G'),
                't (min)': _cm_col(df, 'H'),
            })
            out['Polymer'] = foam
            out['Label'] = out['Label'].fillna("").astype(str).str.strip()
            out = out[out['Label'] != ""]
            out = out[~out['Label'].duplicated(keep='first')]
            out['Additive'] = out['Additive'].fillna("")
            out = _ensure_psat_column(out)
            for col in columns:
                if col not in out.columns:
                    out[col] = pd.NA
            return out[columns]
        except Exception:
            return pd.DataFrame(columns=columns)

    def _read_density_pos(self, path, foam):
        fallback = ['Label'] + DENSITY_DATA_COLUMNS
        if not path or not os.path.exists(path):
            return pd.DataFrame(columns=fallback)
        try:
            df = pd.read_excel(path, sheet_name=foam, engine='openpyxl')
            out = pd.DataFrame({
                'Label': _cm_col(df, 'B').map(self.normalize_label),
                RHO_FOAM_G: _cm_col(df, 'F'),
                DESV_RHO_FOAM_G: _cm_col(df, 'G'),
                PDER_RHO_FOAM: _cm_col(df, 'H'),
                RHO_REL: _cm_col(df, 'I'),
                EXPANSION_COL: _cm_col(df, 'J'),
            })
            out[RHO_FOAM_KG] = _normalize_numeric_series(out[RHO_FOAM_G]) * 1000
            out[DESV_RHO_FOAM_KG] = _normalize_numeric_series(out[DESV_RHO_FOAM_G]) * 1000
            return out[['Label'] + DENSITY_DATA_COLUMNS].dropna(subset=['Label'])
        except Exception:
            return pd.DataFrame(columns=fallback)

    def _read_pdr_pos(self, path):
        if not path or not os.path.exists(path):
            return pd.DataFrame(columns=['Label','Pi (MPa)','Pf (MPa)','PDR (MPa/s)'])
        try:
            with pd.ExcelFile(path) as xls:
                sheet = 'Registros' if 'Registros' in xls.sheet_names else xls.sheet_names[0]
                df = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl')
            out = pd.DataFrame({
                'Label': _cm_col(df, 'A').map(self.normalize_label),
                'Pi (MPa)': _cm_col(df, 'B'),
                'Pf (MPa)': _cm_col(df, 'C'),
                'PDR (MPa/s)': _cm_col(df, 'D'),
            })
            return out.dropna(subset=['Label'])
        except Exception:
            return pd.DataFrame(columns=['Label','Pi (MPa)','Pf (MPa)','PDR (MPa/s)'])

    def _read_oc_pos(self, path):
        if not path or not os.path.exists(path):
            print(f"DEBUG OC: No path or file doesn't exist: {path}")
            return pd.DataFrame(columns=['Label','OC (%)'])
        # Estrategia:
        # 1. Intentar leer valores ya calculados (data_only=True) con openpyxl (si el archivo tiene valores cacheados).
        # 2. Si vienen todos None/NaN, fallback a lectura positional previa (por si hay valores escritos directamente).
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = []
            for c in ws[1]:
                v = c.value
                if isinstance(v, str):
                    v = v.strip()
                headers.append(v)
            header_map = {str(h).strip().lower(): idx+1 for idx, h in enumerate(headers) if h is not None}
            # Posibles nombres para %OC
            oc_candidates = ['%oc', 'oc (%)', 'oc%', 'oc %', 'oc(%)']
            oc_col_idx = None
            for cand in oc_candidates:
                if cand in header_map:
                    oc_col_idx = header_map[cand]
                    break
            # Fallback posicional (K = 11) si no se detectó
            if oc_col_idx is None and len(headers) >= 11:
                oc_col_idx = 11
            label_idx = header_map.get('label', 1)
            labels = []
            oc_vals = []
            # Recorrer filas
            for row in ws.iter_rows(min_row=2, values_only=True):
                label_val = row[label_idx-1] if label_idx-1 < len(row) else None
                if label_val in (None, "", "nan"):
                    continue
                oc_raw = None
                if oc_col_idx is not None and oc_col_idx-1 < len(row):
                    oc_raw = row[oc_col_idx-1]
                # Normalizar etiqueta
                label_norm = self.normalize_label(str(label_val)) if label_val is not None else None
                # Convertir valor OC
                oc_num = None
                if isinstance(oc_raw, (int, float)):
                    oc_num = float(oc_raw)
                elif isinstance(oc_raw, str):
                    txt = oc_raw.strip().replace('%', '').replace(',', '.')
                    try:
                        oc_num = float(txt)
                    except Exception:
                        oc_num = None
                labels.append(label_norm)
                oc_vals.append(oc_num)
            wb.close()
            df_out = pd.DataFrame({'Label': labels, 'OC (%)': oc_vals})
            print(f"DEBUG OC WB: Leídos {len(df_out)} registros desde cache (data_only). Primeros: {df_out.head().to_dict(orient='records')}")
            # Si todos NaN/None -> fallback
            if not df_out.empty and df_out['OC (%)'].notna().any():
                return df_out.dropna(subset=['Label'])
            print("DEBUG OC WB: Todos los valores %OC son None/NaN tras data_only, fallback a pandas posicional")
        except Exception as e:
            print(f"DEBUG OC WB ERROR: {e}. Fallback a pandas posicional")

        # Fallback original (posicional) - puede devolver NaN si son fórmulas sin cache.
        try:
            with pd.ExcelFile(path, engine='openpyxl') as xls:
                df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], engine='openpyxl')
            print(f"DEBUG OC Fallback: shape={df.shape} cols={list(df.columns)}")
            out = pd.DataFrame({
                'Label': _cm_col(df, 'A').map(self.normalize_label),
                'OC (%)': pd.to_numeric(_cm_col(df, 'K'), errors='coerce'),
            })
            print(f"DEBUG OC Fallback: primeras filas {out.head().to_dict(orient='records')}")
            return out.dropna(subset=['Label'])
        except Exception as e2:
            print(f"DEBUG OC Fallback ERROR: {e2}")
            return pd.DataFrame(columns=['Label','OC (%)'])

    def _read_dsc_pos(self, path):
        if not path or not os.path.exists(path):
            return pd.DataFrame(columns=['Label','DSC Tm (\u00B0C)','DSC Xc (%)','DSC Tg (\u00B0C)'])
        try:
            with pd.ExcelFile(path) as xls:
                names = {n.lower(): n for n in xls.sheet_names}
                sheet = names.get('dsc_results') or names.get('dsc_tg_results') or xls.sheet_names[0]
                df = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl')
            if sheet.lower() == 'dsc_results':
                out = pd.DataFrame({
                    'Label': _cm_col(df, 'A').map(self.normalize_label),
                    'DSC Xc (%)': _cm_col(df, 'C'),
                    'DSC Tm (\u00B0C)': _cm_col(df, 'F'),
                })
            elif sheet.lower() == 'dsc_tg_results':
                out = pd.DataFrame({
                    'Label': _cm_col(df, 'A').map(self.normalize_label),
                    'DSC Tg (\u00B0C)': _cm_col(df, 'C'),
                })
            else:
                out = pd.DataFrame({'Label': _cm_col(df, 'A').map(self.normalize_label)})
            return out.dropna(subset=['Label'])
        except Exception:
            return pd.DataFrame(columns=['Label','DSC Tm (\u00B0C)','DSC Xc (%)','DSC Tg (\u00B0C)'])

    def _read_sem_pos(self, path):
        if not path or not os.path.exists(path):
            return pd.DataFrame(columns=['Label','n SEM images','\u00F8 (\u00B5m)','Desvest \u00F8 (\u00B5m)','RSD \u00F8 (%)','N\u1D65 (cells\u00B7cm^3)','Desvest N\u1D65 (cells\u00B7cm^3)','RSD N\u1D65 (%)'])
        try:
            with pd.ExcelFile(path) as xls:
                df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], engine='openpyxl')
            out = pd.DataFrame({
                'Label': _cm_col(df, 'A').map(self.normalize_label),
                'n SEM images': _cm_col(df, 'B'),
                '\u00F8 (\u00B5m)': _cm_col(df, 'C'),
                'Desvest \u00F8 (\u00B5m)': _cm_col(df, 'D'),
                'RSD \u00F8 (%)': _cm_col(df, 'E'),
                'N\u1D65 (cells\u00B7cm^3)': _cm_col(df, 'F'),
                'Desvest N\u1D65 (cells\u00B7cm^3)': _cm_col(df, 'G'),
                'RSD N\u1D65 (%)': _cm_col(df, 'H'),
            })
            return out.dropna(subset=['Label'])
        except Exception:
            return pd.DataFrame(columns=['Label','n SEM images','\u00F8 (\u00B5m)','Desvest \u00F8 (\u00B5m)','RSD \u00F8 (%)','N\u1D65 (cells\u00B7cm^3)','Desvest N\u1D65 (cells\u00B7cm^3)','RSD N\u1D65 (%)'])

    def merge_for_foam(self, foam, files_map):
        print(f"DEBUG MERGE: Starting merge for foam {foam}")
        print(f"DEBUG MERGE: Files map = {files_map}")
        doe = self._read_doe_pos(files_map.get('doe'), foam)
        density = self._read_density_pos(files_map.get('density'), foam)
        pdr = self._read_pdr_pos(files_map.get('pdr'))
        oc = self._read_oc_pos(files_map.get('oc'))
        print(f"DEBUG MERGE: OC data shape: {oc.shape}")
        print(f"DEBUG MERGE: OC data columns: {oc.columns.tolist()}")
        print(f"DEBUG MERGE: OC data:\\n{oc}")
        dsc = self._read_dsc_pos(files_map.get('dsc'))
        sem = self._read_sem_pos(files_map.get('sem'))

        label_sets = []
        for df in [doe, density, pdr, oc, dsc, sem]:
            if not df.empty and 'Label' in df.columns:
                label_sets.append(set(df['Label'].map(self.normalize_label)))
        all_labels = set().union(*label_sets) if label_sets else set()
        if not all_labels:
            return pd.DataFrame()

        def idx(df):
            if not df.empty and 'Label' in df.columns:
                tmp = df.copy()
                tmp['Label'] = tmp['Label'].map(self.normalize_label)
                return tmp.set_index('Label')
            return df
        i_doe, i_den, i_pdr, i_oc, i_dsc, i_sem = map(idx, [doe, density, pdr, oc, dsc, sem])

        rows = []
        for lbl in sorted(all_labels):
            row = {'Label': lbl, 'Polymer': foam}
            if not doe.empty and lbl in i_doe.index:
                sel = i_doe.loc[lbl]
                if isinstance(sel, pd.DataFrame):
                    sel = sel.iloc[0]
                row.update(sel[['Additive', 'Additive %', 'm(g)', 'Water (g)', 'T (\u00B0C)', 'P CO2 (bar)', 't (min)']].to_dict())
            if not pdr.empty and lbl in i_pdr.index:
                row.update(i_pdr.loc[lbl][['Pi (MPa)', 'Pf (MPa)', 'PDR (MPa/s)']].to_dict())
            if not sem.empty and lbl in i_sem.index:
                sem_cols = [
                    'n SEM images',
                    '\u00F8 (\u00B5m)',
                    'Desvest \u00F8 (\u00B5m)',
                    'RSD \u00F8 (%)',
                    'N\u1D65 (cells\u00B7cm^3)',
                    'Desvest N\u1D65 (cells\u00B7cm^3)',
                    'RSD N\u1D65 (%)'
                ]
                for c in sem_cols:
                    if c in i_sem.columns:
                        row[c] = i_sem.loc[lbl][c]
            if not density.empty and lbl in i_den.index:
                for c in DENSITY_DATA_COLUMNS:
                    if c in i_den.columns:
                        row[c] = i_den.loc[lbl][c]
            if not oc.empty and lbl in i_oc.index:
                print(f"DEBUG MERGE OC: Adding OC data for label {lbl}")
                oc_data = i_oc.loc[lbl][['OC (%)']].to_dict()
                print(f"DEBUG MERGE OC: OC data = {oc_data}")
                row.update(oc_data)
            if not dsc.empty and lbl in i_dsc.index:
                for c in ['DSC Tm (\u00B0C)', 'DSC Xc (%)', 'DSC Tg (\u00B0C)']:
                    if c in i_dsc.columns:
                        row[c] = i_dsc.loc[lbl][c]
            rows.append(row)

        df = pd.DataFrame(rows)
        df = _ensure_psat_column(df)
        for col in self.new_column_order:
            if col not in df.columns:
                df[col] = pd.NA
        return df[self.new_column_order]

    def auto_detect_polymer_folders(self, base_path):
        """Auto-detect foam type folders from a base results directory"""
        foam_folders = {}
        if os.path.exists(base_path):
            for item in os.listdir(base_path):
                item_path = os.path.join(base_path, item)
                if os.path.isdir(item_path) and item in self.foam_types:  # Use dynamic foam types
                    foam_folders[item] = item_path
        return foam_folders

# ===== Position-based helpers (canonical columns) =====
def _cm_excel_col_idx(letters: str) -> int:
    letters = str(letters).strip().upper()
    num = 0
    for ch in letters:
        if 'A' <= ch <= 'Z':
            num = num * 26 + (ord(ch) - ord('A') + 1)
    if num <= 0:
        return 0
    return num - 1

def _cm_col(df: pd.DataFrame, letters: str):
    idx = _cm_excel_col_idx(letters)
    if idx >= df.shape[1]:
        return pd.Series([pd.NA] * len(df))
    return df.iloc[:, idx]

